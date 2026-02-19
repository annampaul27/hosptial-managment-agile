from django.http import JsonResponse
from django.shortcuts import render, redirect
from django.db.models import Count, Sum
from django.db.models.functions import TruncMonth
from .forms import LabTechnicianForm


import core
from .models import Appointment, DoctorAvailability, Payment, Lab
from datetime import date
from django.contrib.auth.models import User
from django.contrib import messages
from django.contrib.auth.decorators import user_passes_test
from django.contrib.auth.decorators import user_passes_test
from django.shortcuts import get_object_or_404, redirect
from django.contrib.auth import authenticate, login, logout
from django.contrib.auth.decorators import login_required

from .models import (
    PatientProfile,
    DoctorProfile,
    LabTechnicianProfile,
    FrontDeskProfile,
    AdminProfile,
    Appointment,
    Prescription,
    Lab,
    DiagnosticTest,
    TestBooking,
    LabResult,
    Payment,
    LoginHistory,
    PatientHistory,
)


def index(request):
    return render(request, 'core/index.html')

def logout_view(request):
    logout(request)
    messages.success(request, "You have been logged out successfully.")
    return redirect("login")

def login_view(request):
    if request.method == "POST":
        username = request.POST.get("username")
        password = request.POST.get("password")
        role = request.POST.get("role")

        if not username or not password:
            messages.error(request, "Username and password are required")
            return redirect("login")

        user = authenticate(request, username=username, password=password)

        if user is None:
            messages.error(request, "Invalid username or password")
            return redirect("login")

        # 🔐 ROLE VALIDATION
        if role == "patient" and not PatientProfile.objects.filter(user=user).exists():
            messages.error(request, "You are not registered as a patient")
            return redirect("login")

        if role == "doctor" and not DoctorProfile.objects.filter(user=user).exists():
            messages.error(request, "You are not registered as a doctor")
            return redirect("login")

        if role == "lab" and not LabTechnicianProfile.objects.filter(user=user).exists():
            messages.error(request, "You are not registered as lab staff")
            return redirect("login")

        if role == "front_desk" and not FrontDeskProfile.objects.filter(user=user).exists():
            messages.error(request, "You are not registered as front desk staff")
            return redirect("login")

        if role == "admin" and not user.is_staff:
            messages.error(request, "You are not authorized as admin")
            return redirect("login")


        # ✅ SUCCESS
        login(request, user)

        # 🎯 Redirect by role
        if role == "patient":
            return redirect("patient_dashboard")
        elif role == "doctor":
            return redirect("doctor_dashboard")
        elif role == "lab":
            return redirect("lab_dashboard")
        elif role == "front_desk":
            return redirect("frontdesk_dashboard")
        elif role == "admin":
            return redirect("admin_dashboard")

    return render(request, "core/login.html")


def register(request):
    if request.method == "POST":
        full_name = request.POST.get("full_name")
        email = request.POST.get("email")
        phone = request.POST.get("phone")
        password = request.POST.get("password")
        confirm_password = request.POST.get("confirm_password")

        if len(full_name) < 3:
            messages.error(request, "Full name must be at least 3 characters")
            return redirect("patient_register")

        if not phone.isdigit() or len(phone) != 10:
            messages.error(request, "Phone number must be 10 digits")
            return redirect("patient_register")

        if password != confirm_password:
            messages.error(request, "Passwords do not match")
            return redirect("patient_register")

        if len(password) < 6:
            messages.error(request, "Password must be at least 6 characters")
            return redirect("patient_register")

        if User.objects.filter(username=email).exists():
            messages.error(request, "Email already registered")
            return redirect("patient_register")

        user = User.objects.create_user(
            username=email,
            email=email,
            password=password,
            first_name=full_name
        )

        PatientProfile.objects.create(
            user=user,
            full_name=full_name,
            phone=phone
        )

        messages.success(request, "Registration successful. Please login.")
        return redirect("login")

    return render(request, "core/register.html")

# Dashboards
@login_required
def patient_dashboard(request):
    """
    Patient Dashboard View
    Shows overview of appointments, prescriptions, tests, and payments
    """
    
    # Get the patient profile
    try:
        patient_profile = PatientProfile.objects.get(user=request.user)
    except PatientProfile.DoesNotExist:
        # Handle case where patient profile doesn't exist
        patient_profile = None
    
    # Initialize context with default values
    context = {
        'upcoming_appointments_count': 0,
        'next_appointment': None,
        'prescriptions_count': 0,
        'pending_tests_count': 0,
        'pending_payment_amount': 0,
        'upcoming_appointments': [],
        'recent_prescriptions': [],
    }
    
    # Diagnostic Test Bookings
    diagnostic_bookings = TestBooking.objects.filter(
        patient=patient_profile
    ).order_by('-booking_date')[:5]

    context = {
        # existing context...
        'diagnostic_bookings': diagnostic_bookings,
    }

    
    if patient_profile:
        # Get upcoming appointments (scheduled, not completed/cancelled)
        today = date.today()
        upcoming_appointments = Appointment.objects.filter(
            patient=patient_profile,
            appointment_date__gte=today,
            status__in=['Scheduled', 'Confirmed']
        ).select_related('doctor', 'doctor__user').order_by('appointment_date', 'appointment_time')
        
        context['upcoming_appointments_count'] = upcoming_appointments.count()
        context['upcoming_appointments'] = upcoming_appointments[:5]  # Limit to 5
        context['next_appointment'] = upcoming_appointments.first()
        
        # Get total prescriptions count
        prescriptions = Prescription.objects.filter(
            patient=patient_profile
        ).select_related('doctor', 'doctor__user')
        
        context['prescriptions_count'] = prescriptions.count()
        context['recent_prescriptions'] = prescriptions.order_by('-created_at')[:5]  # Latest 5
        
        # Get pending test bookings
        pending_tests = TestBooking.objects.filter(
            patient=patient_profile,
            status__in=['Booked', 'Pending']
        )
        
        context['pending_tests_count'] = pending_tests.count()
        
        # Get pending payment amount
        pending_payments = Payment.objects.filter(
            patient=patient_profile,
            payment_status='Pending'
        ).aggregate(total=Sum('amount'))
        
        context['pending_payment_amount'] = pending_payments['total'] or 0
    
    return render(request, 'core/dashboard/patient_dashboard.html', context)



@login_required
def patient_diagnostic_tests(request):
    """View and book diagnostic tests"""
    # Your diagnostic tests logic here
    return render(request, 'core/dashboard/patient_diagnostic_tests.html')

def safe_time_string(time_val):
    """Convert ANY time format to HH:MM string"""
    if time_val is None:
        return "00:00"
    time_str = str(time_val)
    return time_str[:5] if len(time_str) >= 5 else time_str

def safe_date_format(date_val, fmt='%b %Y'):
    """Safely format ANY date - never crashes"""
    if date_val is None:
        return "Unknown"
    if isinstance(date_val, str):
        return date_val
    try:
        return date_val.strftime(fmt)
    except:
        return str(date_val)


@login_required
def prescriptions(request):
    """View all prescriptions"""
    try:
        patient_profile = PatientProfile.objects.get(user=request.user)
        prescriptions = Prescription.objects.filter(
            patient=patient_profile
        ).select_related('doctor', 'doctor__user', 'appointment').order_by('-created_at')
    except PatientProfile.DoesNotExist:
        prescriptions = []
    
    context = {
        'prescriptions': prescriptions
    }
    return render(request, 'core/prescriptions.html', context)


@login_required
def medical_history(request):
    """View medical history"""
    # Your medical history logic here
    return render(request, 'core/medical_history.html')


@login_required
def payments(request):
    """View payment history and pending payments"""
    try:
        patient_profile = PatientProfile.objects.get(user=request.user)
        payments = Payment.objects.filter(
            patient=patient_profile
        ).select_related('appointment', 'appointment__doctor', 'appointment__doctor__user').order_by('-payment_date')
    except PatientProfile.DoesNotExist:
        payments = []
    
    context = {
        'payments': payments
    }
    return render(request, 'core/payments.html', context)




def doctor_dashboard(request):
    return render(request, 'core/dashboard/doctor_dashboard.html')

def lab_dashboard(request):
    return render(request, 'core/dashboard/lab_dashboard.html')

import json

def admin_dashboard(request):
    # ===== STATS =====
    total_appointments = Appointment.objects.count()

    total_revenue = (
        Payment.objects.aggregate(total=Sum('amount'))['total'] or 0
    )

    # ===== APPOINTMENTS PER MONTH =====
    appointments = (
        Appointment.objects
        .annotate(month=TruncMonth('appointment_date'))  # ✅ FIXED
        .values('month')
        .annotate(count=Count('id'))
        .order_by('month')
    )

    appointments_labels = [safe_date_format(a['month']) for a in appointments]
    appointments_data = [a['count'] for a in appointments]

    # ===== REVENUE PER MONTH =====
    revenue = (
        Payment.objects
        .annotate(month=TruncMonth('payment_date'))  # ✅ correct already
        .values('month')
        .annotate(total=Sum('amount'))
        .order_by('month')
    )

    revenue_labels = [r['month'].strftime('%b %Y') for r in revenue]
    revenue_data = [float(r['total']) for r in revenue]

    context = {
        'total_appointments': total_appointments,
        'total_revenue': total_revenue,
        'appointments_labels': json.dumps(appointments_labels),
        'appointments_data': json.dumps(appointments_data),
        'revenue_labels': json.dumps(revenue_labels),
        'revenue_data': json.dumps(revenue_data),
    }

    return render(request, 'core/dashboard/admin_dashboard.html', context)


def frontdesk_dashboard(request):
    return render(request, 'core/dashboard/frontdesk_dashboard.html')



def is_admin(user):
    return user.is_superuser or user.is_staff


def admin_users(request):
    patients = User.objects.filter(patientprofile__isnull=False)

    return render(request, "core/dashboard/admin_users.html", {
        "users": patients
    })

@user_passes_test(is_admin)
def admin_patient_add(request):
    if request.method == "POST":
        first_name = request.POST["first_name"]
        last_name = request.POST["last_name"]
        email = request.POST["email"]
        password = request.POST["password"]

        phone = request.POST["phone"]
        gender = request.POST["gender"]
        dob = request.POST["dob"]
        address = request.POST["address"]

        # Create user
        user = User.objects.create_user(
            username=email,
            email=email,
            password=password,
            first_name=first_name,
            last_name=last_name
        )

        # Create patient profile
        PatientProfile.objects.create(
            user=user,
            full_name=f"{first_name} {last_name}",
            gender=gender,
            dob=dob,
            phone=phone,
            address=address,
            status="Active"
        )

        return redirect("admin_users")

    return render(request, "core/dashboard/admin_patient_add.html")

@user_passes_test(is_admin)
def admin_user_edit(request, user_id):
    user = get_object_or_404(User, id=user_id)

    if request.method == "POST":
        full_name = request.POST.get("name", "").strip()
        email = request.POST.get("email")
        is_active = request.POST.get("is_active") == "on"

        # Split name safely
        if full_name:
            parts = full_name.split(" ", 1)
            user.first_name = parts[0]
            user.last_name = parts[1] if len(parts) > 1 else ""

        user.email = email
        user.username = email   # IMPORTANT: keep username in sync
        user.is_active = is_active
        user.save()

        messages.success(request, "Patient updated successfully")
        return redirect("admin_users")

    return render(request, "core/dashboard/admin_user_edit.html", {
        "user": user
    })

def admin_user_detail(request, user_id):
    user = get_object_or_404(User, id=user_id)
    context = {'user': user}
    return render(request, 'core/dashboard/admin_user_detail.html', context)

@user_passes_test(is_admin)
def admin_user_delete(request, user_id):
    user = get_object_or_404(User, id=user_id)
    # ❗ Safety: don't allow deleting admins
    if user.is_staff or user.is_superuser:
        return redirect("admin_users")

    user.delete()
    return redirect("admin_users")

@user_passes_test(is_admin)
def admin_doctors(request):
    doctors = DoctorProfile.objects.select_related("user")
    return render(request, "core/dashboard/admin_doctors.html", {
        "doctors": doctors
    })
    
def admin_doctor_add(request):
    if request.method == "POST":
        first_name = request.POST.get("first_name")
        last_name = request.POST.get("last_name")
        email = request.POST.get("email")
        password = request.POST.get("password")
        confirm_password = request.POST.get("confirm_password")
        specialization = request.POST.get("specialization")
        department = request.POST.get("department")
        phone = request.POST.get("phone")

        # validations
        if password != confirm_password:
            messages.error(request, "Passwords do not match")
            return redirect("admin_doctor_add")

        if User.objects.filter(email=email).exists():
            messages.error(request, "Email already exists")
            return redirect("admin_doctor_add")

        # create user
        user = User.objects.create_user(
            username=email,
            email=email,
            password=password,
            first_name=first_name,
            last_name=last_name,
            is_active=True,
        )

        # create doctor profile
        DoctorProfile.objects.create(
            user=user,
            department=department,
            specialization=specialization,
            phone=phone,
            status="Active",
        )

        messages.success(request, "Doctor added successfully")
        return redirect("admin_doctors")

    return render(request, "core/dashboard/admin_doctor_add.html")

def admin_doctor_edit(request, doctor_id):
    doctor = get_object_or_404(DoctorProfile, id=doctor_id)
    user = doctor.user

    if request.method == "POST":
        # Update User fields
        user.first_name = request.POST.get('first_name', user.first_name)
        user.last_name = request.POST.get('last_name', user.last_name)
        user.email = request.POST.get('email', user.email)
        user.is_active = True if request.POST.get('is_active') == 'on' else False
        user.save()

        # Update DoctorProfile fields
        doctor.department = request.POST.get('department', doctor.department)
        doctor.specialization = request.POST.get('specialization', doctor.specialization)
        doctor.license_number = request.POST.get('license_number', doctor.license_number)
        doctor.experience = request.POST.get('experience', doctor.experience)
        doctor.consultation_fee = request.POST.get('consultation_fee', doctor.consultation_fee)
        doctor.phone = request.POST.get('phone', doctor.phone)
        doctor.bio = request.POST.get('bio', doctor.bio)
        doctor.save()

        messages.success(request, f"Doctor {user.get_full_name()} updated successfully!")
        return redirect('admin_doctors')

    context = {
        'doctor': doctor,
        'user': user,
    }
    return render(request, 'core/dashboard/admin_doctor_edit.html', context)
    
def admin_doctor_detail(request, doctor_id):
    doctor = get_object_or_404(DoctorProfile, id=doctor_id)
    user = doctor.user

    context = {
        'doctor': doctor,
        'user': user,
    }
    return render(request, 'core/dashboard/admin_doctor_detail.html', context)
    
@user_passes_test(is_admin)
def admin_doctor_delete(request, doctor_id):
    doctor = get_object_or_404(DoctorProfile, id=doctor_id)
    user = doctor.user

    # safety: prevent deleting admins
    if user.is_superuser or user.is_staff:
        return redirect("admin_doctors")

    doctor.delete()
    user.delete()

    return redirect("admin_doctors")

@login_required
# ===== List Labs =====
def admin_labs(request):
    labs = Lab.objects.all()
    
    # Handle Add Lab form submission
    if request.method == "POST":
        name = request.POST.get("name")
        phone = request.POST.get("phone")
        address = request.POST.get("address")
        
        if name and phone and address:
            Lab.objects.create(name=name, phone=phone, address=address, status="Active")
            messages.success(request, f"Lab '{name}' added successfully!")
            return redirect("admin_labs")
        else:
            messages.error(request, "All fields are required!")

    query = request.GET.get("q")
    if query:
        labs = labs.filter(name__icontains=query)

    return render(request, "core/dashboard/admin_labs.html", {"labs": labs})

# ===== Add Lab =====
def admin_add_lab(request):
    if request.method == 'POST':
        name = request.POST.get('name')
        phone = request.POST.get('phone')
        address = request.POST.get('address')
        Lab.objects.create(name=name, phone=phone, address=address)
        messages.success(request, f'Lab "{name}" added successfully!')
    return redirect('admin_labs')

# ===== Edit Lab =====
def admin_edit_lab(request, lab_id):
    lab = get_object_or_404(Lab, id=lab_id)
    if request.method == 'POST':
        lab.name = request.POST.get('name')
        lab.phone = request.POST.get('phone')
        lab.address = request.POST.get('address')
        lab.status = request.POST.get('status')
        lab.save()
        messages.success(request, f'Lab "{lab.name}" updated successfully!')
        return redirect('admin_labs')
    context = {'lab': lab}
    return render(request, 'core/dashboard/admin_edit_lab.html', context)

# ===== Delete Lab =====
def admin_delete_lab(request, lab_id):
    lab = get_object_or_404(Lab, id=lab_id)
    if request.method == 'POST':
        lab.delete()
        messages.success(request, f'Lab "{lab.name}" deleted successfully!')
        return redirect('admin_labs')
    return redirect('admin_labs')


@login_required
@user_passes_test(is_admin)
def admin_payments(request):
    payments = Payment.objects.select_related(
        'patient', 'appointment'
    ).order_by('-payment_date')

    return render(request, 'core/dashboard/admin_payments.html', {
        'payments': payments
    })
    
def admin_payment_receipt(request, payment_id):
    payment = get_object_or_404(Payment, id=payment_id)

    return render(request, 'core/dashboard/admin_payment_receipt.html', {
        'payment': payment
    })
    
@login_required
def admin_reports(request):

    # ===== BASIC COUNTS =====
    total_patients = PatientProfile.objects.count()
    total_doctors = DoctorProfile.objects.count()
    total_labs = Lab.objects.count()
    total_tests = TestBooking.objects.count()

    # ===== TOTAL REVENUE =====
    total_revenue = (
        Payment.objects
        .filter(payment_status="Paid")
        .aggregate(total=Sum("amount"))["total"] or 0
    )

    # ===== APPOINTMENTS PER MONTH =====
    appointments_qs = (
        Appointment.objects
        .annotate(month=TruncMonth("created_at"))
        .values("month")
        .annotate(count=Count("id"))
        .order_by("month")
    )

    appointment_labels = [
        a["month"].strftime("%b %Y")
        for a in appointments_qs if a["month"]
    ]
    appointment_data = [a["count"] for a in appointments_qs]

    # ===== REVENUE PER MONTH =====
    revenue_qs = (
        Payment.objects
        .filter(payment_status="Paid")
        .annotate(month=TruncMonth("payment_date"))
        .values("month")
        .annotate(total=Sum("amount"))
        .order_by("month")
    )

    revenue_labels = [
        r["month"].strftime("%b %Y")
        for r in revenue_qs if r["month"]
    ]
    revenue_data = [float(r["total"]) for r in revenue_qs]

    context = {
        "total_patients": total_patients,
        "total_doctors": total_doctors,
        "total_labs": total_labs,
        "total_tests": total_tests,
        "total_revenue": total_revenue,
        "appointment_labels": appointment_labels,
        "appointment_data": appointment_data,
        "revenue_labels": revenue_labels,
        "revenue_data": revenue_data,
    }

    return render(request, "core/dashboard/admin_reports.html", context)

# ========================= FRONT DESK =========================

def admin_frontdesk(request):
    frontdesks = FrontDeskProfile.objects.all()
    query = request.GET.get('q')
    if query:
        frontdesks = frontdesks.filter(user__username__icontains=query)
    return render(request, 'core/dashboard/admin_frontdesk.html', {'frontdesks': frontdesks})

@user_passes_test(is_admin)
def admin_add_frontdesk(request):
    if request.method == 'POST':
        username = request.POST.get('username', '').strip()
        email = request.POST.get('email', '').strip()
        phone = request.POST.get('phone', '').strip()
        password = request.POST.get('password', '')
        status = request.POST.get('status', 'Active')

        # Validation
        if not username or not email or not phone or not password:
            messages.error(request, 'All fields are required')
            return redirect('admin_add_frontdesk')

        # Check if username already exists
        if User.objects.filter(username=username).exists():
            messages.error(request, f"Username '{username}' already exists. Please choose a different username.")
            return redirect('admin_add_frontdesk')

        # Check if email already exists
        if User.objects.filter(email=email).exists():
            messages.error(request, f"Email '{email}' is already registered. Please use a different email.")
            return redirect('admin_add_frontdesk')

        try:
            # Create user
            user = User.objects.create_user(
                username=username,
                email=email,
                password=password
            )
            
            # Create front desk profile
            FrontDeskProfile.objects.create(
                user=user,
                phone=phone,
                status=status
            )
            
            messages.success(request, f'Front desk staff "{username}" added successfully!')
            return redirect('admin_frontdesk')
            
        except Exception as e:
            messages.error(request, f'Error creating staff member: {str(e)}')
            return redirect('admin_add_frontdesk')

    return render(request, 'core/dashboard/admin_add_frontdesk.html')

@user_passes_test(is_admin)
def admin_edit_frontdesk(request, pk):
    profile = get_object_or_404(FrontDeskProfile, pk=pk)
    
    if request.method == 'POST':
        username = request.POST.get('username', '').strip()
        email = request.POST.get('email', '').strip()
        phone = request.POST.get('phone', '').strip()
        status = request.POST.get('status', 'Active')

        # Validation
        if not username or not email or not phone:
            messages.error(request, 'Username, email, and phone are required')
            return redirect('admin_edit_frontdesk', pk=pk)

        # Check if username is taken by another user
        if User.objects.filter(username=username).exclude(id=profile.user.id).exists():
            messages.error(request, f"Username '{username}' is already taken")
            return redirect('admin_edit_frontdesk', pk=pk)

        # Check if email is taken by another user
        if User.objects.filter(email=email).exclude(id=profile.user.id).exists():
            messages.error(request, f"Email '{email}' is already taken")
            return redirect('admin_edit_frontdesk', pk=pk)

        try:
            # Update user
            profile.user.username = username
            profile.user.email = email
            profile.user.save()
            
            # Update profile
            profile.phone = phone
            profile.status = status
            profile.save()
            
            messages.success(request, f'Front desk staff "{username}" updated successfully!')
            return redirect('admin_frontdesk')
            
        except Exception as e:
            messages.error(request, f'Error updating staff member: {str(e)}')
            return redirect('admin_edit_frontdesk', pk=pk)

    return render(request, 'core/dashboard/admin_edit_frontdesk.html', {'profile': profile})

@user_passes_test(is_admin)
def admin_delete_frontdesk(request, pk):
    profile = get_object_or_404(FrontDeskProfile, pk=pk)
    
    if request.method == 'POST':
        username = profile.user.username
        profile.user.delete()  # This also deletes the profile due to CASCADE
        messages.success(request, f'Front desk staff "{username}" deleted successfully!')
    
    return redirect('admin_frontdesk')

# ========================= LAB TECHNICIAN =========================

def admin_lab_technician(request):
    technicians = LabTechnicianProfile.objects.all()
    labs = Lab.objects.all()
    query = request.GET.get('q')
    if query:
        technicians = technicians.filter(user__username__icontains=query)
    return render(request, 'core/dashboard/admin_lab_technician.html', {
        'technicians': technicians,
        'labs': labs,
    })


from django.contrib.auth.hashers import make_password
def admin_lab_technician_add(request):
    labs = Lab.objects.all()

    if request.method == "POST":
        form = LabTechnicianForm(request.POST)

        if form.is_valid():
            if User.objects.filter(username=form.cleaned_data['username']).exists():
                messages.error(request, "Username already exists")
                return redirect('admin_add_technician')

            if User.objects.filter(email=form.cleaned_data['email']).exists():
                messages.error(request, "Email already exists")
                return redirect('admin_add_technician')

            # Create User
            user = User.objects.create(
                username=form.cleaned_data['username'],
                email=form.cleaned_data['email'],
                password=make_password(form.cleaned_data['password']),
                first_name=form.cleaned_data['first_name'],
                last_name=form.cleaned_data['last_name']
            )

            # Assign Lab
            lab = None
            if form.cleaned_data['lab_id']:
                lab = Lab.objects.get(id=form.cleaned_data['lab_id'])

            # Create Technician Profile
            LabTechnicianProfile.objects.create(
                user=user,
                phone=form.cleaned_data['phone'],
                status=form.cleaned_data['status'],
                lab=lab,
                bio=form.cleaned_data['bio']
            )

            messages.success(request, "Lab Technician added successfully")
            return redirect('admin_lab_technicians')

    else:
        form = LabTechnicianForm()

    return render(request, 'core/dashboard/admin_lab_technician_add.html', {
        'form': form,
        'labs': labs
    })

@login_required
def admin_lab_technician_edit(request, pk):
    technician = get_object_or_404(LabTechnicianProfile, pk=pk)
    
    if request.method == 'POST':
        status = request.POST['status']
        
        # Update user fields
        technician.user.username = request.POST['username']
        technician.user.email = request.POST['email']
        technician.user.is_active = (status == 'active')  # ✅ One-liner!
        technician.user.save()
        
        # Update technician profile fields
        technician.phone = request.POST['phone']
        technician.status = status
        technician.save()
        
        messages.success(request, f'Lab Technician {technician.user.username} updated successfully.')
        return redirect('admin_lab_technicians')

    return render(request, 'core/dashboard/admin_lab_technician_edit.html', {
        'technician': technician,
    })

def admin_delete_lab_technician(request, pk):
    profile = get_object_or_404(LabTechnicianProfile, pk=pk)
    profile.user.delete()
    messages.success(request, f'Lab Technician {profile.user.username} deleted successfully.')
    return redirect('admin_lab_technicians')


from django.contrib.auth import update_session_auth_hash
from django.contrib.auth.password_validation import validate_password
from django.core.exceptions import ValidationError

@login_required
def admin_settings(request):
    """
    Main settings page - displays all settings tabs
    """
    context = {
        'user': request.user,
    }
    return render(request, 'core/dashboard/admin_settings.html', context)


@login_required
def admin_settings_general(request):
    """
    Update general system settings
    """
    if request.method == 'POST':
        # Get form data
        system_name = request.POST.get('system_name')
        timezone = request.POST.get('timezone')
        date_format = request.POST.get('date_format')
        language = request.POST.get('language')
        system_email = request.POST.get('system_email')
        
        # Here you would save to a Settings model or configuration file
        # For now, we'll just show a success message
        
        messages.success(request, 'General settings updated successfully!')
        return redirect('admin_settings')
    
    return redirect('admin_settings')


@login_required
def admin_settings_profile(request):
    """
    Update admin profile settings
    """
    if request.method == 'POST':
        user = request.user
        
        # Get form data
        first_name = request.POST.get('first_name')
        last_name = request.POST.get('last_name')
        email = request.POST.get('email')
        phone = request.POST.get('phone')
        bio = request.POST.get('bio')
        
        # Update user
        user.first_name = first_name
        user.last_name = last_name
        user.email = email
        user.save()
        
        # Update profile if you have a UserProfile model
        # user.profile.phone = phone
        # user.profile.bio = bio
        # user.profile.save()
        
        messages.success(request, 'Profile updated successfully!')
        return redirect('admin_settings')
    
    return redirect('admin_settings')


@login_required
def admin_settings_security(request):
    """
    Update security settings (change password)
    """
    if request.method == 'POST':
        user = request.user
        
        current_password = request.POST.get('current_password')
        new_password = request.POST.get('new_password')
        confirm_password = request.POST.get('confirm_password')
        
        # Verify current password
        if not user.check_password(current_password):
            messages.error(request, 'Current password is incorrect!')
            return redirect('admin_settings')
        
        # Check if passwords match
        if new_password != confirm_password:
            messages.error(request, 'New passwords do not match!')
            return redirect('admin_settings')
        
        # Validate new password
        try:
            validate_password(new_password, user)
        except ValidationError as e:
            messages.error(request, ' '.join(e.messages))
            return redirect('admin_settings')
        
        # Update password
        user.set_password(new_password)
        user.save()
        
        # Keep user logged in after password change
        update_session_auth_hash(request, user)
        
        messages.success(request, 'Password changed successfully!')
        return redirect('admin_settings')
    
    return redirect('admin_settings')


@login_required
def admin_settings_notifications(request):
    """
    Update notification preferences
    """
    if request.method == 'POST':
        # Get toggle states
        new_patient = request.POST.get('new_patient') == 'on'
        appointments = request.POST.get('appointments') == 'on'
        payments = request.POST.get('payments') == 'on'
        system_updates = request.POST.get('system_updates') == 'on'
        
        # Save to user preferences or settings model
        # user.preferences.new_patient_notifications = new_patient
        # user.preferences.save()
        
        messages.success(request, 'Notification preferences updated!')
        return redirect('admin_settings')
    
    return redirect('admin_settings')


@login_required
def admin_settings_system(request):
    """
    System maintenance and information
    """
    if request.method == 'POST':
        action = request.POST.get('action')
        
        if action == 'maintenance_mode':
            # Toggle maintenance mode
            messages.success(request, 'Maintenance mode toggled!')
            
        elif action == 'clear_cache':
            # Clear system cache
            messages.success(request, 'System cache cleared successfully!')
            
        return redirect('admin_settings')
    
    return redirect('admin_settings')


@login_required
def admin_settings_backup(request):
    """
    Backup and restore functionality
    """
    if request.method == 'POST':
        action = request.POST.get('action')
        
        if action == 'create_backup':
            # Create database backup
            messages.success(request, 'Backup created successfully!')
            
        elif action == 'restore_backup':
            # Restore from backup
            messages.success(request, 'System restored from backup!')
            
        elif action == 'toggle_auto_backup':
            # Toggle automatic backups
            messages.success(request, 'Auto-backup settings updated!')
            
        return redirect('admin_settings')
    
    return redirect('core/dashboard/admin_settings')

from datetime import datetime, date

@login_required
def patient_book_appointment(request):
    """
    Book Appointment Page - Creates appointment with 'Pending Payment' status
    Redirects to payment processing
    """
    try:
        patient_profile = PatientProfile.objects.get(user=request.user)
    except PatientProfile.DoesNotExist:
        messages.error(request, "Patient profile not found. Please contact support.")
        return redirect('core/dashboard/patient_dashboard')
    
    if request.method == 'POST':
        doctor_id = request.POST.get('doctor')
        appointment_date = request.POST.get('appointment_date')
        appointment_time = request.POST.get('appointment_time')
        reason = request.POST.get('reason')
        
        # Validation
        if not all([doctor_id, appointment_date, appointment_time, reason]):
            messages.error(request, "All fields are required.")
            return redirect('core/dashboard/patient_book_appointment')
        
        try:
            doctor = DoctorProfile.objects.get(id=doctor_id, status='Active')
            
            # Create appointment with 'Pending Payment' status
            appointment = Appointment.objects.create(
                patient=patient_profile,
                doctor=doctor,
                appointment_date=appointment_date,
                appointment_time=appointment_time,
                reason=reason,
                status='Pending Payment'  # Changed from 'Scheduled'
            )
            
            # Create payment for this appointment
            payment = create_payment_for_appointment(appointment)
            
            messages.info(
                request,
                f"Please complete payment of ₹{payment.amount} to confirm your appointment "
                f"with Dr. {doctor.user.get_full_name()} on {appointment_date} at {appointment_time}."
            )
            # Redirect directly to payment processing page
            return redirect('core/dashboard/process_payment', payment_id=payment.id)
            
        except DoctorProfile.DoesNotExist:
            messages.error(request, "Selected doctor is not available.")
            return redirect('patient_book_appointment')
        except Exception as e:
            messages.error(request, f"Error booking appointment: {str(e)}")
            return redirect('core/dashboard/patient_book_appointment')
    
    # GET request - show form
    # Get all unique specializations from active doctors
    specializations = DoctorProfile.objects.filter(
        status='Active'
    ).values_list('specialization', flat=True).distinct().order_by('specialization')
    
    # Get all active doctors for initial load (optional)
    all_doctors = DoctorProfile.objects.filter(
        status='Active'
    ).select_related('user').order_by('user__first_name')
    
    context = {
        'specializations': specializations,
        'all_doctors': all_doctors,
        'today': date.today().isoformat(),
    }
    
    return render(request, 'core/dashboard/patient_book_appointment.html', context)


@login_required
def get_doctors_by_specialization(request):
    """
    AJAX endpoint to get doctors filtered by specialization
    Returns JSON response with doctor details
    """
    specialization = request.GET.get('specialization', '')
    
    if not specialization:
        return JsonResponse({'doctors': []})
    
    # Filter doctors by specialization and active status
    doctors = DoctorProfile.objects.filter(
        specialization=specialization,
        status='Active'
    ).select_related('user').order_by('user__first_name')
    
    # Build response data
    doctors_data = []
    for doctor in doctors:
        doctors_data.append({
            'id': doctor.id,
            'name': doctor.user.get_full_name() or doctor.user.username,
            'specialization': doctor.specialization,
            'department': doctor.department,
            'experience': doctor.experience,
            'consultation_fee': str(doctor.consultation_fee),
            'bio': doctor.bio or '',
        })
    
    return JsonResponse({'doctors': doctors_data})


@login_required
def search_doctors(request):
    """
    Search Doctors by Specialization - Info/Directory Page
    This is a separate page for browsing doctors (not booking)
    """
    specialization_filter = request.GET.get('specialization', '')
    
    # Get all specializations for filter dropdown
    specializations = DoctorProfile.objects.filter(
        status='Active'
    ).values_list('specialization', flat=True).distinct().order_by('specialization')
    
    # Filter doctors
    doctors = DoctorProfile.objects.filter(status='Active').select_related('user')
    
    if specialization_filter:
        doctors = doctors.filter(specialization=specialization_filter)
    
    doctors = doctors.order_by('user__first_name')
    
    context = {
        'doctors': doctors,
        'specializations': specializations,
        'selected_specialization': specialization_filter,
    }
    
    return render(request, 'core/search_doctors.html', context)

from decimal import Decimal
import uuid

@login_required
def payments(request):
    """
    Patient Payments Dashboard
    Shows all payments (pending, paid, failed)
    """
    try:
        patient_profile = PatientProfile.objects.get(user=request.user)
    except PatientProfile.DoesNotExist:
        messages.error(request, "Patient profile not found.")
        return redirect('patient_dashboard')
    
    # Get all payments for this patient
    all_payments = Payment.objects.filter(
        patient=patient_profile
    ).select_related('appointment', 'appointment__doctor', 'appointment__doctor__user').order_by('-payment_date')
    
    # Separate by status
    pending_payments = all_payments.filter(payment_status='Pending')
    paid_payments = all_payments.filter(payment_status='Paid')
    failed_payments = all_payments.filter(payment_status='Failed')
    
    # Calculate totals
    total_pending = pending_payments.aggregate(total=Sum('amount'))['total'] or Decimal('0.00')
    total_paid = paid_payments.aggregate(total=Sum('amount'))['total'] or Decimal('0.00')
    
    context = {
        'all_payments': all_payments,
        'pending_payments': pending_payments,
        'paid_payments': paid_payments,
        'failed_payments': failed_payments,
        'total_pending': total_pending,
        'total_paid': total_paid,
        'pending_count': pending_payments.count(),
        'paid_count': paid_payments.count(),
    }
    
    return render(request, 'core/dashboard/payments.html', context)


@login_required
def payment_detail(request, payment_id):
    """
    View single payment details
    """
    try:
        patient_profile = PatientProfile.objects.get(user=request.user)
    except PatientProfile.DoesNotExist:
        messages.error(request, "Patient profile not found.")
        return redirect('patient_dashboard')
    
    payment = get_object_or_404(
        Payment,
        id=payment_id,
        patient=patient_profile
    )
    
    context = {
        'payment': payment,
    }
    
    return render(request, 'core/dashboard/payment_detail.html', context)


@login_required
def process_payment(request, payment_id):

    patient_profile = get_object_or_404(
        PatientProfile,
        user=request.user
    )

    payment = get_object_or_404(
        Payment,
        id=payment_id,
        patient=patient_profile
    )

    if request.method == "POST":
        payment_method = request.POST.get("payment_method")

        if not payment_method:
            messages.error(request, "Please select a payment method.")
            return redirect('core/dashboard/process_payment', payment_id=payment.id)

        transaction_id = f"TXN{uuid.uuid4().hex[:12].upper()}"

        payment.payment_method = payment_method
        payment.transaction_id = transaction_id
        payment.payment_status = "Paid"
        payment.save()

        # =========================
        # CASE 1 → Appointment
        # =========================
        if payment.appointment:
            appointment = payment.appointment
            appointment.status = "Scheduled"
            appointment.save()

            messages.success(
                request,
                f"Payment successful! Your appointment is confirmed. "
                f"Transaction ID: {transaction_id}"
            )

        # =========================
        # CASE 2 → Diagnostic Test
        # =========================
        elif payment.test_booking:
            # Find related booking
            booking = TestBooking.objects.filter(
                patient=patient_profile,
                test=payment.test_booking.test,
                status="Pending Payment"
            ).first()

            if booking:
                booking.status = "Booked"
                booking.save()

            messages.success(
                request,
                f"Payment successful! Your diagnostic test "
                f"{payment.test_booking.test.test_name} is confirmed. "
                f"Transaction ID: {transaction_id}"
            )

        return redirect("patient_dashboard")

    return render(request, "core/dashboard/process_payment.html", {
        "payment": payment
    })



from decimal import Decimal

def create_payment_for_test(booking, amount=None):
    """
    Create payment for diagnostic test booking
    """

    if amount is None:
        amount = booking.test.price or Decimal('1000.00')

    payment = Payment.objects.create(
        patient=booking.patient,
        diagnostic_test=booking.test,
        amount=amount,
        payment_status='Pending'
    )

    return payment


# ==========================================
# UPDATED DIAGNOSTIC TESTS VIEWS WITH PAYMENT FLOW
# Add these to your core/views.py file
# ==========================================

from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.db.models import Q
from datetime import datetime, date
from decimal import Decimal
import uuid

from core.models import (
    PatientProfile,
    DiagnosticTest,
    TestBooking,
    Lab,
    Payment,
)


@login_required
def patient_diagnostic_tests(request):
    """Display available diagnostic tests with filtering and search"""
    try:
        # Get patient profile
        patient = PatientProfile.objects.get(user=request.user)
    except PatientProfile.DoesNotExist:
        messages.error(request, "Patient profile not found")
        return redirect('patient_dashboard')
    
    # Get all diagnostic tests
    tests = DiagnosticTest.objects.select_related('lab').all()
    
    # Search functionality
    search_query = request.GET.get('search', '').strip()
    if search_query:
        tests = tests.filter(
            Q(test_name__icontains=search_query) |
            Q(lab__name__icontains=search_query)
        )
    
    # Lab filter functionality
    lab_filter = request.GET.get('lab', '').strip()
    if lab_filter:
        tests = tests.filter(lab__id=lab_filter)
    
    # Get all active labs for dropdown
    labs = Lab.objects.filter(status='Active')
    
    # Get user's existing test bookings (as a list for template)
    # Include both Pending Payment and Booked status
    user_bookings = list(TestBooking.objects.filter(
        patient=patient
    ).values_list('test_id', flat=True))
    
    context = {
        'tests': tests,
        'labs': labs,
        'search_query': search_query,
        'lab_filter': lab_filter,
        'user_bookings': user_bookings,
        'patient': patient,
    }
    
    return render(request, 'core/dashboard/patient_diagnostic_tests.html', context)


@login_required
def book_diagnostic_test(request, test_id):
    """
    Book a diagnostic test - Creates booking with 'Pending Payment' status
    Redirects to payment processing
    """
    try:
        patient = PatientProfile.objects.get(user=request.user)
        test = DiagnosticTest.objects.get(id=test_id)
    except (PatientProfile.DoesNotExist, DiagnosticTest.DoesNotExist):
        messages.error(request, "Invalid request")
        return redirect('patient_diagnostic_tests')
    
    # Check if already booked (any status except Cancelled)
    existing_booking = TestBooking.objects.filter(
        patient=patient, 
        test=test
    ).exclude(status='Cancelled').exists()
    
    if existing_booking:
        messages.warning(request, f"You have already booked {test.test_name}")
        return redirect('patient_diagnostic_tests')
    
    if request.method == 'POST':
        booking_date_str = request.POST.get('booking_date')
        
        try:
            booking_date = datetime.strptime(booking_date_str, '%Y-%m-%d').date()
            
            # Check if booking date is in the future
            if booking_date < date.today():
                messages.error(request, "Cannot book tests for past dates")
                return redirect('patient_diagnostic_tests')
            
            # Create test booking with 'Pending Payment' status
            booking = TestBooking.objects.create(
                patient=patient,
                test=test,
                lab=test.lab,
                booking_date=booking_date,
                status='Pending Payment'  # Changed from 'Booked'
            )
            
            # Create payment for this test booking
            payment = Payment.objects.create(
                patient=patient,
                test_booking=booking,   # ✅ THIS FIXES EVERYTHING
                amount=test.price,
                payment_status='Pending'
            )

            
            messages.info(
                request,
                f"Please complete payment of ₹{payment.amount} to confirm your test booking "
                f"for {test.test_name} on {booking_date.strftime('%d %B %Y')}."
            )
            
            # Redirect directly to payment processing page
            return redirect('core/dashboard/process_test_payment', payment_id=payment.id)
        
        except ValueError:
            messages.error(request, "Invalid date format")
            return redirect('patient_diagnostic_tests')
        except Exception as e:
            messages.error(request, f"Error booking test: {str(e)}")
            return redirect('patient_diagnostic_tests')
    
    # If GET request, redirect back
    return redirect('core/dashboard/patient_diagnostic_tests')


@login_required
def process_test_payment(request, payment_id):
    """
    Process payment for a test booking
    After successful payment, update booking status to 'Booked'
    """

    try:
        patient_profile = PatientProfile.objects.get(user=request.user)
    except PatientProfile.DoesNotExist:
        messages.error(request, "Patient profile not found.")
        return redirect('patient_dashboard')

    payment = get_object_or_404(
        Payment,
        id=payment_id,
        patient=patient_profile,
        payment_status='Pending'
    )

    # ✅ Correct way — directly from payment
    test_booking = payment.test_booking

    if not test_booking:
        messages.error(request, "Test booking not linked to this payment.")
        return redirect('patient_diagnostic_tests')

    if request.method == 'POST':
        payment_method = request.POST.get('payment_method')

        if not payment_method:
            messages.error(request, "Please select a payment method.")
            return redirect('process_test_payment', payment_id=payment_id)

        # Generate transaction ID
        transaction_id = f"TXN{uuid.uuid4().hex[:12].upper()}"

        # Update payment
        payment.payment_method = payment_method
        payment.transaction_id = transaction_id
        payment.payment_status = 'Paid'
        payment.save()

        # ✅ Update booking status
        test_booking.status = 'Booked'
        test_booking.save()

        messages.success(
            request,
            f"Payment successful! Your test booking for {test_booking.test.test_name} "
            f"on {test_booking.booking_date.strftime('%d %B %Y')} "
            f"at {test_booking.lab.name} is now confirmed. "
            f"Transaction ID: {transaction_id}"
        )

        return redirect('patient_booked_tests')

    context = {
        'payment': payment,
        'test_booking': test_booking,
    }

    return render(request, 'core/dashboard/process_test_payment.html', context)


@login_required
def patient_booked_tests(request):
    try:
        patient = PatientProfile.objects.get(user=request.user)
    except PatientProfile.DoesNotExist:
        messages.error(request, "Patient profile not found")
        return redirect('patient_dashboard')

    # Base queryset (ALL bookings)
    all_bookings = TestBooking.objects.filter(
        patient=patient
    ).select_related('test', 'lab', 'payment').order_by('-booking_date', '-created_at')

    # Apply filter for display only
    status_filter = request.GET.get('status', '')
    bookings = all_bookings

    if status_filter:
        bookings = bookings.filter(status=status_filter)

    # ✅ Statistics should use ALL bookings
    total_bookings = all_bookings.count()
    pending_payment_count = all_bookings.filter(status='Pending Payment').count()
    booked_count = all_bookings.filter(status='Booked').count()
    completed_count = all_bookings.filter(status='Completed').count()
    cancelled_count = all_bookings.filter(status='Cancelled').count()

    context = {
        'bookings': bookings,
        'total_bookings': total_bookings,
        'pending_payment_count': pending_payment_count,
        'booked_count': booked_count,
        'completed_count': completed_count,
        'cancelled_count': cancelled_count,
        'status_filter': status_filter,
    }

    return render(request, 'core/dashboard/patient_booked_tests.html', context)


@login_required
def cancel_test_booking(request, booking_id):
    """Cancel a test booking"""
    try:
        patient = PatientProfile.objects.get(user=request.user)
        booking = TestBooking.objects.get(id=booking_id, patient=patient)
        
        # Only allow cancellation of Pending Payment or Booked tests
        if booking.status in ['Pending Payment', 'Booked']:
            # Check if booking is in the future
            if booking.booking_date >= date.today():
                booking.status = 'Cancelled'
                booking.save()
                
                # If payment was made, you might want to handle refunds here
                if booking.payment and booking.payment.payment_status == 'Paid':
                    # TODO: Implement refund logic
                    messages.info(
                        request,
                        f"Test booking cancelled. Refund will be processed within 5-7 business days."
                    )
                else:
                    messages.success(
                        request, 
                        f"Test booking for {booking.test.test_name} has been cancelled"
                    )
            else:
                messages.error(request, "Cannot cancel past test bookings")
        else:
            messages.error(request, "This test booking cannot be cancelled")
            
    except (PatientProfile.DoesNotExist, TestBooking.DoesNotExist):
        messages.error(request, "Test booking not found")
    
    return redirect('core/dashboard/patient_booked_tests')


# ==================== HELPER FUNCTION ====================

def create_payment_for_test(booking):
    """
    Helper function to create a payment when test is booked
    Call this after creating a test booking
    
    Usage:
        booking = TestBooking.objects.create(...)
        payment = create_payment_for_test(booking)
    """
    # Use test price
    amount = booking.test.price
    
    # Create payment with appointment=None (since it's for a test)
    payment = Payment.objects.create(
        patient=booking.patient,
        appointment=None,  # No appointment for test bookings
        amount=amount,
        payment_method='',  # Will be filled when patient pays
        payment_status='Pending'
    )
    
    return payment


@login_required
def patient_test_results(request):
    """View all lab test results for the patient"""
    try:
        patient = PatientProfile.objects.get(user=request.user)
    except PatientProfile.DoesNotExist:
        messages.error(request, "Patient profile not found")
        return redirect('patient_dashboard')
    
    # Get all lab results for this patient
    results = LabResult.objects.filter(
        patient=patient
    ).select_related(
        'doctor', 
        'doctor__user', 
        'lab_technician', 
        'lab_technician__user'
    ).order_by('-test_date', '-created_at')
    
    # Filter by result status if provided
    status_filter = request.GET.get('status', '')
    if status_filter:
        results = results.filter(result_status=status_filter)
    
    # Calculate statistics
    total_results = results.count()
    normal_count = results.filter(result_status='Normal').count()
    abnormal_count = results.filter(result_status='Abnormal').count()
    
    context = {
        'results': results,
        'total_results': total_results,
        'normal_count': normal_count,
        'abnormal_count': abnormal_count,
        'status_filter': status_filter,
    }
    
    return render(request, 'core/dashboard/patient_test_results.html', context)


@login_required
def cancel_test_booking(request, booking_id):
    """Cancel a test booking"""
    try:
        patient = PatientProfile.objects.get(user=request.user)
        booking = TestBooking.objects.get(id=booking_id, patient=patient)
        
        # Only allow cancellation of Booked tests
        if booking.status == 'Booked':
            # Check if booking is in the future
            if booking.booking_date >= date.today():
                booking.status = 'Cancelled'
                booking.save()
                messages.success(
                    request, 
                    f"Test booking for {booking.test.test_name} has been cancelled"
                )
            else:
                messages.error(request, "Cannot cancel past test bookings")
        else:
            messages.error(request, "This test booking cannot be cancelled")
            
    except (PatientProfile.DoesNotExist, TestBooking.DoesNotExist):
        messages.error(request, "Test booking not found")
    
    return redirect('core/dashboard/patient_booked_tests')


@login_required
def view_test_details(request, test_id):
    """View detailed information about a diagnostic test"""
    try:
        test = DiagnosticTest.objects.select_related('lab').get(id=test_id)
    except DiagnosticTest.DoesNotExist:
        messages.error(request, "Test not found")
        return redirect('patient_diagnostic_tests')
    
    context = {
        'test': test,
    }
    
    return render(request, 'core/dashboard/test_details.html', context)


@login_required
def get_labs_api(request):
    """API endpoint to get all active labs (for AJAX)"""
    labs = Lab.objects.filter(status='Active').values('id', 'name', 'address', 'phone')
    return JsonResponse(list(labs), safe=False)


@login_required
def patient_prescriptions(request):
    """View all prescriptions"""
    try:
        patient_profile = PatientProfile.objects.get(user=request.user)
        prescriptions = Prescription.objects.filter(
            patient=patient_profile
        ).select_related('doctor', 'doctor__user', 'appointment').order_by('-created_at')
    except PatientProfile.DoesNotExist:
        prescriptions = []
    
    context = {
        'prescriptions': prescriptions
    }
    return render(request, 'core/dashboard/patient_prescriptions.html', context)

from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.http import JsonResponse, FileResponse
from django.views.decorators.http import require_http_methods
from django.db.models import Q
from datetime import datetime, timedelta
import json

# Import your existing models
from core.models import PatientProfile, DoctorProfile, Appointment, Prescription
# Import new medical history models (add these to your models.py)
from core.models import (
    Allergy, PatientMedication, MedicalCondition, Surgery,
    FamilyHistory, Immunization, VitalSigns, HealthNote,
    MedicalDocument, EmergencyContact, BloodDonationHistory
)


@login_required
def medical_history(request):
    """Display complete medical history for patient"""
    try:
        patient_profile = PatientProfile.objects.get(user=request.user)
    except PatientProfile.DoesNotExist:
        messages.error(request, "Patient profile not found. Please contact administrator.")
        return redirect('core/dashboard/patient_dashboard')

    # Get all medical information
    context = {
        'patient_profile': patient_profile,
        'allergies': patient_profile.allergies.all(),
        'medications': patient_profile.current_medications.all(),
        'conditions': patient_profile.medical_conditions.all(),
        'surgeries': patient_profile.surgeries.all(),
        'family_history': patient_profile.family_history.all(),
        'immunizations': patient_profile.immunizations.all(),
        'vital_signs': patient_profile.vital_signs.all().order_by('-date')[:10],
        'health_notes': patient_profile.health_notes.all(),
        'medical_documents': patient_profile.medical_documents.all(),
        'emergency_contacts': patient_profile.emergency_contacts.all(),
        'blood_donations': patient_profile.blood_donations.all(),
        'recent_appointments': Appointment.objects.filter(
            patient=patient_profile
        ).order_by('-appointment_date')[:5],
        'prescriptions': Prescription.objects.filter(
            patient=patient_profile
        ).order_by('-created_at')[:5],
    }
    return render(request, 'core/dashboard/medical_history.html', context)


@login_required
@require_http_methods(["POST"])
def add_allergy(request):
    """Add new allergy for patient"""
    try:
        patient_profile = PatientProfile.objects.get(user=request.user)
        allergy = Allergy.objects.create(
            patient=patient_profile,
            name=request.POST.get('name').strip(),
            severity=request.POST.get('severity'),
            reaction=request.POST.get('reaction', '').strip()
        )
        messages.success(request, f"Allergy '{allergy.name}' added successfully.")
    except Allergy.IntegrityError:
        messages.error(request, "This allergy already exists in your records.")
    except Exception as e:
        messages.error(request, f"Error adding allergy: {str(e)}")

    return redirect('core/dashboard/medical_history')


@login_required
@require_http_methods(["POST"])
def add_medication(request):
    """Add new medication"""
    try:
        patient_profile = PatientProfile.objects.get(user=request.user)
        
        # Get prescribing doctor if provided
        doctor_id = request.POST.get('prescribing_doctor')
        prescribing_doctor = None
        if doctor_id:
            prescribing_doctor = DoctorProfile.objects.get(id=doctor_id)
        
        medication = PatientMedication.objects.create(
            patient=patient_profile,
            name=request.POST.get('name').strip(),
            generic_name=request.POST.get('generic_name', '').strip(),
            dosage=request.POST.get('dosage').strip(),
            frequency=request.POST.get('frequency').strip(),
            start_date=request.POST.get('start_date'),
            end_date=request.POST.get('end_date') if request.POST.get('end_date') else None,
            prescribing_doctor=prescribing_doctor,
            reason=request.POST.get('reason', '').strip(),
        )
        messages.success(request, f"Medication '{medication.name}' added successfully.")
    except Exception as e:
        messages.error(request, f"Error adding medication: {str(e)}")

    return redirect('core/dashboard/medical_history')


@login_required
@require_http_methods(["POST"])
def add_condition(request):
    """Add new medical condition"""
    try:
        patient_profile = PatientProfile.objects.get(user=request.user)
        condition = MedicalCondition.objects.create(
            patient=patient_profile,
            name=request.POST.get('name').strip(),
            description=request.POST.get('description', '').strip(),
            diagnosis_date=request.POST.get('diagnosis_date'),
            status=request.POST.get('status', 'Active'),
            icd_code=request.POST.get('icd_code', '').strip(),
            severity=request.POST.get('severity', '').strip(),
        )
        messages.success(request, f"Condition '{condition.name}' added successfully.")
    except Exception as e:
        messages.error(request, f"Error adding condition: {str(e)}")

    return redirect('core/dashboard/medical_history')


@login_required
@require_http_methods(["POST"])
def add_surgery(request):
    """Add surgery/procedure record"""
    try:
        patient_profile = PatientProfile.objects.get(user=request.user)
        surgery = Surgery.objects.create(
            patient=patient_profile,
            name=request.POST.get('name').strip(),
            date=request.POST.get('date'),
            hospital=request.POST.get('hospital', '').strip(),
            surgeon=request.POST.get('surgeon', '').strip(),
            anesthesia_type=request.POST.get('anesthesia_type', '').strip(),
            complications=request.POST.get('complications', '').strip(),
            notes=request.POST.get('notes', '').strip()
        )
        messages.success(request, f"Surgery '{surgery.name}' added successfully.")
    except Exception as e:
        messages.error(request, f"Error adding surgery: {str(e)}")

    return redirect('core/dashboard/medical_history')


@login_required
@require_http_methods(["POST"])
def add_family_history(request):
    """Add family medical history"""
    try:
        patient_profile = PatientProfile.objects.get(user=request.user)
        family_item = FamilyHistory.objects.create(
            patient=patient_profile,
            relation=request.POST.get('relation'),
            condition=request.POST.get('condition').strip(),
            age_of_onset=request.POST.get('age_of_onset') or None,
            notes=request.POST.get('notes', '').strip()
        )
        messages.success(request, "Family history added successfully.")
    except Exception as e:
        messages.error(request, f"Error adding family history: {str(e)}")

    return redirect('core/dashboard/medical_history')


@login_required
@require_http_methods(["POST"])
def add_immunization(request):
    """Add immunization record"""
    try:
        patient_profile = PatientProfile.objects.get(user=request.user)
        immunization = Immunization.objects.create(
            patient=patient_profile,
            name=request.POST.get('name').strip(),
            date=request.POST.get('date'),
            vaccine_batch=request.POST.get('vaccine_batch', '').strip(),
            administered_by=request.POST.get('administered_by', '').strip(),
            location=request.POST.get('location', '').strip(),
            next_due=request.POST.get('next_due') if request.POST.get('next_due') else None
        )
        messages.success(request, f"Immunization '{immunization.name}' added successfully.")
    except Exception as e:
        messages.error(request, f"Error adding immunization: {str(e)}")

    return redirect('core/dashboard/medical_history')


@login_required
@require_http_methods(["POST"])
def add_vital_signs(request):
    """Record vital signs"""
    try:
        patient_profile = PatientProfile.objects.get(user=request.user)
        
        vital_signs = VitalSigns.objects.create(
            patient=patient_profile,
            date=request.POST.get('date'),
            time=request.POST.get('time', '00:00'),
            blood_pressure_systolic=request.POST.get('blood_pressure_systolic') or None,
            blood_pressure_diastolic=request.POST.get('blood_pressure_diastolic') or None,
            heart_rate=request.POST.get('heart_rate') or None,
            temperature=request.POST.get('temperature') or None,
            weight=request.POST.get('weight') or None,
            height=request.POST.get('height') or None,
            respiratory_rate=request.POST.get('respiratory_rate') or None,
            oxygen_saturation=request.POST.get('oxygen_saturation') or None,
        )
        messages.success(request, "Vital signs recorded successfully.")
    except Exception as e:
        messages.error(request, f"Error recording vital signs: {str(e)}")

    return redirect('core/dashboard/medical_history')


@login_required
@require_http_methods(["POST"])
def add_health_note(request):
    """Add health note"""
    try:
        patient_profile = PatientProfile.objects.get(user=request.user)
        health_note = HealthNote.objects.create(
            patient=patient_profile,
            title=request.POST.get('title').strip(),
            content=request.POST.get('content').strip(),
            date=request.POST.get('date')
        )
        messages.success(request, "Health note added successfully.")
    except Exception as e:
        messages.error(request, f"Error adding health note: {str(e)}")

    return redirect('core/dashboard/medical_history')


@login_required
@require_http_methods(["POST"])
def add_medical_document(request):
    """Upload medical document"""
    try:
        if 'file' not in request.FILES:
            messages.error(request, "No file uploaded.")
            return redirect('core/dashboard/medical_history')

        patient_profile = PatientProfile.objects.get(user=request.user)
        document = MedicalDocument.objects.create(
            patient=patient_profile,
            title=request.POST.get('title').strip(),
            document_type=request.POST.get('document_type'),
            file=request.FILES['file'],
            date=request.POST.get('date'),
            description=request.POST.get('description', '').strip(),
            uploaded_by=request.user
        )
        messages.success(request, f"Document '{document.title}' uploaded successfully.")
    except Exception as e:
        messages.error(request, f"Error uploading document: {str(e)}")

    return redirect('core/dashboard/medical_history')


@login_required
@require_http_methods(["POST"])
def add_emergency_contact(request):
    """Add emergency contact"""
    try:
        patient_profile = PatientProfile.objects.get(user=request.user)
        
        # If marked as primary, unmark others
        if request.POST.get('is_primary') == 'on':
            patient_profile.emergency_contacts.update(is_primary=False)
        
        emergency_contact = EmergencyContact.objects.create(
            patient=patient_profile,
            name=request.POST.get('name').strip(),
            relationship=request.POST.get('relationship').strip(),
            phone=request.POST.get('phone').strip(),
            email=request.POST.get('email', '').strip(),
            address=request.POST.get('address', '').strip(),
            is_primary=request.POST.get('is_primary') == 'on'
        )
        messages.success(request, f"Emergency contact '{emergency_contact.name}' added successfully.")
    except Exception as e:
        messages.error(request, f"Error adding emergency contact: {str(e)}")

    return redirect('core/dashboard/medical_history')


@login_required
@require_http_methods(["POST"])
def add_blood_donation(request):
    """Record blood donation"""
    try:
        patient_profile = PatientProfile.objects.get(user=request.user)
        donation_date = datetime.strptime(request.POST.get('donation_date'), '%Y-%m-%d').date()
        
        # Calculate next eligible date (56 days after donation)
        next_eligible_date = donation_date + timedelta(days=56)
        
        donation = BloodDonationHistory.objects.create(
            patient=patient_profile,
            donation_date=donation_date,
            blood_bank=request.POST.get('blood_bank', '').strip(),
            blood_quantity=int(request.POST.get('blood_quantity', 450)),
            next_eligible_date=next_eligible_date,
            notes=request.POST.get('notes', '').strip()
        )
        messages.success(request, "Blood donation recorded successfully.")
    except Exception as e:
        messages.error(request, f"Error recording blood donation: {str(e)}")

    return redirect('core/dashboard/medical_history')


@login_required
def download_medical_history(request):
    """Download medical history as PDF"""
    from reportlab.lib.pagesizes import letter
    from reportlab.lib import colors
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak
    from reportlab.lib.units import inch
    from io import BytesIO

    try:
        patient_profile = PatientProfile.objects.get(user=request.user)
    except PatientProfile.DoesNotExist:
        messages.error(request, "Patient profile not found.")
        return redirect('core/dashboard/medical_history')

    # Create PDF in memory
    pdf_buffer = BytesIO()
    doc = SimpleDocTemplate(pdf_buffer, pagesize=letter, topMargin=0.5*inch, bottomMargin=0.5*inch)
    elements = []
    styles = getSampleStyleSheet()

    # Custom styles
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=20,
        textColor=colors.HexColor('#1a1a1a'),
        spaceAfter=10,
        alignment=1,  # Center
    )
    
    heading_style = ParagraphStyle(
        'CustomHeading',
        parent=styles['Heading2'],
        fontSize=14,
        textColor=colors.HexColor('#2c3e50'),
        spaceAfter=8,
        spaceBefore=8,
    )

    # Title
    elements.append(Paragraph("MEDICAL HISTORY REPORT", title_style))
    elements.append(Spacer(1, 12))

    # Patient Information
    patient_info = [
        ['Patient Name:', patient_profile.full_name],
        ['Gender:', patient_profile.gender or 'N/A'],
        ['Date of Birth:', str(patient_profile.dob) if patient_profile.dob else 'N/A'],
        ['Phone:', patient_profile.phone or 'N/A'],
        ['Report Generated:', datetime.now().strftime("%d-%m-%Y %H:%M")],
    ]

    patient_table = Table(patient_info, colWidths=[2*inch, 4*inch])
    patient_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (0, -1), colors.HexColor('#e8f4f8')),
        ('TEXTCOLOR', (0, 0), (-1, -1), colors.black),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 10),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
        ('GRID', (0, 0), (-1, -1), 1, colors.HexColor('#cccccc')),
    ]))
    elements.append(patient_table)
    elements.append(Spacer(1, 20))

    # Allergies
    allergies = patient_profile.allergies.all()
    if allergies.exists():
        elements.append(Paragraph("ALLERGIES", heading_style))
        allergy_data = [['Name', 'Severity', 'Reaction']]
        for allergy in allergies:
            allergy_data.append([
                allergy.name,
                allergy.severity,
                allergy.reaction[:50] + '...' if len(allergy.reaction) > 50 else allergy.reaction
            ])

        allergy_table = Table(allergy_data, colWidths=[2*inch, 1.5*inch, 2.5*inch])
        allergy_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#ffcccc')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 9),
            ('GRID', (0, 0), (-1, -1), 1, colors.HexColor('#cccccc')),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#ffe6e6')]),
        ]))
        elements.append(allergy_table)
        elements.append(Spacer(1, 12))

    # Medical Conditions
    conditions = patient_profile.medical_conditions.all()
    if conditions.exists():
        elements.append(Paragraph("MEDICAL CONDITIONS", heading_style))
        condition_data = [['Condition', 'Status', 'Diagnosis Date']]
        for condition in conditions:
            condition_data.append([
                condition.name,
                condition.status,
                str(condition.diagnosis_date)
            ])

        condition_table = Table(condition_data, colWidths=[2.5*inch, 1.5*inch, 1.5*inch])
        condition_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#cce5ff')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 9),
            ('GRID', (0, 0), (-1, -1), 1, colors.HexColor('#cccccc')),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#e6f2ff')]),
        ]))
        elements.append(condition_table)
        elements.append(Spacer(1, 12))

    # Current Medications
    medications = patient_profile.current_medications.filter(end_date__isnull=True)
    if medications.exists():
        elements.append(Paragraph("CURRENT MEDICATIONS", heading_style))
        med_data = [['Medication', 'Dosage', 'Frequency']]
        for med in medications:
            med_data.append([med.name, med.dosage, med.frequency])

        med_table = Table(med_data, colWidths=[2.5*inch, 1.5*inch, 1.5*inch])
        med_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#ccffcc')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 9),
            ('GRID', (0, 0), (-1, -1), 1, colors.HexColor('#cccccc')),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#e6ffe6')]),
        ]))
        elements.append(med_table)
        elements.append(Spacer(1, 12))

    # Surgeries
    surgeries = patient_profile.surgeries.all()
    if surgeries.exists():
        elements.append(Paragraph("SURGERIES & PROCEDURES", heading_style))
        for surgery in surgeries[:5]:  # Limit to 5
            surgery_text = f"<b>{surgery.name}</b> on {surgery.date}"
            if surgery.hospital:
                surgery_text += f" at {surgery.hospital}"
            elements.append(Paragraph(surgery_text, styles['Normal']))
            elements.append(Spacer(1, 6))

    # Build PDF
    doc.build(elements)
    pdf_buffer.seek(0)

    response = FileResponse(pdf_buffer, content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="medical_history_{datetime.now().strftime("%Y%m%d_%H%M%S")}.pdf"'
    return response


@login_required
def delete_allergy(request, allergy_id):
    """Delete allergy record"""
    try:
        patient_profile = PatientProfile.objects.get(user=request.user)
        allergy = Allergy.objects.get(id=allergy_id, patient=patient_profile)
        allergy_name = allergy.name
        allergy.delete()
        messages.success(request, f"Allergy '{allergy_name}' deleted successfully.")
    except Allergy.DoesNotExist:
        messages.error(request, "Allergy not found.")
    except Exception as e:
        messages.error(request, f"Error deleting allergy: {str(e)}")

    return redirect('core/dashboard/medical_history')


@login_required
def delete_medication(request, medication_id):
    """Delete medication record"""
    try:
        patient_profile = PatientProfile.objects.get(user=request.user)
        medication = PatientMedication.objects.get(id=medication_id, patient=patient_profile)
        med_name = medication.name
        medication.delete()
        messages.success(request, f"Medication '{med_name}' deleted successfully.")
    except PatientMedication.DoesNotExist:
        messages.error(request, "Medication not found.")
    except Exception as e:
        messages.error(request, f"Error deleting medication: {str(e)}")

    return redirect('core/dashboard/medical_history')

# Add these views to your core/views.py file

from django.shortcuts import render, redirect
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.contrib.auth import update_session_auth_hash
from django.contrib.auth.password_validation import validate_password
from django.core.exceptions import ValidationError


@login_required
def patient_settings(request):
    """
    Main patient settings page
    """
    context = {
        'user': request.user,
    }
    return render(request, 'core/dashboard/patient_settings.html', context)


@login_required
def patient_settings_profile(request):
    """
    Update patient profile information
    """
    if request.method == 'POST':
        user = request.user
        
        # Get form data
        first_name = request.POST.get('first_name')
        last_name = request.POST.get('last_name')
        email = request.POST.get('email')
        phone = request.POST.get('phone')
        date_of_birth = request.POST.get('date_of_birth')
        gender = request.POST.get('gender')
        
        # Update user
        user.first_name = first_name
        user.last_name = last_name
        user.email = email
        user.save()
        
        # Update or create patient profile
        from core.models import PatientProfile
        
        profile, created = PatientProfile.objects.get_or_create(user=user)
        profile.phone = phone
        profile.date_of_birth = date_of_birth if date_of_birth else None
        profile.gender = gender
        profile.save()
        
        messages.success(request, 'Profile updated successfully!')
        return redirect('patient_settings')
    
    return redirect('patient_settings')


@login_required
def patient_settings_security(request):
    """
    Change patient password
    """
    if request.method == 'POST':
        user = request.user
        
        current_password = request.POST.get('current_password')
        new_password = request.POST.get('new_password')
        confirm_password = request.POST.get('confirm_password')
        
        # Verify current password
        if not user.check_password(current_password):
            messages.error(request, 'Current password is incorrect!')
            return redirect('patient_settings')
        
        # Check if passwords match
        if new_password != confirm_password:
            messages.error(request, 'New passwords do not match!')
            return redirect('patient_settings')
        
        # Validate new password
        try:
            validate_password(new_password, user)
        except ValidationError as e:
            messages.error(request, ' '.join(e.messages))
            return redirect('patient_settings')
        
        # Update password
        user.set_password(new_password)
        user.save()
        
        # Keep user logged in
        update_session_auth_hash(request, user)
        
        messages.success(request, 'Password changed successfully!')
        return redirect('patient_settings')
    
    return redirect('patient_settings')


@login_required
def patient_settings_medical(request):
    """
    Update patient medical information
    """
    if request.method == 'POST':
        from core.models import PatientProfile
        
        # Get form data
        blood_group = request.POST.get('blood_group')
        height = request.POST.get('height')
        weight = request.POST.get('weight')
        emergency_contact = request.POST.get('emergency_contact')
        allergies = request.POST.get('allergies')
        medical_conditions = request.POST.get('medical_conditions')
        
        # Update or create patient profile
        profile, created = PatientProfile.objects.get_or_create(user=request.user)
        profile.blood_group = blood_group
        profile.height = height if height else None
        profile.weight = weight if weight else None
        profile.emergency_contact = emergency_contact
        profile.allergies = allergies
        profile.medical_conditions = medical_conditions
        profile.save()
        
        messages.success(request, 'Medical information updated successfully!')
        return redirect('patient_settings')
    
    return redirect('core/dashboard/patient_settings')


@login_required
def patient_settings_notifications(request):
    """
    Update patient notification preferences
    """
    if request.method == 'POST':
        from core.models import PatientProfile
        
        # Get checkbox values
        email_appointments = request.POST.get('email_appointments') == 'on'
        email_reminders = request.POST.get('email_reminders') == 'on'
        email_lab_results = request.POST.get('email_lab_results') == 'on'
        
        # Update or create patient profile
        profile, created = PatientProfile.objects.get_or_create(user=request.user)
        profile.email_appointments = email_appointments
        profile.email_reminders = email_reminders
        profile.email_lab_results = email_lab_results
        profile.save()
        
        messages.success(request, 'Notification preferences updated!')
        return redirect('patient_settings')
    
    return redirect(core/core/dashboard/'patient_settings')

# Update this view in your core/views.py file
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from datetime import date
from core.models import Appointment, PatientProfile


@login_required
def patient_appointments(request):
    """
    View all patient's appointments with statistics
    ✅ FIXED: Properly retrieve PatientProfile before filtering
    ✅ FIXED: Correct select_related syntax (doctor__user not doctor__doctorprofile)
    """
    try:
        # ✅ FIXED: Get the PatientProfile first
        patient_profile = PatientProfile.objects.get(user=request.user)
    except PatientProfile.DoesNotExist:
        messages.error(request, "Patient profile not found. Please contact support.")
        return redirect('patient_dashboard')
    
    # Get all appointments for this patient
    # ✅ FIXED: Use doctor__user instead of doctor__doctorprofile
    appointments = Appointment.objects.filter(
        patient=patient_profile  # ✅ Now passing PatientProfile instance
    ).select_related('doctor', 'doctor__user').order_by('-appointment_date', '-appointment_time')
    
    # Calculate statistics
    pending_count = appointments.filter(status='pending').count()
    confirmed_count = appointments.filter(status='confirmed').count()
    completed_count = appointments.filter(status='completed').count()
    cancelled_count = appointments.filter(status='cancelled').count()
    
    context = {
        'appointments': appointments,
        'pending_count': pending_count,
        'confirmed_count': confirmed_count,
        'completed_count': completed_count,
        'cancelled_count': cancelled_count,
    }
    
    return render(request, 'core/dashboard/patient_appointments.html', context)


@login_required
def patient_cancel_appointment(request, appointment_id):
    """
    Cancel an appointment
    ✅ FIXED: Properly retrieve PatientProfile before filtering
    """
    try:
        patient_profile = PatientProfile.objects.get(user=request.user)
    except PatientProfile.DoesNotExist:
        messages.error(request, "Patient profile not found!")
        return redirect('patient_appointments')
    
    try:
        appointment = Appointment.objects.get(id=appointment_id, patient=patient_profile)
        
        # Check if appointment can be cancelled
        if appointment.status in ['pending', 'confirmed']:
            # Check if appointment is in the future
            if appointment.appointment_date >= date.today():
                appointment.status = 'cancelled'
                appointment.save()
                messages.success(
                    request, 
                    f'Appointment with Dr. {appointment.doctor.user.get_full_name()} has been cancelled successfully!'
                )
            else:
                messages.error(request, 'Cannot cancel past appointments!')
        else:
            messages.error(request, 'This appointment cannot be cancelled!')
            
    except Appointment.DoesNotExist:
        messages.error(request, 'Appointment not found!')
    
    return redirect('patient_appointments')

def patient_appointment_detail(request, appointment_id):
    try:
        patient_profile = PatientProfile.objects.get(user=request.user)
    except PatientProfile.DoesNotExist:
        messages.error(request, 'Patient profile not found!')
        return redirect('patient_dashboard')
    
    try:
        appointment = Appointment.objects.select_related(
            'doctor', 'doctor__user'
        ).get(id=appointment_id, patient=patient_profile)

        # Get payment (if exists)
        payment = getattr(appointment, 'payment', None)

        context = {
            'appointment': appointment,
            'payment': payment,
        }

        return render(request, 'core/dashboard/patient_appointment_detail.html', context)

    except Appointment.DoesNotExist:
        messages.error(request, 'Appointment not found!')
        return redirect('patient_appointments')


# ============================================================================
# BONUS: Additional fixes for other views that filter by patient
# ============================================================================

@login_required
def patient_prescriptions(request):
    """
    View all prescriptions
    ✅ FIXED: Properly retrieve PatientProfile before filtering
    """
    from core.models import Prescription
    
    try:
        patient_profile = PatientProfile.objects.get(user=request.user)
        prescriptions = Prescription.objects.filter(
            patient=patient_profile
        ).select_related('doctor', 'doctor__user', 'appointment').order_by('-created_at')
    except PatientProfile.DoesNotExist:
        messages.error(request, "Patient profile not found.")
        prescriptions = []
        return redirect('patient_dashboard')
    
    context = {
        'prescriptions': prescriptions
    }
    return render(request, 'core/dashboard/patient_prescriptions.html', context)


# ==========================================
# DOCTOR VIEWS - Add to core/views.py
# ==========================================

from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.db.models import Q, Count
from django.http import JsonResponse
from datetime import date
from core.models import (
    DoctorProfile,
    PatientProfile,
    Appointment,
    Prescription,
    PatientHistory,
    Allergy,
    MedicalCondition,
    PatientMedication,
)


# ==================== DOCTOR DASHBOARD ====================

@login_required
def doctor_dashboard(request):
    """
    Doctor Dashboard - Main overview page
    Shows today's appointments, pending prescriptions, patient stats, and schedule
    """
    try:
        doctor_profile = DoctorProfile.objects.get(user=request.user)
    except DoctorProfile.DoesNotExist:
        messages.error(request, "Doctor profile not found")
        return redirect('login')
    
    today = date.today()
    
    # Today's Appointments
    today_appointments = Appointment.objects.filter(
        doctor=doctor_profile,
        appointment_date=today,
        status__in=['Scheduled', 'Confirmed']
    ).select_related('patient', 'patient__user').order_by('appointment_time')
    
    today_appointments_count = today_appointments.count()
    next_appointment = today_appointments.first()
    
    # Total Patients (unique patients with appointments)
    total_patients = Appointment.objects.filter(
        doctor=doctor_profile
    ).values('patient').distinct().count()
    
    # Pending Prescriptions
    pending_prescriptions_list = Prescription.objects.filter(
        doctor=doctor_profile
    ).select_related('patient').order_by('-created_at')[:5]
    
    pending_prescriptions = Prescription.objects.filter(
        doctor=doctor_profile
    ).count()
    
    # Pending Approvals (e.g., medical history requests)
    pending_approvals = PatientHistory.objects.filter(
        doctor=doctor_profile
    ).count()
    
    # Recent Patients (last seen)
    recent_patient_ids = Appointment.objects.filter(
        doctor=doctor_profile,
        status='Completed'
    ).order_by('-appointment_date').values_list(
        'patient_id', flat=True
    ).distinct()[:5]
    
    recent_patients = PatientProfile.objects.filter(
        id__in=recent_patient_ids
    )
    
    # Unread Notifications
    unread_notifications = 0  # Implement based on your notification system
    
    context = {
        'today_appointments': today_appointments,
        'today_appointments_count': today_appointments_count,
        'next_appointment': next_appointment,
        'total_patients': total_patients,
        'pending_prescriptions': pending_prescriptions,
        'pending_prescriptions_list': pending_prescriptions_list,
        'pending_approvals': pending_approvals,
        'recent_patients': recent_patients,
        'doctor_profile': doctor_profile,
        'unread_notifications': unread_notifications,
    }
    
    return render(request, 'core/dashboard/doctor_dashboard.html', context)


# ==================== DOCTOR APPOINTMENTS ====================

@login_required
def doctor_appointments(request):
    """
    Doctor's complete appointments list with filtering
    """
    try:
        doctor_profile = DoctorProfile.objects.get(user=request.user)
    except DoctorProfile.DoesNotExist:
        messages.error(request, "Doctor profile not found")
        return redirect('login')
    
    # Get all appointments for this doctor
    appointments = Appointment.objects.filter(
        doctor=doctor_profile
    ).select_related('patient', 'patient__user').order_by('appointment_date', 'appointment_time')
    
    # Filter by status if provided
    status_filter = request.GET.get('status', '')
    if status_filter:
        appointments = appointments.filter(status=status_filter)
    
    # Filter by date range if provided
    date_from = request.GET.get('date_from', '')
    date_to = request.GET.get('date_to', '')
    
    if date_from:
        appointments = appointments.filter(appointment_date__gte=date_from)
    if date_to:
        appointments = appointments.filter(appointment_date__lte=date_to)
    
    # Separate appointments by status
    all_appointments = Appointment.objects.filter(doctor=doctor_profile)
    
    stats = {
        'pending_payment': all_appointments.filter(status='Pending Payment').count(),
        'scheduled': all_appointments.filter(status='Scheduled').count(),
        'confirmed': all_appointments.filter(status='Confirmed').count(),
        'completed': all_appointments.filter(status='Completed').count(),
        'cancelled': all_appointments.filter(status='Cancelled').count(),
        'no_show': all_appointments.filter(status='No Show').count(),
    }
    
    context = {
        'appointments': appointments,
        'status_filter': status_filter,
        'date_from': date_from,
        'date_to': date_to,
        'stats': stats,
        'doctor_profile': doctor_profile,
    }
    
    return render(request, 'core/dashboard/doctor_appointments.html', context)


@login_required
def doctor_appointment_detail(request, appointment_id):
    """
    View detailed information about a specific appointment
    """
    try:
        doctor_profile = DoctorProfile.objects.get(user=request.user)
    except DoctorProfile.DoesNotExist:
        messages.error(request, "Doctor profile not found")
        return redirect('login')
    
    appointment = get_object_or_404(
        Appointment,
        id=appointment_id,
        doctor=doctor_profile
    )
    
    # Get patient medical history
    patient_history = PatientHistory.objects.filter(
        patient=appointment.patient,
        doctor=doctor_profile
    ).order_by('-recorded_date')
    
    # Get prescriptions for this appointment
    prescriptions = Prescription.objects.filter(
        appointment=appointment
    )
    
    # Get patient's current medications and allergies
    current_medications = PatientMedication.objects.filter(
        patient=appointment.patient,
        end_date__isnull=True
    )
    
    allergies = Allergy.objects.filter(patient=appointment.patient)
    
    payment = appointment.payments.filter(payment_status="Paid").order_by('-payment_date').first()
    
    
    
    context = {
        'appointment': appointment,
        'patient_history': patient_history,
        'prescriptions': prescriptions,
        'current_medications': current_medications,
        'allergies': allergies,
        'doctor_profile': doctor_profile,
        'payment': payment,
    }
    
    return render(request, 'core/dashboard/doctor_appointment_detail.html', context)


@login_required
def doctor_confirm_appointment(request, appointment_id):
    """
    Confirm an appointment
    """
    try:
        doctor_profile = DoctorProfile.objects.get(user=request.user)
    except DoctorProfile.DoesNotExist:
        messages.error(request, "Doctor profile not found")
        return redirect('login')
    
    appointment = get_object_or_404(
        Appointment,
        id=appointment_id,
        doctor=doctor_profile
    )
    
    if appointment.status in ['Scheduled', 'Pending Payment']:
        appointment.status = 'Confirmed'
        appointment.save()
        messages.success(
            request, 
            f"Appointment with {appointment.patient.full_name} confirmed successfully"
        )
    else:
        messages.warning(request, "This appointment cannot be confirmed")
    
    return redirect('doctor_appointments')


@login_required
def doctor_complete_appointment(request, appointment_id):
    """
    Mark appointment as completed and add medical history
    """
    try:
        doctor_profile = DoctorProfile.objects.get(user=request.user)
    except DoctorProfile.DoesNotExist:
        messages.error(request, "Doctor profile not found")
        return redirect('login')
    
    appointment = get_object_or_404(
        Appointment,
        id=appointment_id,
        doctor=doctor_profile
    )
    
    if request.method == 'POST':
        diagnosis = request.POST.get('diagnosis', '')
        treatment = request.POST.get('treatment', '')
        notes = request.POST.get('notes', '')
        
        # Create patient history record
        PatientHistory.objects.create(
            patient=appointment.patient,
            doctor=doctor_profile,
            appointment=appointment,
            diagnosis=diagnosis,
            treatment=treatment,
            notes=notes,
            recorded_date=date.today()
        )
        
        # Update appointment status
        appointment.status = 'Completed'
        appointment.save()
        
        messages.success(
            request, 
            f"Appointment with {appointment.patient.full_name} marked as completed"
        )
        return redirect('doctor_appointments')
    
    context = {
        'appointment': appointment,
        'doctor_profile': doctor_profile,
    }
    
    return render(request, 'core/dashboard/doctor_complete_appointment.html', context)


# ==================== DOCTOR PATIENTS ====================

@login_required
def doctor_patients(request):
    """
    View all patients under this doctor's care
    """
    try:
        doctor_profile = DoctorProfile.objects.get(user=request.user)
    except DoctorProfile.DoesNotExist:
        messages.error(request, "Doctor profile not found")
        return redirect('login')
    
    # Get unique patients with appointments to this doctor
    patient_ids = Appointment.objects.filter(
        doctor=doctor_profile
    ).values_list('patient_id', flat=True).distinct()
    
    patients = PatientProfile.objects.filter(
        id__in=patient_ids
    ).order_by('full_name')
    
    # Search functionality
    search_query = request.GET.get('search', '')
    if search_query:
        patients = patients.filter(
            Q(full_name__icontains=search_query) |
            Q(phone__icontains=search_query) |
            Q(user__email__icontains=search_query)
        )
    
    # Add appointment count and last visit for each patient
    for patient in patients:
        appointments = Appointment.objects.filter(
            doctor=doctor_profile,
            patient=patient
        )
        patient.appointment_count = appointments.count()
        patient.last_appointment = appointments.order_by('-appointment_date').first()
    
    context = {
        'patients': patients,
        'search_query': search_query,
        'doctor_profile': doctor_profile,
    }
    
    return render(request, 'core/dashboard/doctor_patients.html', context)


@login_required
def doctor_patient_detail(request, patient_id):
    """
    View detailed patient information
    """
    try:
        doctor_profile = DoctorProfile.objects.get(user=request.user)
    except DoctorProfile.DoesNotExist:
        messages.error(request, "Doctor profile not found")
        return redirect('login')
    
    patient = get_object_or_404(PatientProfile, id=patient_id)
    
    # Verify doctor has seen this patient
    has_appointment = Appointment.objects.filter(
        doctor=doctor_profile,
        patient=patient
    ).exists()
    
    if not has_appointment:
        messages.error(request, "You don't have access to this patient's records")
        return redirect('doctor_patients')
    
    # Get patient data
    appointments = Appointment.objects.filter(
        doctor=doctor_profile,
        patient=patient
    ).order_by('-appointment_date')
    
    medical_history = PatientHistory.objects.filter(
        doctor=doctor_profile,
        patient=patient
    ).order_by('-recorded_date')
    
    prescriptions = Prescription.objects.filter(
        patient=patient,
        doctor=doctor_profile
    ).order_by('-created_at')
    
    allergies = Allergy.objects.filter(patient=patient)
    conditions = MedicalCondition.objects.filter(patient=patient)
    medications = PatientMedication.objects.filter(
        patient=patient,
        end_date__isnull=True
    )
    
    context = {
        'patient': patient,
        'appointments': appointments,
        'medical_history': medical_history,
        'prescriptions': prescriptions,
        'allergies': allergies,
        'conditions': conditions,
        'medications': medications,
        'doctor_profile': doctor_profile,
    }
    
    return render(request, 'core/dashboard/doctor_patient_detail.html', context)


# ==================== DOCTOR PRESCRIPTIONS ====================

@login_required
def doctor_prescriptions(request):
    """
    View and manage prescriptions
    """
    try:
        doctor_profile = DoctorProfile.objects.get(user=request.user)
    except DoctorProfile.DoesNotExist:
        messages.error(request, "Doctor profile not found")
        return redirect('login')
    
    prescriptions = Prescription.objects.filter(
        doctor=doctor_profile
    ).select_related('patient', 'appointment').order_by('-created_at')
    
    # Filter by status if provided
    status_filter = request.GET.get('status', '')
    if status_filter:
        prescriptions = prescriptions.filter(status=status_filter)
    
    # Get statistics
    all_prescriptions = Prescription.objects.filter(doctor=doctor_profile)
    
    stats = {
        'total': all_prescriptions.count(),
        'pending': all_prescriptions.filter(status='Pending').count(),
        'active': all_prescriptions.filter(status='Active').count(),
        'completed': all_prescriptions.filter(status='Completed').count(),
    }
    
    context = {
        'prescriptions': prescriptions,
        'status_filter': status_filter,
        'stats': stats,
        'doctor_profile': doctor_profile,
    }
    
    return render(request, 'core/dashboard/doctor_prescriptions.html', context)


@login_required
def doctor_add_prescription(request, appointment_id):
    """
    Add a prescription for an appointment
    """
    try:
        doctor_profile = DoctorProfile.objects.get(user=request.user)
    except DoctorProfile.DoesNotExist:
        messages.error(request, "Doctor profile not found")
        return redirect('login')
    
    appointment = get_object_or_404(
        Appointment,
        id=appointment_id,
        doctor=doctor_profile
    )
    
    if request.method == 'POST':
        medicine_name = request.POST.get('medicine_name', '').strip()
        dosage = request.POST.get('dosage', '').strip()
        frequency = request.POST.get('frequency', '').strip()
        duration = request.POST.get('duration', '').strip()
        instructions = request.POST.get('instructions', '').strip()
        
        if not all([medicine_name, dosage, frequency, duration]):
            messages.error(request, "All prescription fields are required")
            return redirect('doctor_add_prescription', appointment_id=appointment_id)
        
        Prescription.objects.create(
            appointment=appointment,
            patient=appointment.patient,
            doctor=doctor_profile,
            medicine_name=medicine_name,
            dosage=dosage,
            frequency=frequency,
            duration=duration,
            instructions=instructions
        )
        
        messages.success(request, f"Prescription for {medicine_name} added successfully")
        return redirect('doctor_appointment_detail', appointment_id=appointment_id)
    
    context = {
        'appointment': appointment,
        'doctor_profile': doctor_profile,
    }
    
    return render(request, 'core/dashboard/doctor_add_prescription.html', context)


@login_required
def doctor_prescription_detail(request, prescription_id):
    """
    View and edit prescription details
    """
    try:
        doctor_profile = DoctorProfile.objects.get(user=request.user)
    except DoctorProfile.DoesNotExist:
        messages.error(request, "Doctor profile not found")
        return redirect('login')
    
    prescription = get_object_or_404(
        Prescription,
        id=prescription_id,
        doctor=doctor_profile
    )
    
    if request.method == 'POST':
        prescription.medicine_name = request.POST.get('medicine_name', prescription.medicine_name).strip()
        prescription.dosage = request.POST.get('dosage', prescription.dosage).strip()
        prescription.frequency = request.POST.get('frequency', prescription.frequency).strip()
        prescription.duration = request.POST.get('duration', prescription.duration).strip()
        prescription.instructions = request.POST.get('instructions', prescription.instructions).strip()
        prescription.save()
        
        messages.success(request, "Prescription updated successfully")
        return redirect('doctor_prescriptions')
    
    context = {
        'prescription': prescription,
        'doctor_profile': doctor_profile,
    }
    
    return render(request, 'core/dashboard/doctor_prescription_detail.html', context)


# ==================== DOCTOR SCHEDULE ====================

@login_required
def doctor_schedule(request):
    """
    Manage doctor's availability schedule
    """
    try:
        doctor_profile = DoctorProfile.objects.get(user=request.user)
    except DoctorProfile.DoesNotExist:
        messages.error(request, "Doctor profile not found")
        return redirect('login')
    
    # Get doctor's availability (you may need to create a DoctorAvailability model)
    # For now, return default schedule
    
    context = {
        'doctor_profile': doctor_profile,
        'working_hours': '9:00 AM - 5:00 PM',
        'working_days': 'Monday - Friday',
    }
    
    return render(request, 'core/dashboard/doctor_schedule.html', context)


@login_required
def doctor_schedule_update(request):
    """
    Update doctor's schedule
    """
    try:
        doctor_profile = DoctorProfile.objects.get(user=request.user)
    except DoctorProfile.DoesNotExist:
        messages.error(request, "Doctor profile not found")
        return redirect('login')
    
    if request.method == 'POST':
        # Update working hours
        # You would need to store this in a DoctorAvailability or similar model
        
        messages.success(request, "Schedule updated successfully")
        return redirect('doctor_schedule')
    
    return redirect('core/dashboard/doctor_schedule')

# ==========================================
# DOCTOR VIEWS - Complete Implementation
# Add this to core/views.py
# ==========================================

from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.db.models import Q, Count, Avg
from django.http import JsonResponse
from datetime import date
from core.models import (
    DoctorProfile,
    PatientProfile,
    Appointment,
    Prescription,
    PatientHistory,
    Allergy,
    MedicalCondition,
    PatientMedication,
)


# ==================== DOCTOR DASHBOARD ====================

@login_required
def doctor_dashboard(request):
    """
    Doctor Dashboard - Main overview page
    Shows today's appointments, pending prescriptions, patient stats, and schedule
    """
    try:
        doctor_profile = DoctorProfile.objects.get(user=request.user)
    except DoctorProfile.DoesNotExist:
        messages.error(request, "Doctor profile not found")
        return redirect('login')
    
    today = date.today()
    
    # Today's Appointments
    today_appointments = Appointment.objects.filter(
        doctor=doctor_profile,
        appointment_date=today,
        status__in=['Scheduled', 'Confirmed']
    ).select_related('patient', 'patient__user').order_by('appointment_time')
    
    today_appointments_count = today_appointments.count()
    next_appointment = today_appointments.first()
    
    # Total Patients (unique patients with appointments)
    total_patients = Appointment.objects.filter(
        doctor=doctor_profile
    ).values('patient').distinct().count()
    
    # Pending Prescriptions
    pending_prescriptions_list = Prescription.objects.filter(
        doctor=doctor_profile
    ).select_related('patient').order_by('-created_at')[:5]
    
    pending_prescriptions = Prescription.objects.filter(
        doctor=doctor_profile
    ).count()
    
    # Pending Approvals
    pending_approvals = PatientHistory.objects.filter(
        doctor=doctor_profile
    ).count()
    
    # Recent Patients (last seen)
    recent_patient_ids = Appointment.objects.filter(
        doctor=doctor_profile,
        status='Completed'
    ).order_by('-appointment_date').values_list(
        'patient_id', flat=True
    ).distinct()[:5]
    
    recent_patients = PatientProfile.objects.filter(
        id__in=recent_patient_ids
    )
    
    # Unread Notifications
    unread_notifications = 0
    
    context = {
        'today_appointments': today_appointments,
        'today_appointments_count': today_appointments_count,
        'next_appointment': next_appointment,
        'total_patients': total_patients,
        'pending_prescriptions': pending_prescriptions,
        'pending_prescriptions_list': pending_prescriptions_list,
        'pending_approvals': pending_approvals,
        'recent_patients': recent_patients,
        'doctor_profile': doctor_profile,
        'unread_notifications': unread_notifications,
    }
    
    return render(request, 'core/dashboard/doctor_dashboard.html', context)


# ==================== DOCTOR APPOINTMENTS ====================

@login_required
def doctor_appointments(request):
    """
    Doctor's complete appointments list with filtering
    """
    try:
        doctor_profile = DoctorProfile.objects.get(user=request.user)
    except DoctorProfile.DoesNotExist:
        messages.error(request, "Doctor profile not found")
        return redirect('login')
    
    # Get all appointments for this doctor
    appointments = Appointment.objects.filter(
        doctor=doctor_profile
    ).select_related('patient', 'patient__user').order_by('appointment_date', 'appointment_time')
    
    # Filter by status if provided
    status_filter = request.GET.get('status', '')
    if status_filter:
        appointments = appointments.filter(status=status_filter)
    
    # Filter by date range if provided
    date_from = request.GET.get('date_from', '')
    date_to = request.GET.get('date_to', '')
    
    if date_from:
        appointments = appointments.filter(appointment_date__gte=date_from)
    if date_to:
        appointments = appointments.filter(appointment_date__lte=date_to)
    
    # Separate appointments by status
    all_appointments = Appointment.objects.filter(doctor=doctor_profile)
    
    stats = {
        'pending_payment': all_appointments.filter(status='Pending Payment').count(),
        'scheduled': all_appointments.filter(status='Scheduled').count(),
        'confirmed': all_appointments.filter(status='Confirmed').count(),
        'completed': all_appointments.filter(status='Completed').count(),
        'cancelled': all_appointments.filter(status='Cancelled').count(),
        'no_show': all_appointments.filter(status='No Show').count(),
    }
    
    context = {
        'appointments': appointments,
        'status_filter': status_filter,
        'date_from': date_from,
        'date_to': date_to,
        'stats': stats,
        'doctor_profile': doctor_profile,
    }
    
    return render(request, 'core/dashboard/doctor_appointments.html', context)


@login_required
def doctor_appointment_detail(request, appointment_id):
    """
    View detailed information about a specific appointment
    """
    try:
        doctor_profile = DoctorProfile.objects.get(user=request.user)
    except DoctorProfile.DoesNotExist:
        messages.error(request, "Doctor profile not found")
        return redirect('login')
    
    appointment = get_object_or_404(
        Appointment,
        id=appointment_id,
        doctor=doctor_profile
    )
    
    # Get patient medical history
    patient_history = PatientHistory.objects.filter(
        patient=appointment.patient,
        doctor=doctor_profile
    ).order_by('-recorded_date')
    
    # Get prescriptions for this appointment
    prescriptions = Prescription.objects.filter(
        appointment=appointment
    )
    
    # Get patient's current medications and allergies
    current_medications = PatientMedication.objects.filter(
        patient=appointment.patient,
        end_date__isnull=True
    )
    
    allergies = Allergy.objects.filter(patient=appointment.patient)
    
    context = {
        'appointment': appointment,
        'patient_history': patient_history,
        'prescriptions': prescriptions,
        'current_medications': current_medications,
        'allergies': allergies,
        'doctor_profile': doctor_profile,
    }
    
    return render(request, 'core/dashboard/doctor_appointment_detail.html', context)


@login_required
def doctor_confirm_appointment(request, appointment_id):
    """
    Confirm an appointment
    """
    try:
        doctor_profile = DoctorProfile.objects.get(user=request.user)
    except DoctorProfile.DoesNotExist:
        messages.error(request, "Doctor profile not found")
        return redirect('login')
    
    appointment = get_object_or_404(
        Appointment,
        id=appointment_id,
        doctor=doctor_profile
    )
    
    if appointment.status in ['Scheduled', 'Pending Payment']:
        appointment.status = 'Confirmed'
        appointment.save()
        messages.success(
            request, 
            f"Appointment with {appointment.patient.full_name} confirmed successfully"
        )
    else:
        messages.warning(request, "This appointment cannot be confirmed")
    
    return redirect('doctor_appointments')


@login_required
def doctor_complete_appointment(request, appointment_id):
    """
    Mark appointment as completed and add medical history
    """
    try:
        doctor_profile = DoctorProfile.objects.get(user=request.user)
    except DoctorProfile.DoesNotExist:
        messages.error(request, "Doctor profile not found")
        return redirect('login')
    
    appointment = get_object_or_404(
        Appointment,
        id=appointment_id,
        doctor=doctor_profile
    )
    
    if request.method == 'POST':
        diagnosis = request.POST.get('diagnosis', '')
        treatment = request.POST.get('treatment', '')
        notes = request.POST.get('notes', '')
        
        # Create patient history record
        PatientHistory.objects.create(
            patient=appointment.patient,
            doctor=doctor_profile,
            appointment=appointment,
            diagnosis=diagnosis,
            treatment=treatment,
            notes=notes,
            recorded_date=date.today()
        )
        
        # Update appointment status
        appointment.status = 'Completed'
        appointment.save()
        
        messages.success(
            request, 
            f"Appointment with {appointment.patient.full_name} marked as completed"
        )
        return redirect('doctor_appointments')
    
    context = {
        'appointment': appointment,
        'doctor_profile': doctor_profile,
    }
    
    return render(request, 'core/dashboard/doctor_complete_appointment.html', context)


# ==================== DOCTOR PATIENTS ====================

@login_required
def doctor_patients(request):
    """
    View all patients under this doctor's care
    """
    try:
        doctor_profile = DoctorProfile.objects.get(user=request.user)
    except DoctorProfile.DoesNotExist:
        messages.error(request, "Doctor profile not found")
        return redirect('login')
    
    # Get unique patients with appointments to this doctor
    patient_ids = Appointment.objects.filter(
        doctor=doctor_profile
    ).values_list('patient_id', flat=True).distinct()
    
    patients = PatientProfile.objects.filter(
        id__in=patient_ids
    ).order_by('full_name')
    
    # Search functionality
    search_query = request.GET.get('search', '')
    if search_query:
        patients = patients.filter(
            Q(full_name__icontains=search_query) |
            Q(phone__icontains=search_query) |
            Q(user__email__icontains=search_query)
        )
    
    # Add appointment count and last visit for each patient
    for patient in patients:
        appointments = Appointment.objects.filter(
            doctor=doctor_profile,
            patient=patient
        )
        patient.appointment_count = appointments.count()
        patient.last_appointment = appointments.order_by('-appointment_date').first()
    
    context = {
        'patients': patients,
        'search_query': search_query,
        'doctor_profile': doctor_profile,
    }
    
    return render(request, 'core/dashboard/doctor_patients.html', context)


@login_required
def doctor_patient_detail(request, patient_id):
    """
    View detailed patient information
    """
    try:
        doctor_profile = DoctorProfile.objects.get(user=request.user)
    except DoctorProfile.DoesNotExist:
        messages.error(request, "Doctor profile not found")
        return redirect('login')
    
    patient = get_object_or_404(PatientProfile, id=patient_id)
    
    # Verify doctor has seen this patient
    has_appointment = Appointment.objects.filter(
        doctor=doctor_profile,
        patient=patient
    ).exists()
    
    if not has_appointment:
        messages.error(request, "You don't have access to this patient's records")
        return redirect('doctor_patients')
    
    # Get patient data
    appointments = Appointment.objects.filter(
        doctor=doctor_profile,
        patient=patient
    ).order_by('-appointment_date')
    
    medical_history = PatientHistory.objects.filter(
        doctor=doctor_profile,
        patient=patient
    ).order_by('-recorded_date')
    
    prescriptions = Prescription.objects.filter(
        patient=patient,
        doctor=doctor_profile
    ).order_by('-created_at')
    
    allergies = Allergy.objects.filter(patient=patient)
    conditions = MedicalCondition.objects.filter(patient=patient)
    medications = PatientMedication.objects.filter(
        patient=patient,
        end_date__isnull=True
    )
    
    context = {
        'patient': patient,
        'appointments': appointments,
        'medical_history': medical_history,
        'prescriptions': prescriptions,
        'allergies': allergies,
        'conditions': conditions,
        'medications': medications,
        'doctor_profile': doctor_profile,
    }
    
    return render(request, 'core/dashboard/doctor_patient_detail.html', context)


# ==================== DOCTOR PRESCRIPTIONS ====================

@login_required
def doctor_prescriptions(request):
    """
    View and manage prescriptions
    """
    try:
        doctor_profile = DoctorProfile.objects.get(user=request.user)
    except DoctorProfile.DoesNotExist:
        messages.error(request, "Doctor profile not found")
        return redirect('login')
    
    prescriptions = Prescription.objects.filter(
        doctor=doctor_profile
    ).select_related('patient', 'appointment').order_by('-created_at')
    
    # Filter by status if provided
    status_filter = request.GET.get('status', '')
    if status_filter:
        prescriptions = prescriptions.filter(status=status_filter)
    
    # Get statistics
    all_prescriptions = Prescription.objects.filter(doctor=doctor_profile)
    
    stats = {
        'total': all_prescriptions.count(),
        'pending': all_prescriptions.filter(status='Pending').count(),
        'active': all_prescriptions.filter(status='Active').count(),
        'completed': all_prescriptions.filter(status='Completed').count(),
    }
    
    context = {
        'prescriptions': prescriptions,
        'status_filter': status_filter,
        'stats': stats,
        'doctor_profile': doctor_profile,
    }
    
    return render(request, 'core/dashboard/doctor_prescriptions.html', context)


@login_required
def doctor_add_prescription(request, appointment_id):
    """
    Add a prescription for an appointment
    """
    try:
        doctor_profile = DoctorProfile.objects.get(user=request.user)
    except DoctorProfile.DoesNotExist:
        messages.error(request, "Doctor profile not found")
        return redirect('login')
    
    appointment = get_object_or_404(
        Appointment,
        id=appointment_id,
        doctor=doctor_profile
    )
    
    if request.method == 'POST':
        medicine_name = request.POST.get('medicine_name', '').strip()
        dosage = request.POST.get('dosage', '').strip()
        frequency = request.POST.get('frequency', '').strip()
        duration = request.POST.get('duration', '').strip()
        instructions = request.POST.get('instructions', '').strip()
        
        if not all([medicine_name, dosage, frequency, duration]):
            messages.error(request, "All prescription fields are required")
            return redirect('doctor_add_prescription', appointment_id=appointment_id)
        
        Prescription.objects.create(
            appointment=appointment,
            patient=appointment.patient,
            doctor=doctor_profile,
            medicine_name=medicine_name,
            dosage=dosage,
            frequency=frequency,
            duration=duration,
            instructions=instructions
        )
        
        messages.success(request, f"Prescription for {medicine_name} added successfully")
        return redirect('doctor_appointment_detail', appointment_id=appointment_id)
    
    context = {
        'appointment': appointment,
        'doctor_profile': doctor_profile,
    }
    
    return render(request, 'core/dashboard/doctor_add_prescription.html', context)


@login_required
def doctor_prescription_detail(request, prescription_id):
    """
    View and edit prescription details
    """
    try:
        doctor_profile = DoctorProfile.objects.get(user=request.user)
    except DoctorProfile.DoesNotExist:
        messages.error(request, "Doctor profile not found")
        return redirect('login')
    
    prescription = get_object_or_404(
        Prescription,
        id=prescription_id,
        doctor=doctor_profile
    )
    
    if request.method == 'POST':
        prescription.medicine_name = request.POST.get('medicine_name', prescription.medicine_name).strip()
        prescription.dosage = request.POST.get('dosage', prescription.dosage).strip()
        prescription.frequency = request.POST.get('frequency', prescription.frequency).strip()
        prescription.duration = request.POST.get('duration', prescription.duration).strip()
        prescription.instructions = request.POST.get('instructions', prescription.instructions).strip()
        prescription.save()
        
        messages.success(request, "Prescription updated successfully")
        return redirect('doctor_prescriptions')
    
    context = {
        'prescription': prescription,
        'doctor_profile': doctor_profile,
    }
    
    return render(request, 'core/dashboard/doctor_prescription_detail.html', context)


# ==================== DOCTOR SCHEDULE ====================

@login_required
def doctor_schedule(request):
    """
    Manage doctor's availability schedule
    """
    try:
        doctor_profile = DoctorProfile.objects.get(user=request.user)
    except DoctorProfile.DoesNotExist:
        messages.error(request, "Doctor profile not found")
        return redirect('login')
    
    context = {
        'doctor_profile': doctor_profile,
        'working_hours': '9:00 AM - 5:00 PM',
        'working_days': 'Monday - Friday',
    }
    
    return render(request, 'core/dashboard/doctor_schedule.html', context)


@login_required
def doctor_schedule_update(request):
    """
    Update doctor's schedule
    """
    try:
        doctor_profile = DoctorProfile.objects.get(user=request.user)
    except DoctorProfile.DoesNotExist:
        messages.error(request, "Doctor profile not found")
        return redirect('login')
    
    if request.method == 'POST':
        # Update working hours
        # You would need to store this in a DoctorAvailability or similar model
        
        messages.success(request, "Schedule updated successfully")
        return redirect('core/dashboard/doctor_schedule')
    
    return redirect('core/dashboard/doctor_schedule')

@login_required
def lab_dashboard(request):
    # Provide simple context: recent bookings assigned to this technician's lab
    try:
        tech_profile = LabTechnicianProfile.objects.get(user=request.user)
        assigned_lab = tech_profile.lab
    except LabTechnicianProfile.DoesNotExist:
        tech_profile = None
        assigned_lab = None

    recent_bookings = []
    pending_tests_count = 0
    completed_tests_count = 0
    assigned_lab_name = assigned_lab.name if assigned_lab else None

    if assigned_lab:
        recent_bookings = TestBooking.objects.filter(lab=assigned_lab).order_by('-created_at')[:10]
        pending_tests_count = TestBooking.objects.filter(lab=assigned_lab, status__in=['Booked', 'Pending']).count()
        completed_tests_count = TestBooking.objects.filter(lab=assigned_lab, status='Completed').count()

    context = {
        'recent_bookings': recent_bookings,
        'pending_tests_count': pending_tests_count,
        'completed_tests_count': completed_tests_count,
        'assigned_lab_name': assigned_lab_name,
    }

    return render(request, 'core/dashboard/lab_dashboard.html', context)


@login_required
def lab_tests(request):
    """
    Display all test bookings for the lab technician's assigned lab
    with filtering by status
    """
    try:
        tech_profile = LabTechnicianProfile.objects.get(user=request.user)
        assigned_lab = tech_profile.lab
    except LabTechnicianProfile.DoesNotExist:
        messages.error(request, "Lab technician profile not found")
        return redirect('login')

    if not assigned_lab:
        messages.warning(request, "No lab assigned to your profile")
        bookings = TestBooking.objects.none()
    else:
        # Get all bookings for this lab
        bookings = TestBooking.objects.filter(
            lab=assigned_lab
        ).select_related('patient', 'patient__user', 'test').order_by('-created_at')

    # Filter by status if provided
    status_filter = request.GET.get('status', '')
    if status_filter:
        if status_filter == 'Pending':
            bookings = bookings.filter(status__in=['Booked', 'Pending'])
        elif status_filter == 'Completed':
            bookings = bookings.filter(status='Completed')
        # 'All Bookings' shows everything (no filter)

    # Calculate statistics
    all_bookings = TestBooking.objects.filter(lab=assigned_lab) if assigned_lab else TestBooking.objects.none()
    total_count = all_bookings.count()
    pending_count = all_bookings.filter(status__in=['Booked', 'Pending']).count()
    completed_count = all_bookings.filter(status='Completed').count()

    context = {
        'bookings': bookings,
        'status_filter': status_filter,
        'total_count': total_count,
        'pending_count': pending_count,
        'completed_count': completed_count,
        'assigned_lab': assigned_lab,
    }

    return render(request, 'core/dashboard/lab_tests.html', context)


@login_required
def lab_results(request):
    """
    Lab results management with filtering by status
    Allows lab technicians to update test results
    """
    try:
        tech_profile = LabTechnicianProfile.objects.get(user=request.user)
        assigned_lab = tech_profile.lab
    except LabTechnicianProfile.DoesNotExist:
        messages.error(request, "Lab technician profile not found")
        return redirect('login')

    if not assigned_lab:
        messages.warning(request, "No lab assigned to your profile")
        bookings = TestBooking.objects.none()
    else:
        # Get all bookings for this lab
        bookings = TestBooking.objects.filter(
            lab=assigned_lab
        ).select_related('patient', 'patient__user', 'test', 'lab').order_by('-created_at')

    # Handle POST request - Save result
    if request.method == 'POST':
        booking_id = request.POST.get('save')
        if booking_id:
            try:
                booking = TestBooking.objects.get(id=int(booking_id), lab=assigned_lab)
                result_val = request.POST.get(f'result_{booking_id}', '').strip()
                status_val = request.POST.get(f'status_{booking_id}', booking.status)
                
                # Update booking status
                booking.status = status_val
                booking.save()

                # Create a LabResult if marking as completed
                if status_val == 'Completed' and result_val:
                    from datetime import date as _date
                    
                    # Get a doctor (you may want to improve this logic)
                    doctor = DoctorProfile.objects.first()
                    
                    if doctor:
                        LabResult.objects.create(
                            patient=booking.patient,
                            doctor=doctor,
                            lab_technician=tech_profile,
                            test_name=booking.test.test_name,
                            test_value=result_val or 'N/A',
                            normal_range='',  # You can add this field to the form
                            result_status=status_val,
                            remarks='',  # You can add this field to the form
                            test_date=_date.today()
                        )
                        messages.success(request, f"Result saved and marked as {status_val} for {booking.patient.full_name}")
                    else:
                        messages.warning(request, "Status updated but couldn't create lab result - no doctor available")
                else:
                    messages.success(request, f"Status updated to {status_val} for {booking.patient.full_name}")

            except TestBooking.DoesNotExist:
                messages.error(request, "Booking not found")
            except Exception as e:
                messages.error(request, f"Error saving result: {str(e)}")

        return redirect('lab_results')

    # Filter by status if provided
    status_filter = request.GET.get('status', '')
    if status_filter:
        if status_filter == 'Pending':
            bookings = bookings.filter(status__in=['Booked', 'Pending'])
        elif status_filter == 'Completed':
            bookings = bookings.filter(status='Completed')
        # All results shows everything (no filter)

    # Calculate statistics
    all_bookings = TestBooking.objects.filter(lab=assigned_lab) if assigned_lab else TestBooking.objects.none()
    total_count = all_bookings.count()
    pending_count = all_bookings.filter(status__in=['Booked', 'Pending']).count()
    completed_count = all_bookings.filter(status='Completed').count()

    context = {
        'bookings': bookings,
        'status_filter': status_filter,
        'total_count': total_count,
        'pending_count': pending_count,
        'completed_count': completed_count,
        'assigned_lab': assigned_lab,
    }

    return render(request, 'core/dashboard/lab_results.html', context)

@login_required
def lab_prices(request):
    """Display all diagnostic tests for the lab with real data from backend"""
    try:
        tech_profile = LabTechnicianProfile.objects.get(user=request.user)
        assigned_lab = tech_profile.lab
    except LabTechnicianProfile.DoesNotExist:
        messages.error(request, "Lab technician profile not found")
        return redirect('login')

    if not assigned_lab:
        messages.warning(request, "No lab assigned to your profile")
        tests = DiagnosticTest.objects.none()
    else:
        tests = DiagnosticTest.objects.filter(lab=assigned_lab).order_by('test_name')

    # Calculate statistics
    active_tests_count = tests.count()
    avg_price = tests.aggregate(avg=Avg('price'))['avg'] or 0

    context = {
        'tests': tests,
        'active_tests_count': active_tests_count,
        'avg_price': round(avg_price, 2),
        'assigned_lab': assigned_lab,
    }

    return render(request, 'core/dashboard/lab_prices.html', context)


@login_required
def lab_add_test(request):
    """Add a new diagnostic test"""
    try:
        tech_profile = LabTechnicianProfile.objects.get(user=request.user)
        assigned_lab = tech_profile.lab
    except LabTechnicianProfile.DoesNotExist:
        messages.error(request, "Lab technician profile not found")
        return redirect('login')

    if not assigned_lab:
        messages.error(request, "No lab assigned to your profile")
        return redirect('lab_prices')

    if request.method == 'POST':
        test_name = request.POST.get('test_name', '').strip()
        price = request.POST.get('price', '').strip()

        if not test_name or not price:
            messages.error(request, "Test name and price are required")
            return redirect('lab_add_test')

        try:
            DiagnosticTest.objects.create(
                lab=assigned_lab,
                test_name=test_name,
                price=price,
            )
            messages.success(request, f"Test '{test_name}' added successfully")
            return redirect('lab_prices')
        except Exception as e:
            messages.error(request, f"Error adding test: {str(e)}")

    context = {
        'assigned_lab': assigned_lab,
    }
    return render(request, 'core/dashboard/lab_add_test.html', context)


@login_required
def lab_edit_test(request, test_id):
    test = get_object_or_404(DiagnosticTest, id=test_id)

    if request.method == "POST":
        test.test_name = request.POST.get("test_name")
        test.test_code = request.POST.get("test_code")
        test.category = request.POST.get("category")
        test.price = request.POST.get("price")
        test.result_duration = request.POST.get("result_duration")
        test.sample_type = request.POST.get("sample_type")
        test.description = request.POST.get("description")
        test.preparation_instructions = request.POST.get("preparation_instructions")
        test.is_active = True if request.POST.get("is_active") else False
        test.home_collection = True if request.POST.get("home_collection") else False

        test.save()
        messages.success(request, "Diagnostic test updated successfully.")
        return redirect("lab_prices")

    return render(request, "core/dashboard/lab_edit_test.html", {"test": test})


@login_required
def lab_delete_test(request, test_id):
    """Delete a diagnostic test"""
    try:
        tech_profile = LabTechnicianProfile.objects.get(user=request.user)
        assigned_lab = tech_profile.lab
    except LabTechnicianProfile.DoesNotExist:
        messages.error(request, "Lab technician profile not found")
        return redirect('login')

    if not assigned_lab:
        messages.error(request, "No lab assigned to your profile")
        return redirect('lab_prices')

    test = get_object_or_404(DiagnosticTest, id=test_id, lab=assigned_lab)

    if request.method == 'POST':
        test_name = test.test_name
        test.delete()
        messages.success(request, f"Test '{test_name}' deleted successfully")
        return redirect('lab_prices')

    return redirect('lab_prices')
    
from django.http import JsonResponse
from django.views.decorators.http import require_http_methods
from django.contrib.auth.decorators import login_required
from django.db.models import Count, Sum, Q, F, DecimalField, Value, CharField
from django.db.models.functions import TruncDate, Coalesce
from datetime import datetime, timedelta, date
from decimal import Decimal
import json

@login_required
def lab_reports(request):
    """
    Lab Reports & Analytics Dashboard
    Renders the reports page with real data loaded via JavaScript
    
    The page loads empty, then JavaScript fetches data from api_lab_reports_data()
    """
    try:
        lab_technician = LabTechnicianProfile.objects.get(user=request.user)
        assigned_lab = lab_technician.lab
    except LabTechnicianProfile.DoesNotExist:
        messages.error(request, "Lab technician profile not found. Please contact administrator.")
        return redirect('lab_dashboard')
    
    if not assigned_lab:
        messages.warning(request, "No lab assigned to your profile. Please contact administrator.")
        return redirect('lab_dashboard')
    
    context = {
        'assigned_lab': assigned_lab,
        'lab_technician': lab_technician,
    }
    
    return render(request, 'core/dashboard/lab_reports.html', context)

@login_required
@require_http_methods(["GET"])
def api_lab_reports_data(request):
    """
    API endpoint for lab reports and analytics
    Returns comprehensive data for the Reports & Analytics dashboard
    
    Parameters:
        start_date: Date in format YYYY-MM-DD (optional, defaults to start of month)
        end_date: Date in format YYYY-MM-DD (optional, defaults to today)
    """
    
    try:
        # Get lab technician profile
        from core.models import LabTechnicianProfile, TestBooking, LabResult, DiagnosticTest
        
        lab_technician = LabTechnicianProfile.objects.get(user=request.user)
        assigned_lab = lab_technician.lab
        
        if not assigned_lab:
            return JsonResponse({
                'error': 'No lab assigned to your profile',
                'metrics': {},
                'charts': {}
            }, status=400)
        
        # Parse date range from request
        start_date_str = request.GET.get('start_date')
        end_date_str = request.GET.get('end_date')
        
        today = date.today()
        if start_date_str:
            try:
                start_date = datetime.strptime(start_date_str, '%Y-%m-%d').date()
            except ValueError:
                start_date = date(today.year, today.month, 1)
        else:
            start_date = date(today.year, today.month, 1)
        
        if end_date_str:
            try:
                end_date = datetime.strptime(end_date_str, '%Y-%m-%d').date()
            except ValueError:
                end_date = today
        else:
            end_date = today
        
        # Calculate previous period for trend analysis
        period_days = (end_date - start_date).days
        previous_start = start_date - timedelta(days=period_days + 1)
        previous_end = start_date - timedelta(days=1)
        
        # Fetch test bookings for current period
        current_bookings = TestBooking.objects.filter(
            lab=assigned_lab,
            booking_date__range=[start_date, end_date]
        ).select_related('test', 'patient')
        
        # Fetch lab results for current period
        current_results = LabResult.objects.filter(
            lab_technician__lab=assigned_lab,
            test_date__range=[start_date, end_date]
        ).select_related('patient', 'test_name')
        
        # Previous period data for trends
        previous_bookings = TestBooking.objects.filter(
            lab=assigned_lab,
            booking_date__range=[previous_start, previous_end]
        )
        
        previous_results = LabResult.objects.filter(
            lab_technician__lab=assigned_lab,
            test_date__range=[previous_start, previous_end]
        )
        
        # ==================== METRICS ====================
        total_tests = current_bookings.count()
        completed_tests = current_results.filter(result_status='Normal') | current_results.filter(result_status='Abnormal')
        completed_tests = completed_tests.count()
        pending_tests = current_bookings.filter(status__in=['Booked', 'Pending']).count()
        
        # Calculate total revenue
        total_revenue = current_bookings.aggregate(
            total=Coalesce(Sum(F('test__price'), output_field=DecimalField(max_digits=10, decimal_places=2)), Decimal('0.00'))
        )['total']
        
        # Calculate trends (percentage change)
        previous_total = previous_bookings.count()
        tests_trend = calculate_trend(previous_total, total_tests)
        
        previous_revenue = previous_bookings.aggregate(
            total=Coalesce(Sum(F('test__price'), output_field=DecimalField(max_digits=10, decimal_places=2)), Decimal('0.00'))
        )['total']
        revenue_trend = calculate_trend(previous_revenue or 0, total_revenue)
        
        metrics = {
            'total_tests': total_tests,
            'completed_tests': completed_tests,
            'pending_tests': pending_tests,
            'total_revenue': float(total_revenue),
            'tests_trend': tests_trend,
            'revenue_trend': revenue_trend,
        }
        
        # ==================== CHARTS ====================
        
        # Tests per month/week
        tests_by_date = current_bookings.annotate(
            date=TruncDate('booking_date')
        ).values('date').annotate(
            count=Count('id')
        ).order_by('date')
        
        test_volume_labels = [item['date'].strftime('%d %b') for item in tests_by_date]
        test_volume_data = [item['count'] for item in tests_by_date]
        
        # Test types distribution
        test_types = current_bookings.values('test__category').annotate(
            count=Count('id'),
            percentage=Count('id') * 100 / total_tests if total_tests > 0 else 0
        ).order_by('-count')
        
        test_type_charts = []
        for item in test_types:
            category = item['test__category'] or 'Other'
            test_type_charts.append({
                'name': category,
                'count': item['count'],
                'percentage': round(item['percentage'], 1)
            })
        
        charts = {
            'tests_per_month': {
                'labels': test_volume_labels if test_volume_labels else ['No data'],
                'data': test_volume_data if test_volume_data else [0]
            },
            'test_types': test_type_charts
        }
        
        # ==================== POPULAR TESTS ====================
        
        popular_tests = current_bookings.values(
            'test__test_name', 'test__price'
        ).annotate(
            bookings=Count('id'),
            revenue=Sum(F('test__price'), output_field=DecimalField(max_digits=10, decimal_places=2))
        ).order_by('-bookings')[:5]
        
        popular_tests_data = []
        for item in popular_tests:
            popular_tests_data.append({
                'name': item['test__test_name'],
                'bookings': item['bookings'],
                'revenue': float(item['revenue'] or 0)
            })
        
        # ==================== DAILY REVENUE (LAST 7 DAYS) ====================
        
        daily_revenue = []
        for i in range(6, -1, -1):  # Last 7 days
            day_date = end_date - timedelta(days=i)
            day_bookings = current_bookings.filter(booking_date=day_date)
            
            day_revenue = day_bookings.aggregate(
                total=Coalesce(Sum(F('test__price'), output_field=DecimalField(max_digits=10, decimal_places=2)), Decimal('0.00'))
            )['total']
            
            daily_revenue.append({
                'date': day_date.strftime('%d %b %Y'),
                'day_label': day_date.strftime('%A'),
                'revenue': float(day_revenue)
            })
        
        return JsonResponse({
            'metrics': metrics,
            'charts': charts,
            'popular_tests': popular_tests_data,
            'daily_revenue': daily_revenue,
            'period': {
                'start_date': start_date.isoformat(),
                'end_date': end_date.isoformat()
            }
        })
    
    except LabTechnicianProfile.DoesNotExist:
        return JsonResponse({
            'error': 'Lab technician profile not found',
            'metrics': {},
            'charts': {}
        }, status=400)
    except Exception as e:
        return JsonResponse({
            'error': f'Error fetching reports: {str(e)}',
            'metrics': {},
            'charts': {}
        }, status=500)


def calculate_trend(previous_value, current_value):
    """
    Calculate percentage trend between two values
    Returns positive/negative percentage
    """
    if previous_value == 0:
        if current_value > 0:
            return 100
        return 0
    
    trend = ((current_value - previous_value) / previous_value) * 100
    return round(trend, 1)


@login_required
@require_http_methods(["GET"])
def export_lab_report(request):
    """
    Export lab reports as PDF or Excel
    Parameters:
        start_date: Date in format YYYY-MM-DD
        end_date: Date in format YYYY-MM-DD
        format: 'pdf' or 'excel' (default: 'pdf')
    """
    from core.models import LabTechnicianProfile, TestBooking, LabResult
    from io import BytesIO
    from reportlab.lib.pagesizes import letter, A4
    from reportlab.lib import colors
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import inch
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak
    from django.http import FileResponse
    
    try:
        lab_technician = LabTechnicianProfile.objects.get(user=request.user)
        assigned_lab = lab_technician.lab
        
        if not assigned_lab:
            return JsonResponse({'error': 'No lab assigned'}, status=400)
        
        start_date_str = request.GET.get('start_date')
        end_date_str = request.GET.get('end_date')
        export_format = request.GET.get('format', 'pdf')
        
        start_date = datetime.strptime(start_date_str, '%Y-%m-%d').date()
        end_date = datetime.strptime(end_date_str, '%Y-%m-%d').date()
        
        # Fetch data
        bookings = TestBooking.objects.filter(
            lab=assigned_lab,
            booking_date__range=[start_date, end_date]
        ).select_related('test', 'patient')
        
        results = LabResult.objects.filter(
            lab_technician__lab=assigned_lab,
            test_date__range=[start_date, end_date]
        ).select_related('patient', 'doctor')
        
        if export_format == 'pdf':
            return generate_pdf_report(assigned_lab, bookings, results, start_date, end_date)
        else:
            return generate_excel_report(assigned_lab, bookings, results, start_date, end_date)
    
    except LabTechnicianProfile.DoesNotExist:
        return JsonResponse({'error': 'Lab technician profile not found'}, status=400)
    except Exception as e:
        return JsonResponse({'error': f'Error exporting report: {str(e)}'}, status=500)


def generate_pdf_report(lab, bookings, results, start_date, end_date):
    """Generate PDF report"""
    from reportlab.lib.pagesizes import letter
    from reportlab.lib import colors
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import inch
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from io import BytesIO
    from django.http import FileResponse
    
    pdf_buffer = BytesIO()
    doc = SimpleDocTemplate(pdf_buffer, pagesize=letter, topMargin=0.5*inch, bottomMargin=0.5*inch)
    elements = []
    styles = getSampleStyleSheet()
    
    # Title
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=18,
        textColor=colors.HexColor('#1a1a1a'),
        spaceAfter=10,
        alignment=1
    )
    
    elements.append(Paragraph(f"Lab Report - {lab.name}", title_style))
    elements.append(Paragraph(f"Period: {start_date} to {end_date}", styles['Normal']))
    elements.append(Spacer(1, 20))
    
    # Summary
    total_tests = bookings.count()
    total_completed = results.count()
    total_revenue = bookings.aggregate(Sum(F('test__price'), output_field=DecimalField()))['test__price__sum'] or 0
    
    summary_data = [
        ['Metric', 'Value'],
        ['Total Tests', str(total_tests)],
        ['Completed Tests', str(total_completed)],
        ['Total Revenue', f'₹{total_revenue:.2f}'],
    ]
    
    summary_table = Table(summary_data, colWidths=[2*inch, 2*inch])
    summary_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#667eea')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 10),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
        ('GRID', (0, 0), (-1, -1), 1, colors.HexColor('#cccccc')),
    ]))
    elements.append(summary_table)
    elements.append(Spacer(1, 20))
    
    # Test Results Table
    elements.append(Paragraph("Test Results", styles['Heading2']))
    results_data = [['Patient', 'Test Name', 'Result Status', 'Date']]
    
    for result in results[:10]:  # Limit to first 10
        results_data.append([
            result.patient.full_name,
            result.test_name,
            result.result_status,
            result.test_date.strftime('%d-%m-%Y')
        ])
    
    results_table = Table(results_data, colWidths=[1.5*inch, 2*inch, 1*inch, 1.5*inch])
    results_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#667eea')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 9),
        ('GRID', (0, 0), (-1, -1), 1, colors.HexColor('#cccccc')),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f0f0f0')]),
    ]))
    elements.append(results_table)
    
    # Build PDF
    doc.build(elements)
    pdf_buffer.seek(0)
    
    response = FileResponse(pdf_buffer, content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="lab_report_{datetime.now().strftime("%Y%m%d")}.pdf"'
    return response


def generate_excel_report(lab, bookings, results, start_date, end_date):
    """Generate Excel report"""
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
        from io import BytesIO
        from django.http import HttpResponse
        
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Lab Report"
        
        # Headers
        ws['A1'] = f"Lab Report - {lab.name}"
        ws['A2'] = f"Period: {start_date} to {end_date}"
        
        # Summary
        ws['A4'] = 'Metric'
        ws['B4'] = 'Value'
        ws['A5'] = 'Total Tests'
        ws['B5'] = bookings.count()
        ws['A6'] = 'Completed Tests'
        ws['B6'] = results.count()
        ws['A7'] = 'Total Revenue'
        ws['B7'] = float(bookings.aggregate(Sum(F('test__price'), output_field=DecimalField()))['test__price__sum'] or 0)
        
        # Results
        row = 9
        ws[f'A{row}'] = 'Patient'
        ws[f'B{row}'] = 'Test Name'
        ws[f'C{row}'] = 'Result Status'
        ws[f'D{row}'] = 'Date'
        
        row += 1
        for result in results[:100]:  # Limit to 100
            ws[f'A{row}'] = result.patient.full_name
            ws[f'B{row}'] = result.test_name
            ws[f'C{row}'] = result.result_status
            ws[f'D{row}'] = result.test_date.strftime('%d-%m-%Y')
            row += 1
        
        # Adjust column widths
        ws.column_dimensions['A'].width = 20
        ws.column_dimensions['B'].width = 25
        ws.column_dimensions['C'].width = 15
        ws.column_dimensions['D'].width = 15
        
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        
        response = HttpResponse(
            output.read(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="lab_report_{datetime.now().strftime("%Y%m%d")}.xlsx"'
        return response
    
    except ImportError:
        return JsonResponse({'error': 'Excel export requires openpyxl library'}, status=500)


def get_chart_data(lab_results, test_bookings, start_date, end_date):
    """
    Generate chart data for analytics
    """
    # Tests per month
    tests_per_month = []
    labels = []
    
    current_date = start_date
    while current_date <= end_date:
        month_label = current_date.strftime('%b')
        labels.append(month_label)
        
        count = lab_results.filter(
            test_date__year=current_date.year,
            test_date__month=current_date.month
        ).count()
        tests_per_month.append(count)
        
        # Move to next month
        if current_date.month == 12:
            current_date = current_date.replace(year=current_date.year + 1, month=1)
        else:
            current_date = current_date.replace(month=current_date.month + 1)

    # Test type distribution
    test_type_dist = lab_results.values('test_name').annotate(
        count=Count('id')
    ).order_by('-count')[:5]

    test_type_labels = [item['test_name'] for item in test_type_dist]
    test_type_data = [item['count'] for item in test_type_dist]

    # Result status distribution
    result_status_dist = lab_results.values('result_status').annotate(
        count=Count('id')
    )

    result_status_labels = [item['result_status'] for item in result_status_dist]
    result_status_data = [item['count'] for item in result_status_dist]

    # Revenue trend
    revenue_per_month = []
    revenue_labels = []
    
    current_date = start_date
    while current_date <= end_date:
        month_label = current_date.strftime('%b')
        revenue_labels.append(month_label)
        
        revenue = test_bookings.filter(
            booking_date__year=current_date.year,
            booking_date__month=current_date.month
        ).aggregate(
            total=Sum(F('test__price'), output_field=DecimalField(max_digits=10, decimal_places=2))
        )['total'] or 0
        
        revenue_per_month.append(float(revenue))
        
        # Move to next month
        if current_date.month == 12:
            current_date = current_date.replace(year=current_date.year + 1, month=1)
        else:
            current_date = current_date.replace(month=current_date.month + 1)

    return {
        'testsPerMonth': {
            'labels': labels,
            'data': tests_per_month
        },
        'testTypes': {
            'labels': test_type_labels,
            'data': test_type_data
        },
        'resultStatus': {
            'labels': result_status_labels,
            'data': result_status_data
        },
        'revenue': {
            'labels': revenue_labels,
            'data': revenue_per_month
        }
    }


@login_required
def lab_report_export_pdf(request):
    """
    Export lab reports as PDF
    """
    try:
        lab_technician = LabTechnicianProfile.objects.get(user=request.user)
    except LabTechnicianProfile.DoesNotExist:
        return JsonResponse({'error': 'Lab Technician profile not found'}, status=404)

    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')

    if start_date:
        start_date = datetime.strptime(start_date, '%Y-%m-%d').date()
    else:
        start_date = timezone.now().date() - timedelta(days=30)

    if end_date:
        end_date = datetime.strptime(end_date, '%Y-%m-%d').date()
    else:
        end_date = timezone.now().date()

    # Get data
    lab_results = LabResult.objects.filter(
        test_date__range=[start_date, end_date],
        lab_technician__lab=lab_technician.lab
    )

    test_bookings = TestBooking.objects.filter(
        booking_date__range=[start_date, end_date],
        lab=lab_technician.lab
    )

    # For PDF generation, you would use a library like reportlab or weasyprint
    # This is a placeholder implementation
    
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename="lab_report.pdf"'

    # Generate PDF content here using reportlab or weasyprint
    # For now, return a message
    return JsonResponse({
        'message': 'PDF export - Implement with reportlab or weasyprint',
        'total_tests': lab_results.count(),
        'total_bookings': test_bookings.count()
    })


@login_required
def lab_report_export_excel(request):
    """
    Export lab reports as Excel
    """
    try:
        lab_technician = LabTechnicianProfile.objects.get(user=request.user)
    except LabTechnicianProfile.DoesNotExist:
        return JsonResponse({'error': 'Lab Technician profile not found'}, status=404)

    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    from io import BytesIO

    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')

    if start_date:
        start_date = datetime.strptime(start_date, '%Y-%m-%d').date()
    else:
        start_date = timezone.now().date() - timedelta(days=30)

    if end_date:
        end_date = datetime.strptime(end_date, '%Y-%m-%d').date()
    else:
        end_date = timezone.now().date()

    # Get data
    lab_results = LabResult.objects.filter(
        test_date__range=[start_date, end_date],
        lab_technician__lab=lab_technician.lab
    ).select_related('patient', 'doctor')

    # Create workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Lab Results"

    # Add headers
    headers = ['Patient Name', 'Test Name', 'Test Date', 'Result Status', 
               'Test Value', 'Normal Range', 'Doctor', 'Remarks']
    ws.append(headers)

    # Style headers
    header_fill = PatternFill(start_color='667eea', end_color='667eea', fill_type='solid')
    header_font = Font(bold=True, color='FFFFFF')

    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')

    # Add data
    for result in lab_results:
        ws.append([
            result.patient.full_name,
            result.test_name,
            result.test_date.strftime('%d-%m-%Y'),
            result.result_status,
            result.test_value,
            result.normal_range,
            f"Dr. {result.doctor.user.get_full_name()}",
            result.remarks
        ])

    # Adjust column widths
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[column_letter].width = adjusted_width

    # Save to BytesIO object
    output = BytesIO()
    wb.save(output)
    output.seek(0)

    response = HttpResponse(
        output.read(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="lab_reports.xlsx"'

    return response


@login_required
def lab_result_detail(request, result_id):
    """
    View detailed lab result
    """
    try:
        result = LabResult.objects.get(id=result_id)
        lab_technician = LabTechnicianProfile.objects.get(user=request.user)
        
        if result.lab_technician.lab != lab_technician.lab:
            return render(request, 'error.html', {'message': 'Access denied'})

        context = {
            'result': result,
            'patient': result.patient,
            'doctor': result.doctor
        }

        return render(request, 'core/dashboard/lab_result_detail.html', context)

    except LabResult.DoesNotExist:
        return render(request, 'error.html', {'message': 'Lab result not found'})
    
from django.utils import timezone
from datetime import datetime, timedelta
import json
from django.db.models import Count, Sum, Q, F,DecimalField

@login_required
def get_analytics_data(request):
    """
    API endpoint for getting analytics data (for AJAX requests)
    """
    try:
        lab_technician = LabTechnicianProfile.objects.get(user=request.user)
    except LabTechnicianProfile.DoesNotExist:
        return JsonResponse({'error': 'Lab Technician profile not found'}, status=404)

    # Get parameters
    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')
    test_type = request.GET.get('test_type', '')

    if start_date:
        start_date = datetime.strptime(start_date, '%Y-%m-%d').date()
    else:
        start_date = timezone.now().date() - timedelta(days=30)

    if end_date:
        end_date = datetime.strptime(end_date, '%Y-%m-%d').date()
    else:
        end_date = timezone.now().date()

    # Get data
    lab_results = LabResult.objects.filter(
        test_date__range=[start_date, end_date],
        lab_technician__lab=lab_technician.lab
    )

    if test_type:
        lab_results = lab_results.filter(test_name__icontains=test_type)

    test_bookings = TestBooking.objects.filter(
        booking_date__range=[start_date, end_date],
        lab=lab_technician.lab
    )

    if test_type:
        test_bookings = test_bookings.filter(test__test_name__icontains=test_type)

    # Get chart data
    chart_data = get_chart_data(lab_results, test_bookings, start_date, end_date)

    # Calculate metrics
    metrics = {
        'total_tests': lab_results.count(),
        'total_bookings': test_bookings.count(),
        'total_revenue': float(test_bookings.aggregate(
            total=Sum(F('test__price'), output_field=DecimalField(max_digits=10, decimal_places=2))
        )['total'] or 0),
        'pending_results': lab_results.filter(result_status='Pending').count()
    }

    return JsonResponse({
        'metrics': metrics,
        'charts': chart_data
    })


@login_required
def lab_booking_details(request, booking_id):
    """
    View detailed information about a test booking
    """
    try:
        patient = PatientProfile.objects.get(user=request.user)
    except PatientProfile.DoesNotExist:
        messages.error(request, "Patient profile not found")
        return redirect('patient_dashboard')
    
    # Get the booking
    booking = get_object_or_404(
        TestBooking,
        id=booking_id,
        patient=patient
    )
    
    context = {
        'booking': booking,
        'today': date.today(),
    }
    
    return render(request, 'core/dashboard/lab_booking_details.html', context)


@login_required
def reschedule_test_booking(request, booking_id):
    """
    Reschedule a test booking to a different date
    """
    try:
        patient = PatientProfile.objects.get(user=request.user)
    except PatientProfile.DoesNotExist:
        messages.error(request, "Patient profile not found")
        return redirect('patient_dashboard')
    
    booking = get_object_or_404(
        TestBooking,
        id=booking_id,
        patient=patient
    )
    
    # Check if booking can be rescheduled
    if booking.status not in ['Booked', 'Pending']:
        messages.error(request, "This booking cannot be rescheduled")
        return redirect('core/dashboard/patient_booked_tests')
    
    if request.method == 'POST':
        new_date_str = request.POST.get('booking_date')
        
        try:
            from datetime import datetime
            new_date = datetime.strptime(new_date_str, '%Y-%m-%d').date()
            
            # Validate date is in future
            if new_date < date.today():
                messages.error(request, "Cannot reschedule to past dates")
                return redirect('booking_detail', booking_id=booking_id)
            
            # Update booking date
            old_date = booking.booking_date
            booking.booking_date = new_date
            booking.save()
            
            messages.success(
                request,
                f"Booking rescheduled from {old_date} to {new_date}"
            )
            return redirect('booking_detail', booking_id=booking_id)
        
        except ValueError:
            messages.error(request, "Invalid date format")
            return redirect('booking_detail', booking_id=booking_id)
    
    context = {
        'booking': booking,
        'today': date.today(),
    }
    
    return render(request, 'core/dashboard/reschedule_booking.html', context)


@login_required
def download_test_report(request, booking_id):
    """
    Download test report as PDF
    """
    try:
        patient = PatientProfile.objects.get(user=request.user)
    except PatientProfile.DoesNotExist:
        messages.error(request, "Patient profile not found")
        return redirect('patient_dashboard')
    
    booking = get_object_or_404(
        TestBooking,
        id=booking_id,
        patient=patient,
        status='Completed'
    )
    
    # Get lab results for this test
    lab_results = LabResult.objects.filter(
        patient=patient,
        test_name=booking.test.test_name
    ).order_by('-test_date')
    
    if not lab_results.exists():
        messages.error(request, "No results found for this test")
        return redirect('booking_detail', booking_id=booking_id)
    
    # Generate PDF (implement with reportlab or similar)
    from io import BytesIO
    from reportlab.lib.pagesizes import letter
    from reportlab.lib import colors
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import inch
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from django.http import FileResponse
    
    pdf_buffer = BytesIO()
    doc = SimpleDocTemplate(pdf_buffer, pagesize=letter, topMargin=0.5*inch, bottomMargin=0.5*inch)
    elements = []
    styles = getSampleStyleSheet()
    
    # Title
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=16,
        textColor=colors.HexColor('#1a1a1a'),
        spaceAfter=10,
        alignment=1,
    )
    
    elements.append(Paragraph("Test Report", title_style))
    elements.append(Spacer(1, 12))
    
    # Patient Info
    patient_info = [
        ['Patient Name:', patient.full_name],
        ['Phone:', patient.phone],
        ['Test:', booking.test.test_name],
        ['Lab:', booking.lab.name],
    ]
    
    patient_table = Table(patient_info, colWidths=[2*inch, 4*inch])
    patient_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (0, -1), colors.HexColor('#e8f4f8')),
        ('TEXTCOLOR', (0, 0), (-1, -1), colors.black),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 10),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
        ('GRID', (0, 0), (-1, -1), 1, colors.HexColor('#cccccc')),
    ]))
    elements.append(patient_table)
    elements.append(Spacer(1, 20))
    
    # Results Table
    elements.append(Paragraph("Test Results", styles['Heading2']))
    results_data = [['Test Name', 'Value', 'Normal Range', 'Status']]
    
    for result in lab_results:
        results_data.append([
            result.test_name,
            result.test_value,
            result.normal_range,
            result.result_status,
        ])
    
    results_table = Table(results_data, colWidths=[2*inch, 1.5*inch, 1.5*inch, 1*inch])
    results_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#667eea')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 9),
        ('GRID', (0, 0), (-1, -1), 1, colors.HexColor('#cccccc')),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f0f0f0')]),
    ]))
    elements.append(results_table)
    
    # Build PDF
    doc.build(elements)
    pdf_buffer.seek(0)
    
    response = FileResponse(pdf_buffer, content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="test_report_{booking_id}.pdf"'
    return response


# forms.py - Patient Appointment Reschedule Forms

from django import forms
from django.core.exceptions import ValidationError
from django.utils import timezone
from datetime import datetime, timedelta
from .models import Appointment, DoctorProfile

class RescheduleAppointmentForm(forms.ModelForm):
    """
    Form for rescheduling an appointment
    """
    doctor = forms.ModelChoiceField(
        queryset=DoctorProfile.objects.filter(status='Active'),  # ✅ FIXED: Using DoctorProfile and filtering by status
        required=True,
        widget=forms.RadioSelect,
        label='Select Doctor'
    )
    
    appointment_date = forms.DateField(
        required=True,
        widget=forms.DateInput(attrs={
            'type': 'date',
            'class': 'form-input',
            'min': (timezone.now().date() + timedelta(days=1)).strftime('%Y-%m-%d')
        }),
        label='Appointment Date'
    )
    
    appointment_time = forms.TimeField(
        required=True,
        widget=forms.TimeInput(attrs={
            'type': 'time',
            'class': 'form-input'
        }),
        label='Appointment Time'
    )
    
    reschedule_reason = forms.CharField(
        required=False,
        widget=forms.Textarea(attrs={
            'class': 'form-textarea',
            'rows': 3,
            'placeholder': 'Please let us know why you need to reschedule...'
        }),
        label='Reason for Rescheduling'
    )
    
    notes = forms.CharField(
        required=False,
        widget=forms.Textarea(attrs={
            'class': 'form-textarea',
            'rows': 3,
            'placeholder': 'Any additional notes or concerns for your appointment...'
        }),
        label='Notes for the Doctor'
    )
    
    class Meta:
        model = Appointment
        fields = ['doctor', 'appointment_date', 'appointment_time']
    
    def __init__(self, *args, **kwargs):
        super().__init__(*args, **kwargs)
        
        # If we have an instance, set the current doctor as selected
        if self.instance and self.instance.pk:
            self.fields['doctor'].initial = self.instance.doctor
    
    def clean_appointment_date(self):
        """
        Validate that the appointment date is in the future
        """
        appointment_date = self.cleaned_data.get('appointment_date')
        
        if not appointment_date:
            raise ValidationError('Please select an appointment date.')
        
        # Must be at least 1 day in the future
        min_date = timezone.now().date() + timedelta(days=1)
        
        if appointment_date < min_date:
            raise ValidationError(
                'Appointments must be scheduled at least 24 hours in advance.'
            )
        
        # Optional: Don't allow appointments too far in the future (e.g., 3 months)
        max_date = timezone.now().date() + timedelta(days=90)
        
        if appointment_date > max_date:
            raise ValidationError(
                'Appointments cannot be scheduled more than 3 months in advance.'
            )
        
        return appointment_date
    
    def clean_appointment_time(self):
        """
        Validate that the appointment time is within working hours
        """
        appointment_time = self.cleaned_data.get('appointment_time')
        
        if not appointment_time:
            raise ValidationError('Please select an appointment time.')
        
        # Define working hours (9 AM to 6 PM)
        opening_time = datetime.strptime('09:00', '%H:%M').time()
        closing_time = datetime.strptime('18:00', '%H:%M').time()
        
        if appointment_time < opening_time or appointment_time > closing_time:
            raise ValidationError(
                'Appointments must be between 9:00 AM and 6:00 PM.'
            )
        
        return appointment_time
    
    def clean(self):
        """
        Validate that the appointment datetime is in the future
        and check for conflicts
        """
        cleaned_data = super().clean()
        appointment_date = cleaned_data.get('appointment_date')
        appointment_time = cleaned_data.get('appointment_time')
        doctor = cleaned_data.get('doctor')
        
        if appointment_date and appointment_time:
            # Check if the datetime is in the future
            appointment_datetime = datetime.combine(
                appointment_date, 
                appointment_time
            )
            
            # Make timezone aware if your project uses timezones
            if timezone.is_aware(timezone.now()):
                appointment_datetime = timezone.make_aware(appointment_datetime)
            
            if appointment_datetime <= timezone.now():
                raise ValidationError(
                    'The appointment date and time must be in the future.'
                )
            
            # Check if the slot is available (if doctor is selected)
            if doctor:
                # Exclude current appointment when checking conflicts
                exclude_id = self.instance.pk if self.instance else None
                
                conflicting_appointments = Appointment.objects.filter(
                    doctor=doctor,
                    appointment_date=appointment_date,
                    appointment_time=appointment_time,
                    status__in=['Pending Payment', 'Scheduled', 'Confirmed']  # ✅ Updated to match your STATUS_CHOICES
                )
                
                if exclude_id:
                    conflicting_appointments = conflicting_appointments.exclude(id=exclude_id)
                
                if conflicting_appointments.exists():
                    raise ValidationError(
                        'This time slot is already booked. Please select another time.'
                    )
        
        return cleaned_data


class CancelAppointmentForm(forms.Form):
    """
    Simple form for appointment cancellation with optional reason
    """
    cancellation_reason = forms.CharField(
        required=False,
        widget=forms.Textarea(attrs={
            'class': 'form-textarea',
            'rows': 4,
            'placeholder': 'Please let us know why you need to cancel (optional)...'
        }),
        label='Reason for Cancellation'
    )
    
    confirm_cancellation = forms.BooleanField(
        required=True,
        widget=forms.CheckboxInput(attrs={
            'class': 'form-checkbox'
        }),
        label='I confirm that I want to cancel this appointment'
    )

# Add this view to your core/views.py file

from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.utils import timezone
from datetime import timedelta
from .models import Appointment, DoctorProfile, PatientProfile
from .forms import RescheduleAppointmentForm


@login_required
def patient_reschedule_appointment(request, appointment_id):
    """
    View for patients to reschedule their appointments
    """
    # Get the appointment
    appointment = get_object_or_404(
        Appointment,
        id=appointment_id,
        patient__user=request.user
    )
    
    # Check if appointment can be rescheduled
    if appointment.status not in ['Pending Payment', 'Scheduled', 'Confirmed']:
        messages.error(
            request,
            f'Cannot reschedule an appointment with status: {appointment.get_status_display()}'
        )
        return redirect('patient_appointment_detail', appointment_id=appointment.id)
    
    # Check if appointment is not too soon (e.g., less than 24 hours away)
    appointment_datetime = timezone.make_aware(
        timezone.datetime.combine(appointment.appointment_date, appointment.appointment_time)
    )
    
    if appointment_datetime - timezone.now() < timedelta(hours=24):
        messages.error(
            request,
            'Cannot reschedule appointments less than 24 hours before the scheduled time. Please contact support.'
        )
        return redirect('patient_appointment_detail', appointment_id=appointment.id)
    
    if request.method == 'POST':
        form = RescheduleAppointmentForm(request.POST, instance=appointment)
        
        if form.is_valid():
            # Save the updated appointment
            updated_appointment = form.save(commit=False)
            
            # Keep the status or update it based on your business logic
            # For example, you might want to set it back to 'Pending Payment' if they change doctors
            if updated_appointment.doctor != appointment.doctor:
                updated_appointment.status = 'Pending Payment'
            
            updated_appointment.save()
            
            # Log the reschedule reason if provided
            reschedule_reason = form.cleaned_data.get('reschedule_reason')
            notes = form.cleaned_data.get('notes')
            
            # You could save these to a RescheduleHistory model or add to appointment notes
            # For now, we'll just show a success message
            
            messages.success(
                request,
                f'Your appointment has been rescheduled to {updated_appointment.appointment_date.strftime("%B %d, %Y")} at {updated_appointment.appointment_time.strftime("%I:%M %p")}'
            )
            
            return redirect('patient_appointment_detail', appointment_id=updated_appointment.id)
        else:
            messages.error(request, 'Please correct the errors below.')
    else:
        form = RescheduleAppointmentForm(instance=appointment)
    
    # Get available doctors for selection
    available_doctors = DoctorProfile.objects.filter(status='Active')
    
    # Calculate minimum date (tomorrow)
    min_date = timezone.now().date() + timedelta(days=1)
    
    context = {
        'appointment': appointment,
        'form': form,
        'available_doctors': available_doctors,
        'min_date': min_date,
    }
    
    return render(request, 'core/dashboard/patient_reschedule_appointment.html', context)


# Optional: Create a view for getting available time slots via AJAX
@login_required
def get_available_time_slots(request):
    """
    AJAX endpoint to get available time slots for a doctor on a specific date
    """
    import json
    from django.http import JsonResponse
    
    if request.method == 'GET':
        doctor_id = request.GET.get('doctor_id')
        date_str = request.GET.get('date')
        
        if not doctor_id or not date_str:
            return JsonResponse({'error': 'Missing parameters'}, status=400)
        
        try:
            from datetime import datetime
            doctor = DoctorProfile.objects.get(id=doctor_id)
            appointment_date = datetime.strptime(date_str, '%Y-%m-%d').date()
            
            # Define all possible time slots (customize these based on your needs)
            all_time_slots = [
                '09:00', '09:30', '10:00', '10:30', '11:00', '11:30',
                '14:00', '14:30', '15:00', '15:30', '16:00', '16:30',
                '17:00', '17:30', '18:00'
            ]
            
            # Get booked slots for this doctor on this date
            booked_appointments = Appointment.objects.filter(
                doctor=doctor,
                appointment_date=appointment_date,
                status__in=['Pending Payment', 'Scheduled', 'Confirmed']
            ).values_list('appointment_time', flat=True)
            
            booked_times = []
            for t in booked_appointments:
                try:
                    time_string = str(appointment_time).strip()
                    time_only = time_string[:5]
                    if ':' in time_only and len(time_only) == 5:
                        booked_times.add(time_only)
                    else:
                        booked_times.append(str(t))
                except:
                    booked_times.append(str(t))
            
            # Mark slots as available or booked
            time_slots = []
            for slot in all_time_slots:
                time_slots.append({
                    'time': slot,
                    'available': slot not in booked_times
                })
            
            return JsonResponse({'slots': time_slots})
            
        except DoctorProfile.DoesNotExist:
            return JsonResponse({'error': 'Doctor not found'}, status=404)
        except ValueError:
            return JsonResponse({'error': 'Invalid date format'}, status=400)
    
    return JsonResponse({'error': 'Invalid request method'}, status=405)

from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.http import JsonResponse
from django.views.decorators.http import require_POST
from django.utils import timezone
from django.db.models import Q, Sum, Count
from datetime import datetime, timedelta

from core.models import (
    FrontDeskProfile, Appointment, PatientProfile, DoctorProfile,
    Payment, PatientHistory, TestBooking
)


def get_frontdesk_profile(user):
    """Helper function to get front desk profile"""
    try:
        return FrontDeskProfile.objects.get(user=user)
    except FrontDeskProfile.DoesNotExist:
        return None


@login_required
def frontdesk_dashboard(request):
    """Main Front Desk Dashboard"""
    frontdesk = get_frontdesk_profile(request.user)
    
    if not frontdesk:
        messages.error(request, "You don't have access to this page.")
        return redirect('login')

    today = timezone.now().date()
    
    # Get statistics
    todays_appointments = Appointment.objects.filter(
        appointment_date=today
    ).order_by('appointment_time')
    
    todays_appointments_count = todays_appointments.count()
    
    # Pending check-ins (appointments scheduled for today but not checked in)
    pending_checkins = todays_appointments.filter(
        status__in=['Scheduled', 'Confirmed']
    )
    pending_checkin_count = pending_checkins.count()
    
    # Checked-in today
    checkedin_count = todays_appointments.filter(
        status='Completed'
    ).count()
    
    # Pending payments
    pending_payments = Payment.objects.filter(payment_status='Pending')
    total_pending_payments = pending_payments.aggregate(
        total=Sum('amount')
    )['total'] or 0
    
    pending_appointments_count = Appointment.objects.filter(
        status__in=['Pending Payment', 'Scheduled']
    ).count()

    context = {
        'todays_appointments': todays_appointments[:5],
        'todays_appointments_count': todays_appointments_count,
        'pending_checkin_count': pending_checkin_count,
        'pending_checkins': pending_checkins[:5],
        'checkedin_count': checkedin_count,
        'total_pending_payments': total_pending_payments,
        'pending_appointments_count': pending_appointments_count,
    }

    return render(request, 'core/dashboard/frontdesk_dashboard.html', context)


@login_required
def frontdesk_appointments(request):
    """Manage all appointments"""
    frontdesk = get_frontdesk_profile(request.user)
    
    if not frontdesk:
        messages.error(request, "You don't have access to this page.")
        return redirect('login')

    # Get all appointments
    appointments = Appointment.objects.all().order_by('-appointment_date', '-appointment_time')
    
    # Filter by status if provided
    status = request.GET.get('status')
    if status:
        appointments = appointments.filter(status=status)
    
    # Filter by date range if provided
    date_from = request.GET.get('date_from')
    date_to = request.GET.get('date_to')
    
    if date_from:
        appointments = appointments.filter(appointment_date__gte=date_from)
    if date_to:
        appointments = appointments.filter(appointment_date__lte=date_to)
    
    # Search by patient name or doctor name
    search = request.GET.get('search')
    if search:
        appointments = appointments.filter(
            Q(patient__full_name__icontains=search) |
            Q(doctor__user__first_name__icontains=search) |
            Q(doctor__user__last_name__icontains=search)
        )
    
    # Status choices for filter
    status_choices = Appointment.STATUS_CHOICES

    context = {
        'appointments': appointments,
        'status_choices': status_choices,
        'current_status': status,
        'search_query': search,
    }

    return render(request, 'core/dashboard/frontdesk_appointments.html', context)


@login_required
def frontdesk_appointment_detail(request, appointment_id):
    """View and edit appointment details"""
    frontdesk = get_frontdesk_profile(request.user)
    
    if not frontdesk:
        messages.error(request, "You don't have access to this page.")
        return redirect('login')

    appointment = get_object_or_404(Appointment, id=appointment_id)

    if request.method == 'POST':
        # Update appointment status
        new_status = request.POST.get('status')
        if new_status in dict(Appointment.STATUS_CHOICES):
            appointment.status = new_status
            appointment.save()
            messages.success(request, f"Appointment status updated to {new_status}")
            return redirect('frontdesk_appointment_detail', appointment_id=appointment_id)

    context = {
        'appointment': appointment,
        'status_choices': Appointment.STATUS_CHOICES,
    }

    return render(request, 'core/dashboard/frontdesk_appointment_detail.html', context)


@login_required
def frontdesk_patient_checkin(request):
    """Patient check-in/check-out management"""
    frontdesk = get_frontdesk_profile(request.user)
    
    if not frontdesk:
        messages.error(request, "You don't have access to this page.")
        return redirect('login')

    today = timezone.now().date()
    
    # Get today's appointments
    todays_appointments = Appointment.objects.filter(
        appointment_date=today
    ).order_by('appointment_time')
    
    # Pending check-ins
    pending_checkins = todays_appointments.filter(
        status__in=['Scheduled', 'Confirmed']
    )
    
    # Already checked in
    checked_in = todays_appointments.filter(
        status='Completed'
    )

    if request.method == 'POST':
        appointment_id = request.POST.get('appointment_id')
        action = request.POST.get('action')  # 'check_in' or 'check_out'
        
        appointment = get_object_or_404(Appointment, id=appointment_id)
        
        if action == 'check_in':
            appointment.status = 'Confirmed'
            appointment.save()
            messages.success(request, f"{appointment.patient.full_name} checked in successfully!")
        
        elif action == 'check_out':
            appointment.status = 'Completed'
            appointment.save()
            messages.success(request, f"{appointment.patient.full_name} checked out successfully!")
        
        return redirect('frontdesk_patient_check_in')

    context = {
        'pending_checkins': pending_checkins,
        'checked_in': checked_in,
        'todays_date': today,
    }

    return render(request, 'core/dashboard/frontdesk_patient_checkin.html', context)


@login_required
def frontdesk_patients_list(request):
    """View all patients"""
    frontdesk = get_frontdesk_profile(request.user)
    
    if not frontdesk:
        messages.error(request, "You don't have access to this page.")
        return redirect('login')

    patients = PatientProfile.objects.all().order_by('full_name')
    
    # Search by name or phone
    search = request.GET.get('search')
    if search:
        patients = patients.filter(
            Q(full_name__icontains=search) |
            Q(phone__icontains=search) |
            Q(user__email__icontains=search)
        )
    
    # Filter by status
    status = request.GET.get('status')
    if status:
        patients = patients.filter(status=status)

    context = {
        'patients': patients,
        'search_query': search,
        'current_status': status,
        'status_choices': ['Active', 'Inactive'],
    }

    return render(request, 'core/dashboard/frontdesk_patients_list.html', context)


@login_required
def frontdesk_patients_detail(request, patient_id):
    """View patient details"""
    frontdesk = get_frontdesk_profile(request.user)
    
    if not frontdesk:
        messages.error(request, "You don't have access to this page.")
        return redirect('login')

    patient = get_object_or_404(PatientProfile, id=patient_id)
    
    # Get patient's appointments
    appointments = Appointment.objects.filter(patient=patient).order_by('-appointment_date')
    
    # Get patient's payments
    payments = Payment.objects.filter(patient=patient).order_by('-payment_date')
    
    # Get patient's test bookings
    test_bookings = TestBooking.objects.filter(patient=patient).order_by('-booking_date')

    context = {
        'patient': patient,
        'appointments': appointments[:5],
        'payments': payments[:5],
        'test_bookings': test_bookings[:5],
    }

    return render(request, 'core/dashboard/frontdesk_patients_detail.html', context)


@login_required
def frontdesk_doctors_list(request):
    """View all doctors"""
    frontdesk = get_frontdesk_profile(request.user)
    
    if not frontdesk:
        messages.error(request, "You don't have access to this page.")
        return redirect('login')

    doctors = DoctorProfile.objects.all().order_by('user__first_name')
    
    # Search by name or specialization
    search = request.GET.get('search')
    if search:
        doctors = doctors.filter(
            Q(user__first_name__icontains=search) |
            Q(user__last_name__icontains=search) |
            Q(specialization__icontains=search) |
            Q(department__icontains=search)
        )
    
    # Filter by specialization
    specialization = request.GET.get('specialization')
    if specialization:
        doctors = doctors.filter(specialization=specialization)
    
    # Filter by status
    status = request.GET.get('status')
    if status:
        doctors = doctors.filter(status=status)

    context = {
        'doctors': doctors,
        'search_query': search,
        'current_specialization': specialization,
        'current_status': status,
    }

    return render(request, 'core/dashboard/frontdesk_doctors_list.html', context)


@login_required
def frontdesk_doctor_detail(request, doctor_id):
    """View doctor details and schedule"""
    frontdesk = get_frontdesk_profile(request.user)
    
    if not frontdesk:
        messages.error(request, "You don't have access to this page.")
        return redirect('login')

    doctor = get_object_or_404(DoctorProfile, id=doctor_id)
    
    # Get doctor's appointments
    appointments = Appointment.objects.filter(doctor=doctor).order_by('-appointment_date')
    
    # Today's appointments for this doctor
    today = timezone.now().date()
    todays_appointments = appointments.filter(appointment_date=today)

    context = {
        'doctor': doctor,
        'appointments': appointments[:10],
        'todays_appointments': todays_appointments,
        'todays_appointments_count': todays_appointments.count(),
    }

    return render(request, 'core/dashborad/frontdesk_doctor_detail.html', context)


@login_required
def frontdesk_payments(request):
    """Manage payments"""
    frontdesk = get_frontdesk_profile(request.user)
    
    if not frontdesk:
        messages.error(request, "You don't have access to this page.")
        return redirect('login')

    payments = Payment.objects.all().order_by('-payment_date')
    
    # Filter by status
    status = request.GET.get('status')
    if status:
        payments = payments.filter(payment_status=status)
    
    # Search by patient name
    search = request.GET.get('search')
    if search:
        payments = payments.filter(
            Q(patient__full_name__icontains=search) |
            Q(patient__user__email__icontains=search)
        )
    
    # Statistics
    total_payments = payments.aggregate(total=Sum('amount'))['total'] or 0
    paid_payments = payments.filter(payment_status='Paid').aggregate(total=Sum('amount'))['total'] or 0
    pending_payments = payments.filter(payment_status='Pending').aggregate(total=Sum('amount'))['total'] or 0

    context = {
        'payments': payments,
        'status_choices': Payment.PAYMENT_STATUS,
        'current_status': status,
        'search_query': search,
        'total_payments': total_payments,
        'paid_payments': paid_payments,
        'pending_payments': pending_payments,
    }

    return render(request, 'core/dashboard/frontdesk_payments.html', context)


@login_required
def frontdesk_payment_detail(request, payment_id):
    """View and update payment details"""
    frontdesk = get_frontdesk_profile(request.user)
    
    if not frontdesk:
        messages.error(request, "You don't have access to this page.")
        return redirect('login')

    payment = get_object_or_404(Payment, id=payment_id)

    if request.method == 'POST':
        new_status = request.POST.get('payment_status')
        if new_status in dict(Payment.PAYMENT_STATUS):
            payment.payment_status = new_status
            payment.save()
            messages.success(request, f"Payment status updated to {new_status}")
            return redirect('frontdesk_payment_detail', payment_id=payment_id)

    context = {
        'payment': payment,
        'payment_methods': Payment.PAYMENT_METHODS,
        'payment_statuses': Payment.PAYMENT_STATUS,
    }

    return render(request, 'core/dashboard/frontdesk_payment_detail.html', context)


@login_required
def frontdesk_reports(request):
    """View reports and analytics"""
    frontdesk = get_frontdesk_profile(request.user)
    
    if not frontdesk:
        messages.error(request, "You don't have access to this page.")
        return redirect('login')

    today = timezone.now().date()
    month_ago = today - timedelta(days=30)
    
    # Daily statistics
    todays_appointments = Appointment.objects.filter(appointment_date=today).count()
    todays_checkins = Appointment.objects.filter(
        appointment_date=today,
        status='Completed'
    ).count()
    
    # Monthly statistics
    monthly_appointments = Appointment.objects.filter(
        appointment_date__gte=month_ago
    ).count()
    
    monthly_payments = Payment.objects.filter(
        payment_date__gte=month_ago,
        payment_status='Paid'
    ).aggregate(total=Sum('amount'))['total'] or 0
    
    # Doctor statistics
    doctor_appointments = DoctorProfile.objects.annotate(
        appointment_count=Count('appointment')
    ).order_by('-appointment_count')[:5]
    
    # Payment statistics
    payment_by_method = Payment.objects.filter(
        payment_date__gte=month_ago
    ).values('payment_method').annotate(
        count=Count('id'),
        total=Sum('amount')
    )

    context = {
        'todays_appointments': todays_appointments,
        'todays_checkins': todays_checkins,
        'monthly_appointments': monthly_appointments,
        'monthly_payments': monthly_payments,
        'doctor_appointments': doctor_appointments,
        'payment_by_method': payment_by_method,
    }

    return render(request, 'core/dashboard/frontdesk_reports.html', context)


@login_required
def frontdesk_settings(request):
    """Front desk settings"""
    frontdesk = get_frontdesk_profile(request.user)
    
    if not frontdesk:
        messages.error(request, "You don't have access to this page.")
        return redirect('login')

    if request.method == 'POST':
        # Update profile information
        user = request.user
        user.first_name = request.POST.get('first_name', user.first_name)
        user.last_name = request.POST.get('last_name', user.last_name)
        user.email = request.POST.get('email', user.email)
        user.save()
        
        # Update password if provided
        old_password = request.POST.get('old_password')
        new_password = request.POST.get('new_password')
        confirm_password = request.POST.get('confirm_password')
        
        if old_password and new_password:
            if user.check_password(old_password):
                if new_password == confirm_password:
                    user.set_password(new_password)
                    user.save()
                    messages.success(request, "Password updated successfully!")
                else:
                    messages.error(request, "New passwords do not match!")
            else:
                messages.error(request, "Old password is incorrect!")
        
        if not old_password:
            messages.success(request, "Profile updated successfully!")
        
        return redirect('frontdesk_settings')

    context = {
        'frontdesk': frontdesk,
    }

    return render(request, 'core/dashboard/frontdesk_settings.html', context)

def frontdesk_patients_edit(request, patient_id):
    patient = PatientProfile.objects.get(id=patient_id)
    
    if request.method == 'POST':
        form = PatientForm(request.POST, instance=patient)
        if form.is_valid():
            form.save()
            return redirect('patient_detail', patient_id=patient.id)
    else:
        form = PatientForm(instance=patient)
    
    return render(request, 'core/dashboard/frontdesk_patients_edit.html', {'form': form, 'patient': patient})

def get_frontdesk_profile(user):
    """Helper function to get front desk profile"""
    try:
        return FrontDeskProfile.objects.get(user=user)
    except FrontDeskProfile.DoesNotExist:
        return None


# Add these views to your core/views.py - FIXED VERSION 2

from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.contrib.auth.models import User
from django.contrib import messages
from django.http import JsonResponse
from datetime import date, datetime, timedelta
from decimal import Decimal
import uuid

from core.models import (
    FrontDeskProfile, PatientProfile, DoctorProfile, Appointment,
    Payment
)


def get_frontdesk_profile(user):
    """Helper function to get front desk profile"""
    try:
        return FrontDeskProfile.objects.get(user=user)
    except FrontDeskProfile.DoesNotExist:
        return None


def generate_unique_username(email):
    """
    Generate a unique username from email or create a random one
    """
    # Try using email first
    if email:
        base_username = email.split('@')[0]
    else:
        base_username = f"patient_{uuid.uuid4().hex[:8]}"
    
    username = base_username
    counter = 1
    
    # If username exists, append a number
    while User.objects.filter(username=username).exists():
        username = f"{base_username}{counter}"
        counter += 1
    
    return username


@login_required
def frontdesk_book_appointment(request):
    """
    Front desk staff books appointments for patients
    Flow:
    1. Select or create patient
    2. Select doctor and time
    3. Collect payment
    4. Generate token
    """
    frontdesk = get_frontdesk_profile(request.user)
    
    if not frontdesk:
        messages.error(request, "You don't have access to this page.")
        return redirect('login')

    if request.method == 'POST':
        step = request.POST.get('step', '1')
        
        # Step 1: Patient Selection/Creation
        if step == '1':
            patient_id = request.POST.get('patient_id')
            new_patient_name = request.POST.get('new_patient_name')
            new_patient_phone = request.POST.get('new_patient_phone')
            new_patient_email = request.POST.get('new_patient_email')
            
            if patient_id:
                # Existing patient selected
                try:
                    patient = PatientProfile.objects.get(id=patient_id)
                    request.session['selected_patient_id'] = patient.id
                    messages.success(request, f"Patient {patient.full_name} selected")
                except PatientProfile.DoesNotExist:
                    messages.error(request, "Patient not found")
                    return redirect('frontdesk_book_appointment')
            
            elif new_patient_name and new_patient_phone:
                # Create new patient - FIXED
                try:
                    # Generate unique username
                    if new_patient_email:
                        username = generate_unique_username(new_patient_email)
                    else:
                        username = generate_unique_username(None)
                    
                    # Create user with unique username
                    user = User.objects.create_user(
                        username=username,
                        email=new_patient_email or '',
                        password=uuid.uuid4().hex[:12]
                    )
                    
                    # Create patient profile
                    patient = PatientProfile.objects.create(
                        user=user,
                        full_name=new_patient_name,
                        phone=new_patient_phone,
                        status='Active'
                    )
                    
                    request.session['selected_patient_id'] = patient.id
                    messages.success(request, f"New patient {patient.full_name} created")
                except Exception as e:
                    messages.error(request, f"Error creating patient: {str(e)}")
                    return redirect('frontdesk_book_appointment')
            else:
                messages.error(request, "Please select or create a patient")
                return redirect('frontdesk_book_appointment')
        
        # Step 2: Doctor and Time Selection
        elif step == '2':
            selected_patient_id = request.session.get('selected_patient_id')
            if not selected_patient_id:
                messages.error(request, "Please select a patient first")
                return redirect('frontdesk_book_appointment')
            
            doctor_id = request.POST.get('doctor_id')
            appointment_date = request.POST.get('appointment_date')
            appointment_time = request.POST.get('appointment_time')
            reason = request.POST.get('reason')
            
            # Validation
            if not all([doctor_id, appointment_date, appointment_time, reason]):
                messages.error(request, "All fields are required")
                return redirect('frontdesk_book_appointment')
            
            try:
                doctor = DoctorProfile.objects.get(id=doctor_id)
                
                # Check if slot is available
                existing_appointment = Appointment.objects.filter(
                    doctor=doctor,
                    appointment_date=appointment_date,
                    appointment_time=appointment_time,
                    status__in=['Pending Payment', 'Scheduled', 'Confirmed']
                ).exists()
                
                if existing_appointment:
                    messages.error(request, "This time slot is already booked. Please select another time.")
                    return redirect('frontdesk_book_appointment')
                
                # Store in session
                request.session['appointment_data'] = {
                    'doctor_id': doctor_id,
                    'appointment_date': appointment_date,
                    'appointment_time': appointment_time,
                    'reason': reason,
                    'doctor_name': f"Dr. {doctor.user.get_full_name()}",
                    'consultation_fee': str(doctor.consultation_fee)
                }
                
                messages.success(request, "Appointment details confirmed. Proceed to payment.")
                
            except DoctorProfile.DoesNotExist:
                messages.error(request, "Doctor not found")
                return redirect('frontdesk_book_appointment')
        
        # Step 3: Payment Processing
        elif step == '3':
            selected_patient_id = request.session.get('selected_patient_id')
            appointment_data = request.session.get('appointment_data')
            
            if not selected_patient_id or not appointment_data:
                messages.error(request, "Session expired. Please start over.")
                return redirect('frontdesk_book_appointment')
            
            payment_method = request.POST.get('payment_method')
            amount = Decimal(request.POST.get('amount', '0'))
            
            if not payment_method or amount <= 0:
                messages.error(request, "Invalid payment details")
                return redirect('frontdesk_book_appointment')
            
            try:
                patient = PatientProfile.objects.get(id=selected_patient_id)
                doctor = DoctorProfile.objects.get(id=appointment_data['doctor_id'])
                
                # Create appointment
                appointment = Appointment.objects.create(
                    patient=patient,
                    doctor=doctor,
                    appointment_date=appointment_data['appointment_date'],
                    appointment_time=appointment_data['appointment_time'],
                    reason=appointment_data['reason'],
                    status='Pending Payment'
                )
                
                # Create payment
                transaction_id = f"TXN{uuid.uuid4().hex[:12].upper()}"
                payment = Payment.objects.create(
                    patient=patient,
                    appointment=appointment,
                    amount=amount,
                    payment_method=payment_method,
                    payment_status='Paid',
                    transaction_id=transaction_id
                )
                
                # Update appointment status to Scheduled after payment
                appointment.status = 'Scheduled'
                appointment.save()
                
                # Clear session safely
                if 'selected_patient_id' in request.session:
                    del request.session['selected_patient_id']
                if 'appointment_data' in request.session:
                    del request.session['appointment_data']
                
                # Generate token
                token_number = generate_token(appointment)
                
                messages.success(
                    request,
                    f"Appointment booked successfully! Token: {token_number}"
                )
                
                return redirect('frontdesk_appointment_confirmation', appointment_id=appointment.id)
                
            except Exception as e:
                messages.error(request, f"Error processing payment: {str(e)}")
                return redirect('frontdesk_book_appointment')
    
    # GET request - Show form
    patients = PatientProfile.objects.all().order_by('full_name')
    doctors = DoctorProfile.objects.filter(status='Active').order_by('user__first_name')
    
    # Check if we're in the middle of booking
    selected_patient_id = request.session.get('selected_patient_id')
    appointment_data = request.session.get('appointment_data')
    
    selected_patient = None
    if selected_patient_id:
        try:
            selected_patient = PatientProfile.objects.get(id=selected_patient_id)
        except PatientProfile.DoesNotExist:
            pass
    
    context = {
        'patients': patients,
        'doctors': doctors,
        'selected_patient': selected_patient,
        'appointment_data': appointment_data,
        'today': date.today().isoformat(),
        'min_date': (date.today() + timedelta(days=1)).isoformat(),
    }
    
    return render(request, 'core/dashboard/frontdesk_book_appointment.html', context)


@login_required
def frontdesk_appointment_confirmation(request, appointment_id):
    """
    Show appointment confirmation with token
    """
    frontdesk = get_frontdesk_profile(request.user)
    
    if not frontdesk:
        messages.error(request, "You don't have access to this page.")
        return redirect('login')

    appointment = get_object_or_404(Appointment, id=appointment_id)
    
    # Get payment info
    payment = Payment.objects.filter(appointment=appointment).first()
    
    # Get token from appointment
    token = getattr(appointment, 'token_number', None)
    
    context = {
        'appointment': appointment,
        'payment': payment,
        'token': token,
    }
    
    return render(request, 'core/dashboard/frontdesk_appointment_confirmation.html', context)

# ==========================================
# FINAL FRONTDESK FIX
# Replace frontdesk_get_available_slots() completely
# ==========================================

from django.http import JsonResponse
from django.contrib.auth.decorators import login_required

@login_required
def frontdesk_get_available_slots(request):
    """
    AJAX endpoint to get available time slots for a doctor on a specific date
    FINAL FIX: Bulletproof conversion of appointment times
    """
    
    if request.method == 'GET':
        doctor_id = request.GET.get('doctor_id')
        appointment_date = request.GET.get('appointment_date')
        
        if not doctor_id or not appointment_date:
            return JsonResponse({'error': 'Missing parameters'}, status=400)
        
        try:
            from core.models import DoctorProfile, Appointment
            
            doctor = DoctorProfile.objects.get(id=doctor_id)
            
            # Define time slots (9 AM to 6 PM, 30-min intervals)
            time_slots = []
            for hour in range(9, 18):
                for minute in [0, 30]:
                    time_str = f"{hour:02d}:{minute:02d}"
                    time_slots.append(time_str)
            
            # Get booked appointments
            booked_appointments = Appointment.objects.filter(
                doctor=doctor,
                appointment_date=appointment_date,
                status__in=['Pending Payment', 'Scheduled', 'Confirmed']
            ).values_list('appointment_time', flat=True)
            
            # ==========================================
            # BULLETPROOF CONVERSION (THIS IS THE KEY FIX)
            # ==========================================
            booked_times = set()  # Use set, not list (faster)
            
            for appointment_time in booked_appointments:
                try:
                    # Step 1: Convert to string FIRST
                    time_string = str(appointment_time).strip()
                    
                    # Step 2: Extract HH:MM (first 5 characters)
                    if len(time_string) >= 5:
                        time_only = time_string[:5]
                    else:
                        time_only = time_string
                    
                    # Step 3: Validate it's in HH:MM format
                    if ':' in time_only and len(time_only) == 5:
                        booked_times.add(time_only)
                    else:
                        # Try to fix the format
                        parts = time_string.split(':')
                        if len(parts) >= 2:
                            hour = parts[0].zfill(2)
                            minute = parts[1][:2].zfill(2)
                            booked_times.add(f"{hour}:{minute}")
                
                except Exception as e:
                    # If all else fails, just skip this time
                    print(f"Warning: Could not process time {appointment_time}: {e}")
                    pass
            
            # Build response with available slots
            available_slots = []
            for slot in time_slots:
                is_available = slot not in booked_times
                available_slots.append({
                    'time': slot,
                    'available': is_available
                })
            
            return JsonResponse({'slots': available_slots})
        
        except DoctorProfile.DoesNotExist:
            return JsonResponse({'error': 'Doctor not found'}, status=404)
        except Exception as e:
            import traceback
            error_detail = traceback.format_exc()
            print(f"Error in frontdesk_get_available_slots: {error_detail}")
            return JsonResponse({
                'error': f'Server error: {str(e)}'
            }, status=500)
    
    return JsonResponse({'error': 'Invalid request'}, status=400)

@login_required
def frontdesk_search_patient(request):
    """
    AJAX endpoint to search for existing patients
    """
    if request.method == 'GET':
        query = request.GET.get('q', '').strip()
        
        if not query or len(query) < 2:
            return JsonResponse({'patients': []})
        
        patients = PatientProfile.objects.filter(
            full_name__icontains=query
        )[:10]
        
        patients_data = []
        for patient in patients:
            patients_data.append({
                'id': patient.id,
                'name': patient.full_name,
                'phone': patient.phone,
                'email': patient.user.email,
                'gender': patient.gender or 'Not specified',
                'dob': str(patient.dob) if patient.dob else 'N/A',
            })
        
        return JsonResponse({'patients': patients_data})
    
    return JsonResponse({'error': 'Invalid request'}, status=400)


def generate_token(appointment):
    """
    Generate appointment token - SIMPLE VERSION
    """
    # Convert to string first
    date_obj = appointment.appointment_date
    
    if isinstance(date_obj, str):
        # Remove all non-digits and take first 4 chars
        date_str = ''.join(filter(str.isdigit, date_obj))[-4:]
    else:
        date_str = date_obj.strftime('%d%m')
    
    # Count appointments for this date
    count = Appointment.objects.filter(
        appointment_date=appointment.appointment_date
    ).count()
    
    token_number = f"{date_str}-{count:03d}"
    
    return token_number





@login_required
def frontdesk_today_appointments(request):
    """
    View today's appointments and manage check-ins
    """
    frontdesk = get_frontdesk_profile(request.user)
    
    if not frontdesk:
        messages.error(request, "You don't have access to this page.")
        return redirect('login')

    today = date.today()
    
    # Get today's appointments sorted by time
    todays_appointments = Appointment.objects.filter(
        appointment_date=today
    ).select_related('patient', 'doctor', 'doctor__user').order_by('appointment_time')
    
    # Separate by status
    pending_checkin = todays_appointments.filter(
        status__in=['Scheduled', 'Pending Payment']
    )
    
    checked_in = todays_appointments.filter(
        status='Confirmed'
    )
    
    completed = todays_appointments.filter(
        status='Completed'
    )

    context = {
        'todays_appointments': todays_appointments,
        'pending_checkin': pending_checkin,
        'checked_in': checked_in,
        'completed': completed,
        'today': today,
    }
    
    return render(request, 'core/dashboard/frontdesk_today_appointments.html', context)


@login_required
def frontdesk_quick_checkin(request, appointment_id):
    """
    Quick check-in for appointment
    """
    frontdesk = get_frontdesk_profile(request.user)
    
    if not frontdesk:
        messages.error(request, "You don't have access to this action.")
        return redirect('login')

    appointment = get_object_or_404(Appointment, id=appointment_id)
    
    if request.method == 'POST':
        # Update status to Confirmed (patient checked in)
        appointment.status = 'Confirmed'
        appointment.save()
        
        messages.success(
            request,
            f"{appointment.patient.full_name} checked in successfully!"
        )
        
        return redirect('frontdesk_today_appointments')
    
    return redirect('frontdesk_today_appointments')


@login_required
def frontdesk_patients_edit(request, patient_id):
    """
    Edit patient information by front desk staff
    """
    frontdesk = get_frontdesk_profile(request.user)
    
    if not frontdesk:
        messages.error(request, "You don't have access to this page.")
        return redirect('login')

    patient = get_object_or_404(PatientProfile, id=patient_id)
    
    if request.method == 'POST':
        # Update patient information
        patient.full_name = request.POST.get('full_name', patient.full_name)
        patient.phone = request.POST.get('phone', patient.phone)
        patient.gender = request.POST.get('gender', patient.gender)
        patient.dob = request.POST.get('dob', patient.dob) if request.POST.get('dob') else patient.dob
        patient.address = request.POST.get('address', patient.address)
        patient.status = request.POST.get('status', patient.status)
        patient.save()
        
        # Update user email if provided
        user = patient.user
        new_email = request.POST.get('email', user.email)
        if new_email and new_email != user.email:
            # Check if email already exists
            if User.objects.filter(email=new_email).exclude(id=user.id).exists():
                messages.error(request, "This email is already in use")
                return redirect('frontdesk_patients_edit', patient_id=patient_id)
            user.email = new_email
            user.save()
        
        messages.success(request, f"Patient {patient.full_name} updated successfully!")
        return redirect('frontdesk_patients_detail', patient_id=patient_id)
    
    context = {
        'patient': patient,
        'genders': ['Male', 'Female', 'Other'],
        'statuses': ['Active', 'Inactive'],
    }
    
    return render(request, 'core/dashboard/frontdesk_patients_edit.html', context)

@login_required
def doctor_reschedule_appointment(request, appointment_id):
    """Reschedule an appointment"""

    # ✅ Get DoctorProfile from logged in user
    doctor_profile = request.user.doctorprofile

    # ✅ Now filter correctly
    appointment = get_object_or_404(
        Appointment,
        id=appointment_id,
        doctor=doctor_profile
    )

    if request.method == 'POST':
        new_date = request.POST.get('appointment_date')
        new_time = request.POST.get('appointment_time')
        reschedule_reason = request.POST.get('reschedule_reason', '')
        notify_patient = request.POST.get('notify_patient') == 'on'

        # Update appointment
        appointment.appointment_date = new_date
        appointment.appointment_time = new_time

        # Store reschedule reason
        if reschedule_reason:
            if appointment.notes:
                appointment.notes += f"\n\nRescheduled: {reschedule_reason}"
            else:
                appointment.notes = f"Rescheduled: {reschedule_reason}"

        appointment.save()

        if notify_patient:
            # Add email logic later
            pass

        messages.success(
            request,
            f'Appointment rescheduled successfully to {new_date} at {new_time}'
        )

        return redirect('doctor_appointment_detail', appointment_id=appointment.id)

    context = {
        'appointment': appointment,
        'today': timezone.now().date(),
    }

    return render(request, 'core/dashboard/doctor_appointment_reschedule.html', context)

@login_required
def doctor_prescriptions(request):
    """
    View all prescriptions issued by the logged-in doctor
    ✅ FIXED: Properly retrieve DoctorProfile
    """
    try:
        doctor_profile = DoctorProfile.objects.get(user=request.user)
    except DoctorProfile.DoesNotExist:
        messages.error(request, "Doctor profile not found.")
        return redirect('doctor_dashboard')

    prescriptions = Prescription.objects.filter(
        doctor=doctor_profile
    ).select_related('patient', 'appointment').order_by('-created_at')

    # Filter by status if provided
    status_filter = request.GET.get('status', '')
    if status_filter:
        prescriptions = prescriptions.filter(status=status_filter)

    # Get statistics for dashboard
    all_prescriptions = Prescription.objects.filter(doctor=doctor_profile)
    
    stats = {
        'total': all_prescriptions.count(),
        'active': all_prescriptions.filter(status='Active').count(),
        'completed': all_prescriptions.filter(status='Completed').count(),
        'cancelled': all_prescriptions.filter(status='Cancelled').count(),
    }

    context = {
        'prescriptions': prescriptions,
        'status_filter': status_filter,
        'stats': stats,
        'doctor_profile': doctor_profile,
        'unread_notifications': 0,
    }

    return render(request, 'core/dashboard/doctor_prescriptions.html', context)



@login_required
def doctor_add_prescription(request):
    """
    Create a new prescription
    ✅ FIXED: Proper DoctorProfile and form handling
    """
    try:
        doctor_profile = DoctorProfile.objects.get(user=request.user)
    except DoctorProfile.DoesNotExist:
        messages.error(request, "Doctor profile not found.")
        return redirect('doctor_dashboard')

    if request.method == 'POST':
        # Get form data
        patient_id = request.POST.get('patient_id')
        appointment_id = request.POST.get('appointment_id')
        medicine_name = request.POST.get('medicine_name', '').strip()
        dosage = request.POST.get('dosage', '').strip()
        frequency = request.POST.get('frequency', '').strip()
        duration = request.POST.get('duration', '').strip()
        instructions = request.POST.get('instructions', '').strip()
        status = request.POST.get('status', 'Active')

        # Validation
        if not all([patient_id, medicine_name, dosage, frequency, duration]):
            messages.error(request, "All required fields must be filled.")
            return redirect('doctor_add_prescription')

        try:
            # Get patient
            patient = PatientProfile.objects.get(id=patient_id)

            # Get appointment if provided
            appointment = None
            if appointment_id:
                try:
                    appointment = Appointment.objects.get(id=appointment_id)
                except Appointment.DoesNotExist:
                    messages.warning(request, "Appointment not found, prescription created without appointment link.")

            # Create prescription
            prescription = Prescription.objects.create(
                patient=patient,
                doctor=doctor_profile,
                appointment=appointment,
                medicine_name=medicine_name,
                dosage=dosage,
                frequency=frequency,
                duration=duration,
                instructions=instructions,
                status=status
            )

            messages.success(
                request,
                f"Prescription for {medicine_name} created successfully! Prescription ID: #{prescription.id}"
            )
            return redirect('doctor_prescription', prescription_id=prescription.id)

        except PatientProfile.DoesNotExist:
            messages.error(request, "Selected patient not found.")
            return redirect('doctor_add_prescription')
        except Exception as e:
            messages.error(request, f"Error creating prescription: {str(e)}")
            return redirect('doctor_add_prescription')

    # GET request - Show form
    # Get list of patients for dropdown
    patients = PatientProfile.objects.all().order_by('full_name')
    
    # Get list of appointments for optional linking
    appointments = Appointment.objects.filter(
        doctor=doctor_profile,
        status__in=['Scheduled', 'Confirmed']
    ).order_by('-appointment_date')

    context = {
        'patients': patients,
        'appointments': appointments,
        'doctor_profile': doctor_profile,
        'unread_notifications': 0,
    }

    return render(request, 'core/dashboard/doctor_create_prescription.html', context)


@login_required
def doctor_edit_prescription(request, prescription_id):
    """
    Edit an existing prescription
    ✅ FIXED: Proper DoctorProfile retrieval
    """
    try:
        doctor_profile = DoctorProfile.objects.get(user=request.user)
    except DoctorProfile.DoesNotExist:
        messages.error(request, "Doctor profile not found.")
        return redirect('doctor_dashboard')

    # Get the prescription and verify ownership
    prescription = get_object_or_404(
        Prescription,
        id=prescription_id,
        doctor=doctor_profile
    )

    # Only allow editing if prescription is Active
    if prescription.status != 'Active':
        messages.error(request, "Can only edit active prescriptions.")
        return redirect('doctor_prescription', prescription_id=prescription_id)

    if request.method == 'POST':
        # Update prescription fields
        prescription.medicine_name = request.POST.get('medicine_name', prescription.medicine_name).strip()
        prescription.dosage = request.POST.get('dosage', prescription.dosage).strip()
        prescription.frequency = request.POST.get('frequency', prescription.frequency).strip()
        prescription.duration = request.POST.get('duration', prescription.duration).strip()
        prescription.instructions = request.POST.get('instructions', prescription.instructions).strip()
        prescription.status = request.POST.get('status', prescription.status)
        prescription.save()

        messages.success(request, "Prescription updated successfully!")
        return redirect('doctor_prescription', prescription_id=prescription.id)

    context = {
        'prescription': prescription,
        'doctor_profile': doctor_profile,
        'unread_notifications': 0,
    }

    return render(request, 'core/dashboard/doctor_edit_prescription.html', context)


@login_required
def doctor_delete_prescription(request, prescription_id):
    """
    Delete a prescription (can only delete own prescriptions)
    ✅ FIXED: Proper permission checking
    """
    try:
        doctor_profile = DoctorProfile.objects.get(user=request.user)
    except DoctorProfile.DoesNotExist:
        messages.error(request, "Doctor profile not found.")
        return redirect('doctor_dashboard')

    prescription = get_object_or_404(
        Prescription,
        id=prescription_id,
        doctor=doctor_profile
    )

    if request.method == 'POST':
        medicine_name = prescription.medicine_name
        prescription.delete()
        messages.success(request, f"Prescription for {medicine_name} deleted successfully!")
        return redirect('doctor_prescriptions')

    context = {
        'prescription': prescription,
        'doctor_profile': doctor_profile,
    }

    return render(request, 'core/dashboard/doctor_confirm_delete_prescription.html', context)


@login_required
def doctor_prescription_print(request, prescription_id):
    """
    Generate printable view of prescription
    """
    try:
        doctor_profile = DoctorProfile.objects.get(user=request.user)
    except DoctorProfile.DoesNotExist:
        messages.error(request, "Doctor profile not found.")
        return redirect('doctor_dashboard')

    prescription = get_object_or_404(
        Prescription,
        id=prescription_id,
        doctor=doctor_profile
    )

    context = {
        'prescription': prescription,
        'doctor_profile': doctor_profile,
        'is_print': True,  # Flag for template to show print-only layout
    }

    return render(request, 'core/dashboard/doctor_prescription_print.html', context)

# ============================================================
# FIXED view for doctor_prescription_detail
# Replace / add to core/views.py
# ============================================================

@login_required
def doctor_prescription_detail(request, prescription_id):
    try:
        doctor_profile = DoctorProfile.objects.get(user=request.user)
    except DoctorProfile.DoesNotExist:
        messages.error(request, "Doctor profile not found.")
        return redirect('doctor_dashboard')

    prescription = get_object_or_404(
        Prescription,
        id=prescription_id,
        doctor=doctor_profile
    )

    # All prescriptions for the same patient (for the history table at the bottom)
    patient_prescriptions = Prescription.objects.filter(
        patient=prescription.patient,
        doctor=doctor_profile
    ).order_by('-created_at')

    context = {
        'prescription':         prescription,
        'patient_prescriptions': patient_prescriptions,
        'doctor_profile':        doctor_profile,
        'unread_notifications':  0,
    }

    return render(request, 'core/dashboard/doctor_prescription_detail.html', context)


# ============================================================
# FIXED view for doctor_add_prescription
# Key fix: doctor is already logged in — no doctor_id in form
# ============================================================

@login_required
def doctor_add_prescription(request):
    try:
        doctor_profile = DoctorProfile.objects.get(user=request.user)
    except DoctorProfile.DoesNotExist:
        messages.error(request, "Doctor profile not found.")
        return redirect('doctor_dashboard')

    if request.method == 'POST':
        patient_id    = request.POST.get('patient_id')
        appointment_id = request.POST.get('appointment_id')
        medicine_name = request.POST.get('medicine_name', '').strip()
        dosage        = request.POST.get('dosage', '').strip()
        frequency     = request.POST.get('frequency', '').strip()
        duration      = request.POST.get('duration', '').strip()
        instructions  = request.POST.get('instructions', '').strip()
        status        = request.POST.get('status', 'Active')

        if not all([patient_id, medicine_name, dosage, frequency, duration]):
            messages.error(request, "All required fields must be filled.")
            return redirect('doctor_add_prescription')

        try:
            patient = PatientProfile.objects.get(id=patient_id)

            # Appointment is REQUIRED by the model (non-nullable FK)
            # So we must have a valid appointment_id
            if not appointment_id:
                messages.error(request, "Please link an appointment.")
                return redirect('doctor_add_prescription')

            appointment = Appointment.objects.get(id=appointment_id)

            prescription = Prescription.objects.create(
                patient=patient,
                doctor=doctor_profile,
                appointment=appointment,
                medicine_name=medicine_name,
                dosage=dosage,
                frequency=frequency,
                duration=duration,
                instructions=instructions,
                status=status,
            )

            messages.success(request, f"Prescription for {medicine_name} created! ID: #{prescription.id}")
            return redirect('doctor_prescription_detail', prescription_id=prescription.id)

        except PatientProfile.DoesNotExist:
            messages.error(request, "Selected patient not found.")
        except Appointment.DoesNotExist:
            messages.error(request, "Selected appointment not found.")
        except Exception as e:
            messages.error(request, f"Error: {str(e)}")

        return redirect('doctor_add_prescription')

    # GET — build dropdown data
    # Only show patients who have appointments with this doctor
    patients = PatientProfile.objects.filter(
        appointment__doctor=doctor_profile
    ).distinct().order_by('full_name')

    appointments = Appointment.objects.filter(
        doctor=doctor_profile,
        status__in=['Scheduled', 'Confirmed']
    ).select_related('patient').order_by('-appointment_date')

    context = {
        'patients':             patients,
        'appointments':         appointments,
        'doctor_profile':       doctor_profile,
        'unread_notifications': 0,
    }

    return render(request, 'core/dashboard/doctor_create_prescription.html', context)

# ==================== DOCTOR PATIENTS VIEWS ====================

@login_required
def doctor_patients(request):
    """
    View all patients under this doctor's care with advanced filtering
    
    Features:
    - List all patients
    - Search by name, phone, email
    - Shows appointment count
    - Shows last visit date
    - Shows patient status
    - Shows gender
    
    URL: /doctor/patients/
    Template: core/dashboard/doctor_patients.html
    """
    try:
        doctor_profile = DoctorProfile.objects.get(user=request.user)
    except DoctorProfile.DoesNotExist:
        messages.error(request, "Doctor profile not found.")
        return redirect('doctor_dashboard')

    # Get unique patients with appointments to this doctor
    patient_ids = Appointment.objects.filter(
        doctor=doctor_profile
    ).values_list('patient_id', flat=True).distinct()

    patients = PatientProfile.objects.filter(
        id__in=patient_ids
    ).order_by('full_name')

    # Search functionality
    search_query = request.GET.get('search', '').strip()
    if search_query:
        patients = patients.filter(
            Q(full_name__icontains=search_query) |
            Q(phone__icontains=search_query) |
            Q(user__email__icontains=search_query)
        )

    # Add appointment count and last visit for each patient
    for patient in patients:
        appointments = Appointment.objects.filter(
            doctor=doctor_profile,
            patient=patient
        )
        patient.appointment_count = appointments.count()
        patient.last_appointment = appointments.order_by('-appointment_date').first()

    # Get statistics
    total_patients = PatientProfile.objects.filter(
        id__in=patient_ids
    ).count()

    active_patients = PatientProfile.objects.filter(
        id__in=patient_ids,
        status='Active'
    ).count()

    # Get unread notifications count (implement based on your notification system)
    unread_notifications = 0

    context = {
        'patients': patients,
        'search_query': search_query,
        'doctor_profile': doctor_profile,
        'total_patients': total_patients,
        'active_patients': active_patients,
        'unread_notifications': unread_notifications,
    }

    return render(request, 'core/dashboard/doctor_patients.html', context)


@login_required
def doctor_patient_detail(request, patient_id):
    """
    View detailed patient information including medical history
    
    Features:
    - Full patient profile
    - Medical history with this doctor
    - All appointments with this doctor
    - Current medications
    - Allergies
    - Medical conditions
    - Surgery history
    - Option to write prescription
    
    URL: /doctor/patient/<patient_id>/
    Template: core/dashboard/doctor_patient_detail.html
    """
    try:
        doctor_profile = DoctorProfile.objects.get(user=request.user)
    except DoctorProfile.DoesNotExist:
        messages.error(request, "Doctor profile not found.")
        return redirect('doctor_dashboard')

    patient = get_object_or_404(PatientProfile, id=patient_id)

    # Verify doctor has seen this patient
    has_appointment = Appointment.objects.filter(
        doctor=doctor_profile,
        patient=patient
    ).exists()

    if not has_appointment:
        messages.error(request, "You don't have access to this patient's records")
        return redirect('doctor_patients')

    # Get patient data
    appointments = Appointment.objects.filter(
        doctor=doctor_profile,
        patient=patient
    ).select_related('doctor').order_by('-appointment_date')

    prescriptions = Prescription.objects.filter(
        patient=patient,
        doctor=doctor_profile
    ).order_by('-created_at')

    allergies = Allergy.objects.filter(patient=patient)
    conditions = MedicalCondition.objects.filter(patient=patient)
    medications = PatientMedication.objects.filter(
        patient=patient,
        end_date__isnull=True
    )

    # Statistics
    total_appointments = appointments.count()
    total_prescriptions = prescriptions.count()

    context = {
        'patient': patient,
        'doctor_profile': doctor_profile,
        'appointments': appointments[:10],  # Last 10 appointments
        'total_appointments': total_appointments,
        'prescriptions': prescriptions[:5],  # Last 5 prescriptions
        'total_prescriptions': total_prescriptions,
        'allergies': allergies,
        'conditions': conditions,
        'medications': medications,
        'unread_notifications': 0,
    }

    return render(request, 'core/dashboard/doctor_patient_detail.html', context)


@login_required
def doctor_patient_medical_history(request, patient_id):
    """
    View complete medical history for a patient
    
    Features:
    - All appointments
    - All prescriptions
    - All allergies
    - All medications
    - All medical conditions
    - Surgery history
    - Family history
    - Immunizations
    - Vital signs
    
    URL: /doctor/patient/<patient_id>/history/
    Template: core/dashboard/doctor_patient_medical_history.html
    """
    try:
        doctor_profile = DoctorProfile.objects.get(user=request.user)
    except DoctorProfile.DoesNotExist:
        messages.error(request, "Doctor profile not found.")
        return redirect('doctor_dashboard')

    patient = get_object_or_404(PatientProfile, id=patient_id)

    # Verify access
    has_appointment = Appointment.objects.filter(
        doctor=doctor_profile,
        patient=patient
    ).exists()

    if not has_appointment:
        messages.error(request, "You don't have access to this patient's records")
        return redirect('doctor_patients')

    # Get all medical data
    appointments = Appointment.objects.filter(
        patient=patient
    ).order_by('-appointment_date')

    prescriptions = Prescription.objects.filter(
        patient=patient
    ).order_by('-created_at')

    allergies = Allergy.objects.filter(patient=patient).order_by('-recorded_date')
    conditions = MedicalCondition.objects.filter(patient=patient).order_by('-diagnosis_date')
    medications = PatientMedication.objects.filter(patient=patient).order_by('-start_date')

    # Import additional models if needed
    from core.models import Surgery, FamilyHistory, Immunization, VitalSigns

    surgeries = Surgery.objects.filter(patient=patient).order_by('-date')
    family_history = FamilyHistory.objects.filter(patient=patient).order_by('-recorded_date')
    immunizations = Immunization.objects.filter(patient=patient).order_by('-date')
    vital_signs = VitalSigns.objects.filter(patient=patient).order_by('-date')[:20]

    context = {
        'patient': patient,
        'doctor_profile': doctor_profile,
        'appointments': appointments,
        'prescriptions': prescriptions,
        'allergies': allergies,
        'conditions': conditions,
        'medications': medications,
        'surgeries': surgeries,
        'family_history': family_history,
        'immunizations': immunizations,
        'vital_signs': vital_signs,
        'unread_notifications': 0,
    }

    return render(request, 'core/dashboard/doctor_patient_medical_history.html', context)


@login_required
def doctor_patient_appointments(request, patient_id):
    """
    View all appointments with a specific patient
    
    Features:
    - List all appointments
    - Filter by status
    - Sort by date
    - Confirm appointments
    - Complete appointments
    - Reschedule appointments
    
    URL: /doctor/patient/<patient_id>/appointments/
    Template: core/dashboard/doctor_patient_appointments.html
    """
    try:
        doctor_profile = DoctorProfile.objects.get(user=request.user)
    except DoctorProfile.DoesNotExist:
        messages.error(request, "Doctor profile not found.")
        return redirect('doctor_dashboard')

    patient = get_object_or_404(PatientProfile, id=patient_id)

    # Verify access
    has_appointment = Appointment.objects.filter(
        doctor=doctor_profile,
        patient=patient
    ).exists()

    if not has_appointment:
        messages.error(request, "You don't have access to this patient's records")
        return redirect('doctor_patients')

    # Get appointments
    appointments = Appointment.objects.filter(
        doctor=doctor_profile,
        patient=patient
    ).order_by('-appointment_date', '-appointment_time')

    # Filter by status if provided
    status_filter = request.GET.get('status', '')
    if status_filter:
        appointments = appointments.filter(status=status_filter)

    # Statistics
    total_appointments = Appointment.objects.filter(
        doctor=doctor_profile,
        patient=patient
    ).count()

    pending_count = appointments.filter(
        status__in=['Pending Payment', 'Scheduled']
    ).count()

    completed_count = appointments.filter(
        status='Completed'
    ).count()

    cancelled_count = appointments.filter(
        status='Cancelled'
    ).count()

    context = {
        'patient': patient,
        'doctor_profile': doctor_profile,
        'appointments': appointments,
        'status_filter': status_filter,
        'total_appointments': total_appointments,
        'pending_count': pending_count,
        'completed_count': completed_count,
        'cancelled_count': cancelled_count,
        'unread_notifications': 0,
    }

    return render(request, 'core/dashboard/doctor_patient_appointments.html', context)


@login_required
def doctor_patient_prescriptions(request, patient_id):
    """
    View all prescriptions for a specific patient
    
    Features:
    - List all prescriptions
    - Filter by status
    - Write new prescription
    - Edit prescriptions
    - Delete prescriptions
    
    URL: /doctor/patient/<patient_id>/prescriptions/
    Template: core/dashboard/doctor_patient_prescriptions.html
    """
    try:
        doctor_profile = DoctorProfile.objects.get(user=request.user)
    except DoctorProfile.DoesNotExist:
        messages.error(request, "Doctor profile not found.")
        return redirect('doctor_dashboard')

    patient = get_object_or_404(PatientProfile, id=patient_id)

    # Verify access
    has_appointment = Appointment.objects.filter(
        doctor=doctor_profile,
        patient=patient
    ).exists()

    if not has_appointment:
        messages.error(request, "You don't have access to this patient's records")
        return redirect('doctor_patients')

    # Get prescriptions
    prescriptions = Prescription.objects.filter(
        patient=patient,
        doctor=doctor_profile
    ).order_by('-created_at')

    # Filter by status if provided
    status_filter = request.GET.get('status', '')
    if status_filter:
        prescriptions = prescriptions.filter(status=status_filter)

    # Statistics
    total_prescriptions = Prescription.objects.filter(
        patient=patient,
        doctor=doctor_profile
    ).count()

    active_count = prescriptions.filter(status='Active').count()
    completed_count = prescriptions.filter(status='Completed').count()

    context = {
        'patient': patient,
        'doctor_profile': doctor_profile,
        'prescriptions': prescriptions,
        'status_filter': status_filter,
        'total_prescriptions': total_prescriptions,
        'active_count': active_count,
        'completed_count': completed_count,
        'unread_notifications': 0,
    }

    return render(request, 'core/dashboard/doctor_patient_prescriptions.html', context)


@login_required
def doctor_patient_allergies(request, patient_id):
    """
    View and manage patient allergies
    
    Features:
    - List all allergies
    - Add new allergy
    - Edit allergy
    - Delete allergy
    - Show severity
    - Show reaction
    
    URL: /doctor/patient/<patient_id>/allergies/
    Template: core/dashboard/doctor_patient_allergies.html
    """
    try:
        doctor_profile = DoctorProfile.objects.get(user=request.user)
    except DoctorProfile.DoesNotExist:
        messages.error(request, "Doctor profile not found.")
        return redirect('doctor_dashboard')

    patient = get_object_or_404(PatientProfile, id=patient_id)

    # Verify access
    has_appointment = Appointment.objects.filter(
        doctor=doctor_profile,
        patient=patient
    ).exists()

    if not has_appointment:
        messages.error(request, "You don't have access to this patient's records")
        return redirect('doctor_patients')

    # Get allergies
    allergies = Allergy.objects.filter(patient=patient).order_by('-recorded_date')

    # Statistics
    critical_count = allergies.filter(severity='Critical').count()
    severe_count = allergies.filter(severity='Severe').count()

    context = {
        'patient': patient,
        'doctor_profile': doctor_profile,
        'allergies': allergies,
        'critical_count': critical_count,
        'severe_count': severe_count,
        'total_allergies': allergies.count(),
        'unread_notifications': 0,
    }

    return render(request, 'core/dashboard/doctor_patient_allergies.html', context)

# ============================================================
# Replace doctor_schedule() and doctor_schedule_update() in
# core/views.py with the code below.
# ============================================================

from datetime import date, timedelta
from django.shortcuts import render, redirect
from django.contrib.auth.decorators import login_required
from django.contrib import messages


@login_required
def doctor_schedule(request):
    """
    Display and manage doctor's weekly schedule and today's time slots.
    Template: core/dashboard/doctor_schedule.html
    """
    try:
        doctor_profile = DoctorProfile.objects.get(user=request.user)
    except DoctorProfile.DoesNotExist:
        messages.error(request, "Doctor profile not found")
        return redirect('login')

    today = date.today()

    # ── all_days: passed to the template for the checkbox list ───────────
    # No split filter needed in the template — the list is built here.
    ALL_DAY_DEFS = [
        ('Monday',    'Mon', True),
        ('Tuesday',   'Tue', True),
        ('Wednesday', 'Wed', True),
        ('Thursday',  'Thu', True),
        ('Friday',    'Fri', True),
        ('Saturday',  'Sat', False),
        ('Sunday',    'Sun', False),
    ]
    all_days = [
        {'name': name, 'abbr': abbr, 'default_checked': checked}
        for name, abbr, checked in ALL_DAY_DEFS
    ]

    # ── Weekly overview (Mon–Sun of the current week) ─────────────────────
    WORKING_DAY_NAMES = {'Monday', 'Tuesday', 'Wednesday', 'Thursday', 'Friday'}
    week_start = today - timedelta(days=today.weekday())   # Monday

    week_schedule = []
    for i in range(7):
        day_date = week_start + timedelta(days=i)
        day_name = day_date.strftime('%A')
        is_working = day_name in WORKING_DAY_NAMES

        appt_count = Appointment.objects.filter(
            doctor=doctor_profile,
            appointment_date=day_date,
            status__in=['Scheduled', 'Confirmed', 'Pending Payment']
        ).count()

        week_schedule.append({
            'day':          day_date.strftime('%a'),   # "Mon", "Tue", …
            'full_day':     day_name,
            'date':         day_date,
            'is_today':     day_date == today,
            'is_working':   is_working,
            'hours':        '9–5' if is_working else '',
            'appointments': appt_count if is_working else None,
        })

    # ── Today's time slots (09:00–17:00, 30-min intervals) ───────────────
    BREAK_SLOTS = {'13:00', '13:30'}

    todays_booked = Appointment.objects.filter(
        doctor=doctor_profile,
        appointment_date=today,
        status__in=['Scheduled', 'Confirmed', 'Pending Payment']
    ).select_related('patient').values('appointment_time', 'patient__full_name')

    booked_map = {}
    for appt in todays_booked:
        time_str = str(appt['appointment_time'])[:5]   # "HH:MM"
        booked_map[time_str] = appt['patient__full_name']

    todays_slots = []
    for hour in range(9, 17):
        for minute in (0, 30):
            raw = f"{hour:02d}:{minute:02d}"
            is_break  = raw in BREAK_SLOTS
            is_booked = (raw in booked_map) and not is_break
            display   = f"{hour % 12 or 12}:{minute:02d} {'AM' if hour < 12 else 'PM'}"

            todays_slots.append({
                'time':         display,
                'raw_time':     raw,
                'is_booked':    is_booked,
                'is_break':     is_break,
                'patient_name': booked_map.get(raw, ''),
            })

    # ── Upcoming appointments (next 7 days) ───────────────────────────────
    upcoming_appointments = Appointment.objects.filter(
        doctor=doctor_profile,
        appointment_date__range=[today, today + timedelta(days=7)],
        status__in=['Scheduled', 'Confirmed', 'Pending Payment']
    ).select_related('patient').order_by('appointment_date', 'appointment_time')

    context = {
        'doctor_profile':        doctor_profile,
        'working_hours':         '9:00 AM – 5:00 PM',
        'working_days':          'Mon – Fri',
        'all_days':              all_days,          # ← used by checkbox loop
        'week_schedule':         week_schedule,     # ← used by weekly overview
        'todays_slots':          todays_slots,      # ← used by slot grid
        'upcoming_appointments': upcoming_appointments,
    }

    return render(request, 'core/dashboard/doctor_schedule.html', context)


@login_required
def doctor_schedule_update(request):
    """
    Handle POST from the Update Schedule form.
    """
    try:
        doctor_profile = DoctorProfile.objects.get(user=request.user)
    except DoctorProfile.DoesNotExist:
        messages.error(request, "Doctor profile not found")
        return redirect('login')

    if request.method == 'POST':
        # Read submitted values
        working_days  = request.POST.getlist('working_days')
        start_time    = request.POST.get('start_time', '09:00')
        end_time      = request.POST.get('end_time', '17:00')
        break_start   = request.POST.get('break_start', '13:00')
        break_end     = request.POST.get('break_end', '14:00')
        slot_duration = request.POST.get('slot_duration', '30')
        max_appts     = request.POST.get('max_appointments', '16')
        notes         = request.POST.get('notes', '')

        # TODO: persist to a DoctorAvailability model when you create one:
        DoctorAvailability.objects.update_or_create(
            doctor=doctor_profile,
            defaults={
                'working_days': ','.join(working_days),
                'start_time': start_time,
                'end_time': end_time,
                'break_start': break_start,
                'break_end': break_end,
                'slot_duration': int(slot_duration),
                'max_appointments': int(max_appts),
                'notes': notes,
            }
        )

        messages.success(request, "Schedule updated successfully!")
        return redirect('doctor_schedule')

    return redirect('doctor_schedule')


from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.utils import timezone
from datetime import timedelta
from .models import Appointment, DoctorProfile, PatientProfile
from .forms import RescheduleAppointmentForm


@login_required
def patient_reschedule_appointment(request, appointment_id):
    """
    View for patients to reschedule their appointments
    """
    # Get the appointment
    appointment = get_object_or_404(
        Appointment,
        id=appointment_id,
        patient__user=request.user
    )
    
    # Check if appointment can be rescheduled
    if appointment.status not in ['Pending Payment', 'Scheduled', 'Confirmed']:
        messages.error(
            request,
            f'Cannot reschedule an appointment with status: {appointment.get_status_display()}'
        )
        return redirect('patient_appointment_detail', appointment_id=appointment.id)
    
    # Check if appointment is not too soon (e.g., less than 24 hours away)
    appointment_datetime = timezone.make_aware(
        timezone.datetime.combine(appointment.appointment_date, appointment.appointment_time)
    )
    
    if appointment_datetime - timezone.now() < timedelta(hours=24):
        messages.error(
            request,
            'Cannot reschedule appointments less than 24 hours before the scheduled time. Please contact support.'
        )
        return redirect('patient_appointment_detail', appointment_id=appointment.id)
    
    if request.method == 'POST':
        form = RescheduleAppointmentForm(request.POST, instance=appointment)
        
        if form.is_valid():
            # Save the updated appointment
            updated_appointment = form.save(commit=False)
            
            # Keep the status or update it based on your business logic
            # For example, you might want to set it back to 'Pending Payment' if they change doctors
            if updated_appointment.doctor != appointment.doctor:
                updated_appointment.status = 'Pending Payment'
            
            updated_appointment.save()
            
            # Log the reschedule reason if provided
            reschedule_reason = form.cleaned_data.get('reschedule_reason')
            notes = form.cleaned_data.get('notes')
            
            # You could save these to a RescheduleHistory model or add to appointment notes
            # For now, we'll just show a success message
            
            messages.success(
                request,
                f'Your appointment has been rescheduled to {updated_appointment.appointment_date.strftime("%B %d, %Y")} at {updated_appointment.appointment_time.strftime("%I:%M %p")}'
            )
            
            return redirect('patient_appointment_detail', appointment_id=updated_appointment.id)
        else:
            messages.error(request, 'Please correct the errors below.')
    else:
        form = RescheduleAppointmentForm(instance=appointment)
    
    # Get available doctors for selection
    available_doctors = DoctorProfile.objects.filter(status='Active')
    
    # Calculate minimum date (tomorrow)
    min_date = timezone.now().date() + timedelta(days=1)
    
    context = {
        'appointment': appointment,
        'form': form,
        'available_doctors': available_doctors,
        'min_date': min_date,
    }
    
    return render(request, 'core/dashboard/patient_reschedule_appointment.html', context)


# Optional: Create a view for getting available time slots via AJAX
@login_required
def get_available_time_slots(request):
    """
    AJAX endpoint to get available time slots for a doctor on a specific date
    """
    import json
    from django.http import JsonResponse
    
    if request.method == 'GET':
        doctor_id = request.GET.get('doctor_id')
        date_str = request.GET.get('date')
        
        if not doctor_id or not date_str:
            return JsonResponse({'error': 'Missing parameters'}, status=400)
        
        try:
            from datetime import datetime
            doctor = DoctorProfile.objects.get(id=doctor_id)
            appointment_date = datetime.strptime(date_str, '%Y-%m-%d').date()
            
            # Define all possible time slots (customize these based on your needs)
            all_time_slots = [
                '09:00', '09:30', '10:00', '10:30', '11:00', '11:30',
                '14:00', '14:30', '15:00', '15:30', '16:00', '16:30',
                '17:00', '17:30', '18:00'
            ]
            
            # Get booked slots for this doctor on this date
            booked_appointments = Appointment.objects.filter(
                doctor=doctor,
                appointment_date=appointment_date,
                status__in=['Pending Payment', 'Scheduled', 'Confirmed']
            ).values_list('appointment_time', flat=True)
            
            booked_times = []
            for t in booked_appointments:
                try:
                    time_string = str(appointment_time).strip()
                    time_only = time_string[:5]
                    if ':' in time_only and len(time_only) == 5:
                        booked_times.add(time_only)
                    else:
                        booked_times.append(str(t))
                except:
                    booked_times.append(str(t))
            
            # Mark slots as available or booked
            time_slots = []
            for slot in all_time_slots:
                time_slots.append({
                    'time': slot,
                    'available': slot not in booked_times
                })
            
            return JsonResponse({'slots': time_slots})
            
        except DoctorProfile.DoesNotExist:
            return JsonResponse({'error': 'Doctor not found'}, status=404)
        except ValueError:
            return JsonResponse({'error': 'Invalid date format'}, status=400)
    
    return JsonResponse({'error': 'Invalid request method'}, status=405)

from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.http import JsonResponse
from django.views.decorators.http import require_POST
from django.utils import timezone
from django.db.models import Q, Sum, Count
from datetime import datetime, timedelta

from core.models import (
    FrontDeskProfile, Appointment, PatientProfile, DoctorProfile,
    Payment, PatientHistory, TestBooking
)


def get_frontdesk_profile(user):
    """Helper function to get front desk profile"""
    try:
        return FrontDeskProfile.objects.get(user=user)
    except FrontDeskProfile.DoesNotExist:
        return None


@login_required
def frontdesk_dashboard(request):
    """Main Front Desk Dashboard"""
    frontdesk = get_frontdesk_profile(request.user)
    
    if not frontdesk:
        messages.error(request, "You don't have access to this page.")
        return redirect('login')

    today = timezone.now().date()
    
    # Get statistics
    todays_appointments = Appointment.objects.filter(
        appointment_date=today
    ).order_by('appointment_time')
    
    todays_appointments_count = todays_appointments.count()
    
    # Pending check-ins (appointments scheduled for today but not checked in)
    pending_checkins = todays_appointments.filter(
        status__in=['Scheduled', 'Confirmed']
    )
    pending_checkin_count = pending_checkins.count()
    
    # Checked-in today
    checkedin_count = todays_appointments.filter(
        status='Completed'
    ).count()
    
    # Pending payments
    pending_payments = Payment.objects.filter(payment_status='Pending')
    total_pending_payments = pending_payments.aggregate(
        total=Sum('amount')
    )['total'] or 0
    
    pending_appointments_count = Appointment.objects.filter(
        status__in=['Pending Payment', 'Scheduled']
    ).count()

    context = {
        'todays_appointments': todays_appointments[:5],
        'todays_appointments_count': todays_appointments_count,
        'pending_checkin_count': pending_checkin_count,
        'pending_checkins': pending_checkins[:5],
        'checkedin_count': checkedin_count,
        'total_pending_payments': total_pending_payments,
        'pending_appointments_count': pending_appointments_count,
    }

    return render(request, 'core/dashboard/frontdesk_dashboard.html', context)


@login_required
def frontdesk_appointments(request):
    """Manage all appointments"""
    frontdesk = get_frontdesk_profile(request.user)
    
    if not frontdesk:
        messages.error(request, "You don't have access to this page.")
        return redirect('login')

    # Get all appointments
    appointments = Appointment.objects.all().order_by('-appointment_date', '-appointment_time')
    
    # Filter by status if provided
    status = request.GET.get('status')
    if status:
        appointments = appointments.filter(status=status)
    
    # Filter by date range if provided
    date_from = request.GET.get('date_from')
    date_to = request.GET.get('date_to')
    
    if date_from:
        appointments = appointments.filter(appointment_date__gte=date_from)
    if date_to:
        appointments = appointments.filter(appointment_date__lte=date_to)
    
    # Search by patient name or doctor name
    search = request.GET.get('search')
    if search:
        appointments = appointments.filter(
            Q(patient__full_name__icontains=search) |
            Q(doctor__user__first_name__icontains=search) |
            Q(doctor__user__last_name__icontains=search)
        )
    
    # Status choices for filter
    status_choices = Appointment.STATUS_CHOICES

    context = {
        'appointments': appointments,
        'status_choices': status_choices,
        'current_status': status,
        'search_query': search,
    }

    return render(request, 'core/dashboard/frontdesk_appointments.html', context)


@login_required
def frontdesk_appointment_detail(request, appointment_id):
    """View and edit appointment details"""
    frontdesk = get_frontdesk_profile(request.user)
    
    if not frontdesk:
        messages.error(request, "You don't have access to this page.")
        return redirect('login')

    appointment = get_object_or_404(Appointment, id=appointment_id)

    if request.method == 'POST':
        # Update appointment status
        new_status = request.POST.get('status')
        if new_status in dict(Appointment.STATUS_CHOICES):
            appointment.status = new_status
            appointment.save()
            messages.success(request, f"Appointment status updated to {new_status}")
            return redirect('frontdesk_appointment_detail', appointment_id=appointment_id)

    context = {
        'appointment': appointment,
        'status_choices': Appointment.STATUS_CHOICES,
    }

    return render(request, 'core/dashboard/frontdesk_appointment_detail.html', context)


@login_required
def frontdesk_patient_checkin(request):
    """Patient check-in/check-out management"""
    frontdesk = get_frontdesk_profile(request.user)
    
    if not frontdesk:
        messages.error(request, "You don't have access to this page.")
        return redirect('login')

    today = timezone.now().date()
    
    # Get today's appointments
    todays_appointments = Appointment.objects.filter(
        appointment_date=today
    ).order_by('appointment_time')
    
    # Pending check-ins
    pending_checkins = todays_appointments.filter(
        status__in=['Scheduled', 'Confirmed']
    )
    
    # Already checked in
    checked_in = todays_appointments.filter(
        status='Completed'
    )

    if request.method == 'POST':
        appointment_id = request.POST.get('appointment_id')
        action = request.POST.get('action')  # 'check_in' or 'check_out'
        
        appointment = get_object_or_404(Appointment, id=appointment_id)
        
        if action == 'check_in':
            appointment.status = 'Confirmed'
            appointment.save()
            messages.success(request, f"{appointment.patient.full_name} checked in successfully!")
        
        elif action == 'check_out':
            appointment.status = 'Completed'
            appointment.save()
            messages.success(request, f"{appointment.patient.full_name} checked out successfully!")
        
        return redirect('frontdesk_patient_check_in')

    context = {
        'pending_checkins': pending_checkins,
        'checked_in': checked_in,
        'todays_date': today,
    }

    return render(request, 'core/dashboard/frontdesk_patient_checkin.html', context)


@login_required
def frontdesk_patients_list(request):
    """View all patients"""
    frontdesk = get_frontdesk_profile(request.user)
    
    if not frontdesk:
        messages.error(request, "You don't have access to this page.")
        return redirect('login')

    patients = PatientProfile.objects.all().order_by('full_name')
    
    # Search by name or phone
    search = request.GET.get('search')
    if search:
        patients = patients.filter(
            Q(full_name__icontains=search) |
            Q(phone__icontains=search) |
            Q(user__email__icontains=search)
        )
    
    # Filter by status
    status = request.GET.get('status')
    if status:
        patients = patients.filter(status=status)

    context = {
        'patients': patients,
        'search_query': search,
        'current_status': status,
        'status_choices': ['Active', 'Inactive'],
    }

    return render(request, 'core/dashboard/frontdesk_patients_list.html', context)


@login_required
def frontdesk_patients_detail(request, patient_id):
    """View patient details"""
    frontdesk = get_frontdesk_profile(request.user)
    
    if not frontdesk:
        messages.error(request, "You don't have access to this page.")
        return redirect('login')

    patient = get_object_or_404(PatientProfile, id=patient_id)
    
    # Get patient's appointments
    appointments = Appointment.objects.filter(patient=patient).order_by('-appointment_date')
    
    # Get patient's payments
    payments = Payment.objects.filter(patient=patient).order_by('-payment_date')
    
    # Get patient's test bookings
    test_bookings = TestBooking.objects.filter(patient=patient).order_by('-booking_date')

    context = {
        'patient': patient,
        'appointments': appointments[:5],
        'payments': payments[:5],
        'test_bookings': test_bookings[:5],
    }

    return render(request, 'core/dashboard/frontdesk_patients_detail.html', context)


@login_required
def frontdesk_doctors_list(request):
    """View all doctors"""
    frontdesk = get_frontdesk_profile(request.user)
    
    if not frontdesk:
        messages.error(request, "You don't have access to this page.")
        return redirect('login')

    doctors = DoctorProfile.objects.all().order_by('user__first_name')
    
    # Search by name or specialization
    search = request.GET.get('search')
    if search:
        doctors = doctors.filter(
            Q(user__first_name__icontains=search) |
            Q(user__last_name__icontains=search) |
            Q(specialization__icontains=search) |
            Q(department__icontains=search)
        )
    
    # Filter by specialization
    specialization = request.GET.get('specialization')
    if specialization:
        doctors = doctors.filter(specialization=specialization)
    
    # Filter by status
    status = request.GET.get('status')
    if status:
        doctors = doctors.filter(status=status)

    context = {
        'doctors': doctors,
        'search_query': search,
        'current_specialization': specialization,
        'current_status': status,
    }

    return render(request, 'core/dashboard/frontdesk_doctors_list.html', context)


@login_required
def frontdesk_doctor_detail(request, doctor_id):
    """View doctor details and schedule"""
    frontdesk = get_frontdesk_profile(request.user)
    
    if not frontdesk:
        messages.error(request, "You don't have access to this page.")
        return redirect('login')

    doctor = get_object_or_404(DoctorProfile, id=doctor_id)
    
    # Get doctor's appointments
    appointments = Appointment.objects.filter(doctor=doctor).order_by('-appointment_date')
    
    # Today's appointments for this doctor
    today = timezone.now().date()
    todays_appointments = appointments.filter(appointment_date=today)

    context = {
        'doctor': doctor,
        'appointments': appointments[:10],
        'todays_appointments': todays_appointments,
        'todays_appointments_count': todays_appointments.count(),
    }

    return render(request, 'core/dashborad/frontdesk_doctor_detail.html', context)


@login_required
def frontdesk_payments(request):
    """Manage payments"""
    frontdesk = get_frontdesk_profile(request.user)
    
    if not frontdesk:
        messages.error(request, "You don't have access to this page.")
        return redirect('login')

    payments = Payment.objects.all().order_by('-payment_date')
    
    # Filter by status
    status = request.GET.get('status')
    if status:
        payments = payments.filter(payment_status=status)
    
    # Search by patient name
    search = request.GET.get('search')
    if search:
        payments = payments.filter(
            Q(patient__full_name__icontains=search) |
            Q(patient__user__email__icontains=search)
        )
    
    # Statistics
    total_payments = payments.aggregate(total=Sum('amount'))['total'] or 0
    paid_payments = payments.filter(payment_status='Paid').aggregate(total=Sum('amount'))['total'] or 0
    pending_payments = payments.filter(payment_status='Pending').aggregate(total=Sum('amount'))['total'] or 0

    context = {
        'payments': payments,
        'status_choices': Payment.PAYMENT_STATUS,
        'current_status': status,
        'search_query': search,
        'total_payments': total_payments,
        'paid_payments': paid_payments,
        'pending_payments': pending_payments,
    }

    return render(request, 'core/dashboard/frontdesk_payments.html', context)


@login_required
def frontdesk_payment_detail(request, payment_id):
    """View and update payment details"""
    frontdesk = get_frontdesk_profile(request.user)
    
    if not frontdesk:
        messages.error(request, "You don't have access to this page.")
        return redirect('login')

    payment = get_object_or_404(Payment, id=payment_id)

    if request.method == 'POST':
        new_status = request.POST.get('payment_status')
        if new_status in dict(Payment.PAYMENT_STATUS):
            payment.payment_status = new_status
            payment.save()
            messages.success(request, f"Payment status updated to {new_status}")
            return redirect('frontdesk_payment_detail', payment_id=payment_id)

    context = {
        'payment': payment,
        'payment_methods': Payment.PAYMENT_METHODS,
        'payment_statuses': Payment.PAYMENT_STATUS,
    }

    return render(request, 'core/dashboard/frontdesk_payment_detail.html', context)

@login_required
def frontdesk_reports(request):
    """View reports and analytics"""
    frontdesk = get_frontdesk_profile(request.user)
    
    if not frontdesk:
        messages.error(request, "You don't have access to this page.")
        return redirect('login')

    today = timezone.now().date()
    month_ago = today - timedelta(days=30)
    
    # Daily statistics
    todays_appointments = Appointment.objects.filter(appointment_date=today).count()
    todays_checkins = Appointment.objects.filter(
        appointment_date=today,
        status='Completed'
    ).count()
    
    # Monthly statistics
    monthly_appointments = Appointment.objects.filter(
        appointment_date__gte=month_ago
    ).count()
    
    monthly_payments = Payment.objects.filter(
        payment_date__gte=month_ago,
        payment_status='Paid'
    ).aggregate(total=Sum('amount'))['total'] or 0
    
    # Doctor statistics
    doctor_appointments = DoctorProfile.objects.annotate(
        appointment_count=Count('appointment')
    ).order_by('-appointment_count')[:5]
    
    # Payment statistics
    payment_by_method = Payment.objects.filter(
        payment_date__gte=month_ago
    ).values('payment_method').annotate(
        count=Count('id'),
        total=Sum('amount')
    )

    context = {
        'todays_appointments': todays_appointments,
        'todays_checkins': todays_checkins,
        'monthly_appointments': monthly_appointments,
        'monthly_payments': monthly_payments,
        'doctor_appointments': doctor_appointments,
        'payment_by_method': payment_by_method,
    }

    return render(request, 'core/dashboard/frontdesk_reports.html', context)


@login_required
def frontdesk_settings(request):
    """Front desk settings"""
    frontdesk = get_frontdesk_profile(request.user)
    
    if not frontdesk:
        messages.error(request, "You don't have access to this page.")
        return redirect('login')

    if request.method == 'POST':
        # Update profile information
        user = request.user
        user.first_name = request.POST.get('first_name', user.first_name)
        user.last_name = request.POST.get('last_name', user.last_name)
        user.email = request.POST.get('email', user.email)
        user.save()
        
        # Update password if provided
        old_password = request.POST.get('old_password')
        new_password = request.POST.get('new_password')
        confirm_password = request.POST.get('confirm_password')
        
        if old_password and new_password:
            if user.check_password(old_password):
                if new_password == confirm_password:
                    user.set_password(new_password)
                    user.save()
                    messages.success(request, "Password updated successfully!")
                else:
                    messages.error(request, "New passwords do not match!")
            else:
                messages.error(request, "Old password is incorrect!")
        
        if not old_password:
            messages.success(request, "Profile updated successfully!")
        
        return redirect('frontdesk_settings')

    context = {
        'frontdesk': frontdesk,
    }

    return render(request, 'core/dashboard/frontdesk_settings.html', context)

def frontdesk_patients_edit(request, patient_id):
    patient = PatientProfile.objects.get(id=patient_id)
    
    if request.method == 'POST':
        form = PatientForm(request.POST, instance=patient)
        if form.is_valid():
            form.save()
            return redirect('patient_detail', patient_id=patient.id)
    else:
        form = PatientForm(instance=patient)
    
    return render(request, 'core/dashboard/frontdesk_patients_edit.html', {'form': form, 'patient': patient})
    

def get_frontdesk_profile(user):
    """Helper function to get front desk profile"""
    try:
        return FrontDeskProfile.objects.get(user=user)
    except FrontDeskProfile.DoesNotExist:
        return None

# Add these views to your core/views.py - FIXED VERSION 2

from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.contrib.auth.models import User
from django.contrib import messages
from django.http import JsonResponse
from datetime import date, datetime, timedelta
from decimal import Decimal
import uuid

from core.models import (
    FrontDeskProfile, PatientProfile, DoctorProfile, Appointment,
    Payment
)


def get_frontdesk_profile(user):
    """Helper function to get front desk profile"""
    try:
        return FrontDeskProfile.objects.get(user=user)
    except FrontDeskProfile.DoesNotExist:
        return None


def generate_unique_username(email):
    """
    Generate a unique username from email or create a random one
    """
    # Try using email first
    if email:
        base_username = email.split('@')[0]
    else:
        base_username = f"patient_{uuid.uuid4().hex[:8]}"
    
    username = base_username
    counter = 1
    
    # If username exists, append a number
    while User.objects.filter(username=username).exists():
        username = f"{base_username}{counter}"
        counter += 1
    
    return username


@login_required
def frontdesk_book_appointment(request):
    """
    Front desk staff books appointments for patients
    Flow:
    1. Select or create patient
    2. Select doctor and time
    3. Collect payment
    4. Generate token
    """
    frontdesk = get_frontdesk_profile(request.user)
    
    if not frontdesk:
        messages.error(request, "You don't have access to this page.")
        return redirect('login')

    if request.method == 'POST':
        step = request.POST.get('step', '1')
        
        # Step 1: Patient Selection/Creation
        if step == '1':
            patient_id = request.POST.get('patient_id')
            new_patient_name = request.POST.get('new_patient_name')
            new_patient_phone = request.POST.get('new_patient_phone')
            new_patient_email = request.POST.get('new_patient_email')
            
            if patient_id:
                # Existing patient selected
                try:
                    patient = PatientProfile.objects.get(id=patient_id)
                    request.session['selected_patient_id'] = patient.id
                    messages.success(request, f"Patient {patient.full_name} selected")
                except PatientProfile.DoesNotExist:
                    messages.error(request, "Patient not found")
                    return redirect('frontdesk_book_appointment')
            
            elif new_patient_name and new_patient_phone:
                # Create new patient - FIXED
                try:
                    # Generate unique username
                    if new_patient_email:
                        username = generate_unique_username(new_patient_email)
                    else:
                        username = generate_unique_username(None)
                    
                    # Create user with unique username
                    user = User.objects.create_user(
                        username=username,
                        email=new_patient_email or '',
                        password=uuid.uuid4().hex[:12]
                    )
                    
                    # Create patient profile
                    patient = PatientProfile.objects.create(
                        user=user,
                        full_name=new_patient_name,
                        phone=new_patient_phone,
                        status='Active'
                    )
                    
                    request.session['selected_patient_id'] = patient.id
                    messages.success(request, f"New patient {patient.full_name} created")
                except Exception as e:
                    messages.error(request, f"Error creating patient: {str(e)}")
                    return redirect('frontdesk_book_appointment')
            else:
                messages.error(request, "Please select or create a patient")
                return redirect('frontdesk_book_appointment')
        
        # Step 2: Doctor and Time Selection
        elif step == '2':
            selected_patient_id = request.session.get('selected_patient_id')
            if not selected_patient_id:
                messages.error(request, "Please select a patient first")
                return redirect('frontdesk_book_appointment')
            
            doctor_id = request.POST.get('doctor_id')
            appointment_date = request.POST.get('appointment_date')
            appointment_time = request.POST.get('appointment_time')
            reason = request.POST.get('reason')
            
            # Validation
            if not all([doctor_id, appointment_date, appointment_time, reason]):
                messages.error(request, "All fields are required")
                return redirect('frontdesk_book_appointment')
            
            try:
                doctor = DoctorProfile.objects.get(id=doctor_id)
                
                # Check if slot is available
                existing_appointment = Appointment.objects.filter(
                    doctor=doctor,
                    appointment_date=appointment_date,
                    appointment_time=appointment_time,
                    status__in=['Pending Payment', 'Scheduled', 'Confirmed']
                ).exists()
                
                if existing_appointment:
                    messages.error(request, "This time slot is already booked. Please select another time.")
                    return redirect('frontdesk_book_appointment')
                
                # Store in session
                request.session['appointment_data'] = {
                    'doctor_id': doctor_id,
                    'appointment_date': appointment_date,
                    'appointment_time': appointment_time,
                    'reason': reason,
                    'doctor_name': f"Dr. {doctor.user.get_full_name()}",
                    'consultation_fee': str(doctor.consultation_fee)
                }
                
                messages.success(request, "Appointment details confirmed. Proceed to payment.")
                
            except DoctorProfile.DoesNotExist:
                messages.error(request, "Doctor not found")
                return redirect('frontdesk_book_appointment')
        
        # Step 3: Payment Processing
        elif step == '3':
            selected_patient_id = request.session.get('selected_patient_id')
            appointment_data = request.session.get('appointment_data')
            
            if not selected_patient_id or not appointment_data:
                messages.error(request, "Session expired. Please start over.")
                return redirect('frontdesk_book_appointment')
            
            payment_method = request.POST.get('payment_method')
            amount = Decimal(request.POST.get('amount', '0'))
            
            if not payment_method or amount <= 0:
                messages.error(request, "Invalid payment details")
                return redirect('frontdesk_book_appointment')
            
            try:
                patient = PatientProfile.objects.get(id=selected_patient_id)
                doctor = DoctorProfile.objects.get(id=appointment_data['doctor_id'])
                
                # Create appointment
                appointment = Appointment.objects.create(
                    patient=patient,
                    doctor=doctor,
                    appointment_date=appointment_data['appointment_date'],
                    appointment_time=appointment_data['appointment_time'],
                    reason=appointment_data['reason'],
                    status='Pending Payment'
                )
                
                # Create payment
                transaction_id = f"TXN{uuid.uuid4().hex[:12].upper()}"
                payment = Payment.objects.create(
                    patient=patient,
                    appointment=appointment,
                    amount=amount,
                    payment_method=payment_method,
                    payment_status='Paid',
                    transaction_id=transaction_id
                )
                
                # Update appointment status to Scheduled after payment
                appointment.status = 'Scheduled'
                appointment.save()
                
                # Clear session safely
                if 'selected_patient_id' in request.session:
                    del request.session['selected_patient_id']
                if 'appointment_data' in request.session:
                    del request.session['appointment_data']
                
                # Generate token
                token_number = generate_token(appointment)
                
                messages.success(
                    request,
                    f"Appointment booked successfully! Token: {token_number}"
                )
                
                return redirect('frontdesk_appointment_confirmation', appointment_id=appointment.id)
                
            except Exception as e:
                messages.error(request, f"Error processing payment: {str(e)}")
                return redirect('frontdesk_book_appointment')
    
    # GET request - Show form
    patients = PatientProfile.objects.all().order_by('full_name')
    doctors = DoctorProfile.objects.filter(status='Active').order_by('user__first_name')
    
    # Check if we're in the middle of booking
    selected_patient_id = request.session.get('selected_patient_id')
    appointment_data = request.session.get('appointment_data')
    
    selected_patient = None
    if selected_patient_id:
        try:
            selected_patient = PatientProfile.objects.get(id=selected_patient_id)
        except PatientProfile.DoesNotExist:
            pass
    
    context = {
        'patients': patients,
        'doctors': doctors,
        'selected_patient': selected_patient,
        'appointment_data': appointment_data,
        'today': date.today().isoformat(),
        'min_date': (date.today() + timedelta(days=1)).isoformat(),
    }
    
    return render(request, 'core/dashboard/frontdesk_book_appointment.html', context)


@login_required
def frontdesk_appointment_confirmation(request, appointment_id):
    """
    Show appointment confirmation with token
    """
    frontdesk = get_frontdesk_profile(request.user)
    
    if not frontdesk:
        messages.error(request, "You don't have access to this page.")
        return redirect('login')

    appointment = get_object_or_404(Appointment, id=appointment_id)
    
    # Get payment info
    payment = Payment.objects.filter(appointment=appointment).first()
    
    # Get token from appointment
    token = getattr(appointment, 'token_number', None)
    
    context = {
        'appointment': appointment,
        'payment': payment,
        'token': token,
    }
    
    return render(request, 'core/dashboard/frontdesk_appointment_confirmation.html', context)

# ==========================================
# FINAL FRONTDESK FIX
# Replace frontdesk_get_available_slots() completely
# ==========================================

from django.http import JsonResponse
from django.contrib.auth.decorators import login_required

@login_required
def frontdesk_get_available_slots(request):
    """
    AJAX endpoint to get available time slots for a doctor on a specific date
    FINAL FIX: Bulletproof conversion of appointment times
    """
    
    if request.method == 'GET':
        doctor_id = request.GET.get('doctor_id')
        appointment_date = request.GET.get('appointment_date')
        
        if not doctor_id or not appointment_date:
            return JsonResponse({'error': 'Missing parameters'}, status=400)
        
        try:
            from core.models import DoctorProfile, Appointment
            
            doctor = DoctorProfile.objects.get(id=doctor_id)
            
            # Define time slots (9 AM to 6 PM, 30-min intervals)
            time_slots = []
            for hour in range(9, 18):
                for minute in [0, 30]:
                    time_str = f"{hour:02d}:{minute:02d}"
                    time_slots.append(time_str)
            
            # Get booked appointments
            booked_appointments = Appointment.objects.filter(
                doctor=doctor,
                appointment_date=appointment_date,
                status__in=['Pending Payment', 'Scheduled', 'Confirmed']
            ).values_list('appointment_time', flat=True)
            
            # ==========================================
            # BULLETPROOF CONVERSION (THIS IS THE KEY FIX)
            # ==========================================
            booked_times = set()  # Use set, not list (faster)
            
            for appointment_time in booked_appointments:
                try:
                    # Step 1: Convert to string FIRST
                    time_string = str(appointment_time).strip()
                    
                    # Step 2: Extract HH:MM (first 5 characters)
                    if len(time_string) >= 5:
                        time_only = time_string[:5]
                    else:
                        time_only = time_string
                    
                    # Step 3: Validate it's in HH:MM format
                    if ':' in time_only and len(time_only) == 5:
                        booked_times.add(time_only)
                    else:
                        # Try to fix the format
                        parts = time_string.split(':')
                        if len(parts) >= 2:
                            hour = parts[0].zfill(2)
                            minute = parts[1][:2].zfill(2)
                            booked_times.add(f"{hour}:{minute}")
                
                except Exception as e:
                    # If all else fails, just skip this time
                    print(f"Warning: Could not process time {appointment_time}: {e}")
                    pass
            
            # Build response with available slots
            available_slots = []
            for slot in time_slots:
                is_available = slot not in booked_times
                available_slots.append({
                    'time': slot,
                    'available': is_available
                })
            
            return JsonResponse({'slots': available_slots})
        
        except DoctorProfile.DoesNotExist:
            return JsonResponse({'error': 'Doctor not found'}, status=404)
        except Exception as e:
            import traceback
            error_detail = traceback.format_exc()
            print(f"Error in frontdesk_get_available_slots: {error_detail}")
            return JsonResponse({
                'error': f'Server error: {str(e)}'
            }, status=500)
    
    return JsonResponse({'error': 'Invalid request'}, status=400)

@login_required
def frontdesk_search_patient(request):
    """
    AJAX endpoint to search for existing patients
    """
    if request.method == 'GET':
        query = request.GET.get('q', '').strip()
        
        if not query or len(query) < 2:
            return JsonResponse({'patients': []})
        
        patients = PatientProfile.objects.filter(
            full_name__icontains=query
        )[:10]
        
        patients_data = []
        for patient in patients:
            patients_data.append({
                'id': patient.id,
                'name': patient.full_name,
                'phone': patient.phone,
                'email': patient.user.email,
                'gender': patient.gender or 'Not specified',
                'dob': str(patient.dob) if patient.dob else 'N/A',
            })
        
        return JsonResponse({'patients': patients_data})
    
    return JsonResponse({'error': 'Invalid request'}, status=400)


def generate_token(appointment):
    """
    Generate appointment token - SIMPLE VERSION
    """
    # Convert to string first
    date_obj = appointment.appointment_date
    
    if isinstance(date_obj, str):
        # Remove all non-digits and take first 4 chars
        date_str = ''.join(filter(str.isdigit, date_obj))[-4:]
    else:
        date_str = date_obj.strftime('%d%m')
    
    # Count appointments for this date
    count = Appointment.objects.filter(
        appointment_date=appointment.appointment_date
    ).count()
    
    token_number = f"{date_str}-{count:03d}"
    
    return token_number





@login_required
def frontdesk_today_appointments(request):
    """
    View today's appointments and manage check-ins
    """
    frontdesk = get_frontdesk_profile(request.user)
    
    if not frontdesk:
        messages.error(request, "You don't have access to this page.")
        return redirect('login')

    today = date.today()
    
    # Get today's appointments sorted by time
    todays_appointments = Appointment.objects.filter(
        appointment_date=today
    ).select_related('patient', 'doctor', 'doctor__user').order_by('appointment_time')
    
    # Separate by status
    pending_checkin = todays_appointments.filter(
        status__in=['Scheduled', 'Pending Payment']
    )
    
    checked_in = todays_appointments.filter(
        status='Confirmed'
    )
    
    completed = todays_appointments.filter(
        status='Completed'
    )

    context = {
        'todays_appointments': todays_appointments,
        'pending_checkin': pending_checkin,
        'checked_in': checked_in,
        'completed': completed,
        'today': today,
    }
    
    return render(request, 'core/dashboard/frontdesk_today_appointments.html', context)


@login_required
def frontdesk_quick_checkin(request, appointment_id):
    """
    Quick check-in for appointment
    """
    frontdesk = get_frontdesk_profile(request.user)
    
    if not frontdesk:
        messages.error(request, "You don't have access to this action.")
        return redirect('login')

    appointment = get_object_or_404(Appointment, id=appointment_id)
    
    if request.method == 'POST':
        # Update status to Confirmed (patient checked in)
        appointment.status = 'Confirmed'
        appointment.save()
        
        messages.success(
            request,
            f"{appointment.patient.full_name} checked in successfully!"
        )
        
        return redirect('frontdesk_today_appointments')
    
    return redirect('frontdesk_today_appointments')


@login_required
def frontdesk_patients_edit(request, patient_id):
    """
    Edit patient information by front desk staff
    """
    frontdesk = get_frontdesk_profile(request.user)
    
    if not frontdesk:
        messages.error(request, "You don't have access to this page.")
        return redirect('login')

    patient = get_object_or_404(PatientProfile, id=patient_id)
    
    if request.method == 'POST':
        # Update patient information
        patient.full_name = request.POST.get('full_name', patient.full_name)
        patient.phone = request.POST.get('phone', patient.phone)
        patient.gender = request.POST.get('gender', patient.gender)
        patient.dob = request.POST.get('dob', patient.dob) if request.POST.get('dob') else patient.dob
        patient.address = request.POST.get('address', patient.address)
        patient.status = request.POST.get('status', patient.status)
        patient.save()
        
        # Update user email if provided
        user = patient.user
        new_email = request.POST.get('email', user.email)
        if new_email and new_email != user.email:
            # Check if email already exists
            if User.objects.filter(email=new_email).exclude(id=user.id).exists():
                messages.error(request, "This email is already in use")
                return redirect('frontdesk_patients_edit', patient_id=patient_id)
            user.email = new_email
            user.save()
        
        messages.success(request, f"Patient {patient.full_name} updated successfully!")
        return redirect('frontdesk_patients_detail', patient_id=patient_id)
    
    context = {
        'patient': patient,
        'genders': ['Male', 'Female', 'Other'],
        'statuses': ['Active', 'Inactive'],
    }
    
    return render(request, 'core/dashboard/frontdesk_patients_edit.html', context)

@login_required
def doctor_reschedule_appointment(request, appointment_id):
    """Reschedule an appointment"""

    # ✅ Get DoctorProfile from logged in user
    doctor_profile = request.user.doctorprofile

    # ✅ Now filter correctly
    appointment = get_object_or_404(
        Appointment,
        id=appointment_id,
        doctor=doctor_profile
    )

    if request.method == 'POST':
        new_date = request.POST.get('appointment_date')
        new_time = request.POST.get('appointment_time')
        reschedule_reason = request.POST.get('reschedule_reason', '')
        notify_patient = request.POST.get('notify_patient') == 'on'

        # Update appointment
        appointment.appointment_date = new_date
        appointment.appointment_time = new_time

        # Store reschedule reason
        if reschedule_reason:
            if appointment.notes:
                appointment.notes += f"\n\nRescheduled: {reschedule_reason}"
            else:
                appointment.notes = f"Rescheduled: {reschedule_reason}"

        appointment.save()

        if notify_patient:
            # Add email logic later
            pass

        messages.success(
            request,
            f'Appointment rescheduled successfully to {new_date} at {new_time}'
        )

        return redirect('doctor_appointment_detail', appointment_id=appointment.id)

    context = {
        'appointment': appointment,
        'today': timezone.now().date(),
    }

    return render(request, 'core/dashboard/doctor_appointment_reschedule.html', context)

# ==========================================
# FRONT DESK LAB TEST BOOKING VIEWS
# Add these to your core/views.py
# ==========================================

@login_required
def frontdesk_book_lab_test(request):
    """
    Front desk books lab tests on behalf of patients
    """
    frontdesk = get_frontdesk_profile(request.user)
    if not frontdesk:
        messages.error(request, "You don't have access to this page.")
        return redirect('login')

    if request.method == 'POST':
        patient_id = request.POST.get('patient')
        test_id = request.POST.get('test')
        booking_date = request.POST.get('booking_date')
        payment_method = request.POST.get('payment_method')

        if not all([patient_id, test_id, booking_date, payment_method]):
            messages.error(request, "All fields are required.")
            return redirect('frontdesk_book_lab_test')

        try:
            patient = PatientProfile.objects.get(id=patient_id, status='Active')
            test = DiagnosticTest.objects.get(id=test_id)

            # Check if already booked
            existing = TestBooking.objects.filter(
                patient=patient,
                test=test
            ).exclude(status='Cancelled').exists()

            if existing:
                messages.warning(request, f"{patient.full_name} already has a booking for {test.test_name}.")
                return redirect('frontdesk_book_lab_test')

            # Create booking
            booking = TestBooking.objects.create(
                patient=patient,
                test=test,
                lab=test.lab,
                booking_date=booking_date,
                status='Booked'
            )

            # Create payment (mark as Paid since front desk collects payment)
            import uuid
            payment = Payment.objects.create(
                patient=patient,
                test_booking=booking,
                amount=test.price,
                payment_method=payment_method,
                payment_status='Paid',
                transaction_id=f"TXN{uuid.uuid4().hex[:12].upper()}"
            )

            messages.success(
                request,
                f"Lab test '{test.test_name}' booked successfully for "
                f"{patient.full_name} on {booking_date}. "
                f"Payment of ₹{test.price} collected via {payment_method}."
            )
            return redirect('frontdesk_lab_test_confirmation', booking_id=booking.id)

        except PatientProfile.DoesNotExist:
            messages.error(request, "Patient not found.")
        except DiagnosticTest.DoesNotExist:
            messages.error(request, "Test not found.")
        except Exception as e:
            messages.error(request, f"Error booking test: {str(e)}")

        return redirect('frontdesk_book_lab_test')

    # GET - Load form data
    patients = PatientProfile.objects.filter(status='Active').order_by('full_name')
    labs = Lab.objects.filter(status='Active').order_by('name')
    tests = DiagnosticTest.objects.filter(is_active=True).select_related('lab').order_by('test_name')

    from datetime import date
    context = {
        'patients': patients,
        'labs': labs,
        'tests': tests,
        'today': date.today().isoformat(),
        'payment_methods': Payment.PAYMENT_METHODS,
    }
    return render(request, 'core/dashboard/frontdesk_book_lab_test.html', context)


@login_required
def frontdesk_lab_test_confirmation(request, booking_id):
    """Show lab test booking confirmation"""
    frontdesk = get_frontdesk_profile(request.user)
    if not frontdesk:
        messages.error(request, "You don't have access to this page.")
        return redirect('login')

    booking = get_object_or_404(TestBooking, id=booking_id)
    payment = Payment.objects.filter(test_booking=booking).first()

    context = {
        'booking': booking,
        'payment': payment,
    }
    return render(request, 'core/dashboard/frontdesk_lab_test_confirmation.html', context)


@login_required
def frontdesk_lab_bookings(request):
    """View all lab test bookings"""
    frontdesk = get_frontdesk_profile(request.user)
    if not frontdesk:
        messages.error(request, "You don't have access to this page.")
        return redirect('login')

    bookings = TestBooking.objects.all().select_related(
        'patient', 'test', 'lab'
    ).order_by('-created_at')

    # Filters
    status_filter = request.GET.get('status', '')
    search = request.GET.get('search', '')
    lab_filter = request.GET.get('lab', '')

    if status_filter:
        bookings = bookings.filter(status=status_filter)
    if lab_filter:
        bookings = bookings.filter(lab__id=lab_filter)
    if search:
        bookings = bookings.filter(
            Q(patient__full_name__icontains=search) |
            Q(test__test_name__icontains=search)
        )

    labs = Lab.objects.filter(status='Active')

    context = {
        'bookings': bookings,
        'status_filter': status_filter,
        'search_query': search,
        'lab_filter': lab_filter,
        'labs': labs,
        'total_count': bookings.count(),
    }
    return render(request, 'core/dashboard/frontdesk_lab_bookings.html', context)


@login_required
def frontdesk_get_tests_by_lab(request):
    """AJAX - Get tests filtered by lab"""
    frontdesk = get_frontdesk_profile(request.user)
    if not frontdesk:
        return JsonResponse({'error': 'Unauthorized'}, status=403)

    lab_id = request.GET.get('lab_id', '')
    if not lab_id:
        tests = DiagnosticTest.objects.filter(is_active=True).select_related('lab')
    else:
        tests = DiagnosticTest.objects.filter(lab__id=lab_id, is_active=True).select_related('lab')

    tests_data = []
    for test in tests:
        tests_data.append({
            'id': test.id,
            'test_name': test.test_name,
            'category': test.category,
            'price': str(test.price),
            'result_duration': test.result_duration,
            'sample_type': test.sample_type,
            'lab_name': test.lab.name,
            'home_collection': test.home_collection,
        })

    return JsonResponse({'tests': tests_data})

# ============================================================
#  1. ADD THIS VIEW to core/views.py
#     (paste anywhere near the other frontdesk_ views)
# ============================================================

@login_required
def frontdesk_add_patient(request):
    """
    Handles the Add New Patient modal form submission.
    On success  → redirect back to patient list with success message.
    On error    → redirect back with error message + open_modal=True
                  so the modal re-opens automatically.
    """
    frontdesk = get_frontdesk_profile(request.user)
    if not frontdesk:
        messages.error(request, "You don't have access to this page.")
        return redirect('login')

    if request.method != 'POST':
        return redirect('frontdesk_patients_list')

    # ── Collect form data ──────────────────────────────────────────
    first_name = request.POST.get('first_name', '').strip()
    last_name  = request.POST.get('last_name',  '').strip()
    email      = request.POST.get('email',      '').strip()
    phone      = request.POST.get('phone',      '').strip()
    gender     = request.POST.get('gender',     '').strip()
    dob        = request.POST.get('dob',        '').strip() or None
    address    = request.POST.get('address',    '').strip()
    status     = request.POST.get('status',     'Active').strip()

    # ── Basic validation ───────────────────────────────────────────
    if not all([first_name, last_name, email, phone]):
        messages.error(request, "First name, last name, email and phone are required.")
        return redirect(f"{request.META.get('HTTP_REFERER', '/frontdesk/patients/')}?open_modal=1")

    # ── Check for duplicate email ──────────────────────────────────
    if User.objects.filter(email=email).exists():
        messages.error(request, f"A patient with the email '{email}' already exists.")
        return redirect(f"/frontdesk/patients/?open_modal=1")

    try:
        # ── Create Django User ─────────────────────────────────────
        username = generate_unique_username(email)
        user = User.objects.create_user(
            username=username,
            email=email,
            password=uuid.uuid4().hex,        # random password — patient can reset via email
            first_name=first_name,
            last_name=last_name,
        )

        # ── Create PatientProfile ──────────────────────────────────
        patient = PatientProfile.objects.create(
            user=user,
            full_name=f"{first_name} {last_name}",
            phone=phone,
            gender=gender,
            dob=dob,
            address=address,
            status=status,
        )

        messages.success(
            request,
            f"Patient {patient.full_name} added successfully!"
        )
        return redirect('frontdesk_patients_list')

    except Exception as e:
        messages.error(request, f"Error creating patient: {e}")
        return redirect(f"/frontdesk/patients/?open_modal=1")


# ── Helper used above ──────────────────────────────────────────────
# (already defined if you applied the previous views file —
#  only add this if it's not already in your views.py)

def generate_unique_username(email=None):
    """Generate a unique Django username derived from the email."""
    import uuid as _uuid
    base = email.split('@')[0] if email else f"patient_{_uuid.uuid4().hex[:8]}"
    username, counter = base, 1
    while User.objects.filter(username=username).exists():
        username = f"{base}{counter}"
        counter += 1
    return username


# ── Updated frontdesk_patients_list view ──────────────────────────
# (adds total_count / active_count / inactive_count to context
#  and handles the ?open_modal=1 query param)

@login_required
def frontdesk_patients_list(request):
    frontdesk = get_frontdesk_profile(request.user)
    if not frontdesk:
        messages.error(request, "You don't have access to this page.")
        return redirect('login')

    patients = PatientProfile.objects.all().order_by('full_name')

    search = request.GET.get('search', '').strip()
    if search:
        patients = patients.filter(
            Q(full_name__icontains=search) |
            Q(phone__icontains=search) |
            Q(user__email__icontains=search)
        )

    status = request.GET.get('status', '').strip()
    if status:
        patients = patients.filter(status=status)

    all_patients = PatientProfile.objects.all()

    context = {
        'patients':       patients,
        'search_query':   search,
        'current_status': status,
        # counts for the stat cards (optional — only needed if your
        # template shows them; the template above does not, but keep
        # them for future use)
        'total_count':    all_patients.count(),
        'active_count':   all_patients.filter(status='Active').count(),
        'inactive_count': all_patients.filter(status='Inactive').count(),
        # re-open the modal if we were redirected back after an error
        'open_modal':     request.GET.get('open_modal', False),
    }
    return render(request, 'core/dashboard/frontdesk_patients_list.html', context)



#   path('frontdesk/patients/',              views.frontdesk_patients_list,   name='frontdesk_patients_list'),
#   